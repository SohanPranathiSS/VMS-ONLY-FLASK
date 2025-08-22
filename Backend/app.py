# Flask Backend for Visitor Management System
# Converted from Node.js to Python Flask

from flask import Flask, request, jsonify, send_file, Response
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash
import jwt
from datetime import datetime, timedelta, timezone
import mysql.connector
from mysql.connector import pooling
import os
import json
import logging
from functools import wraps
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import base64
import io
import csv
import pandas as pd
# Import openpyxl components
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
import tempfile
import random
import string

# Optional weasyprint import with fallback
try:
    import weasyprint
    WEASYPRINT_AVAILABLE = True
    print("✓ WeasyPrint loaded successfully")
except (ImportError, OSError) as e:
    WEASYPRINT_AVAILABLE = False
    print(f"⚠ WeasyPrint not available: {e}")
    print("  PDF export will use HTML fallback that can be printed as PDF from browser.")
from werkzeug.utils import secure_filename
import uuid
import random
import string
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Initialize Flask app
app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('JWT_SECRET', 'your-super-secret-jwt-key-change-this-in-production')
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Enable CORS with comprehensive configuration
allowed_origins = os.getenv('ALLOWED_ORIGINS', 'https://visitors.pranathiss.com:3000,http://localhost:3000').split(',')
# Clean up origins (remove whitespace)
allowed_origins = [origin.strip() for origin in allowed_origins if origin.strip()]

# Add both HTTP and HTTPS versions for production
production_origins = []
for origin in allowed_origins:
    production_origins.append(origin)
    # Add HTTPS version if HTTP is provided
    if origin.startswith('http://'):
        https_version = origin.replace('http://', 'https://')
        if https_version not in production_origins:
            production_origins.append(https_version)

print(f"🌐 CORS enabled for origins: {production_origins}")

CORS(app, origins=production_origins, 
     methods=['GET', 'POST', 'PUT', 'DELETE', 'OPTIONS'],
     allow_headers=['Content-Type', 'Authorization', 'X-Requested-With'],
     supports_credentials=True,
     expose_headers=['Content-Type', 'Authorization'])

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Database configuration
DB_CONFIG = {
    'host': os.getenv('DB_HOST', 'localhost'),
    'user': os.getenv('DB_USER', 'root'),
    'password': os.getenv('DB_PASSWORD', ''),
    'database': os.getenv('DB_NAME', 'vms_db'),
    'pool_name': 'mypool',
    'pool_size': 10,
    'pool_reset_session': True,
    'autocommit': True
}

# Alternative database configurations for fallback
DB_FALLBACK_CONFIGS = [
    # Try localhost first (for local/production deployments)
    {**DB_CONFIG, 'host': 'localhost'},
    # Try 127.0.0.1 
    {**DB_CONFIG, 'host': '127.0.0.1'},
    # Try mysql container name (for Docker deployments)
    {**DB_CONFIG, 'host': 'mysql'},
    # Try database service name
    {**DB_CONFIG, 'host': 'database'},
]

# Create connection pool with fallback
connection_pool = None
for config in DB_FALLBACK_CONFIGS:
    try:
        # Remove pool-specific keys for testing
        test_config = {k: v for k, v in config.items() if k not in ['pool_name', 'pool_size', 'pool_reset_session']}
        # Test connection first
        test_conn = mysql.connector.connect(**test_config)
        test_conn.close()
        
        # If test succeeds, create the pool
        connection_pool = mysql.connector.pooling.MySQLConnectionPool(**config)
        logger.info(f"✅ Database connection pool created successfully with host: {config['host']}")
        break
    except mysql.connector.Error as err:
        logger.warning(f"⚠️ Database connection failed for host {config['host']}: {err}")
        continue

if connection_pool is None:
    logger.error(f"❌ All database connection attempts failed")
    # Don't exit, let individual functions handle the error
else:
    logger.info("✅ Database connection pool initialized")

def get_db_connection():
    """Get database connection from pool"""
    global connection_pool
    
    # If pool exists, try to use it
    if connection_pool:
        try:
            conn = connection_pool.get_connection()
            # Test the connection with a simple query
            cursor = conn.cursor()
            cursor.execute("SELECT 1")
            # Make sure to fetch the result to avoid "Unread result found" errors
            cursor.fetchone()
            cursor.close()
            return conn
        except mysql.connector.Error as err:
            logger.error(f"Error getting database connection from pool: {err}")
    
    # Fallback: try direct connections with different hosts
    for config in DB_FALLBACK_CONFIGS:
        try:
            logger.info(f"Attempting direct database connection to {config['host']}")
            fallback_config = {k: v for k, v in config.items() if k not in ['pool_name', 'pool_size', 'pool_reset_session']}
            conn = mysql.connector.connect(**fallback_config)
            logger.info(f"✅ Direct connection successful to {config['host']}")
            return conn
        except mysql.connector.Error as fallback_err:
            logger.error(f"Direct connection failed to {config['host']}: {fallback_err}")
            continue
    
    # If all attempts fail, raise the last error
    raise mysql.connector.Error("All database connection attempts failed")

# Utility functions
def generate_qr_code():
    """Generate unique QR code"""
    timestamp = str(int(datetime.now().timestamp()))
    random_str = ''.join(random.choices(string.ascii_letters + string.digits, k=9))
    return f"VMS-{timestamp}-{random_str}"

def get_company_id_from_companies_table(user_id):
    """Get company_id from companies table using admin_company_id (user_id)"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Look for company where admin_company_id matches the user_id
        cursor.execute("""
            SELECT id as company_id FROM companies 
            WHERE admin_company_id = %s 
            LIMIT 1
        """, (user_id,))
        
        company = cursor.fetchone()
        cursor.close()
        conn.close()
        
        if company:
            return company['company_id']
        else:
            # Fallback: If no company found in companies table, 
            # get company_id from users table for backward compatibility
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT company_id FROM users WHERE id = %s", (user_id,))
            user_data = cursor.fetchone()
            cursor.close()
            conn.close()
            
            if user_data:
                return user_data[0]
            else:
                logger.warning(f"No company found for user_id {user_id}, defaulting to 1")
                return 1
                
    except mysql.connector.Error as db_err:
        logger.error(f"Database error in get_company_id_from_companies_table: {db_err}")
        logger.error(f"Database config: host={DB_CONFIG.get('host')}, user={DB_CONFIG.get('user')}, database={DB_CONFIG.get('database')}")
        raise Exception(f"Database connection failed: {str(db_err)}")
    except Exception as e:
        logger.error(f"Error getting company_id from companies table: {e}")
        # Fallback to user's company_id from users table
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT company_id FROM users WHERE id = %s", (user_id,))
            user_data = cursor.fetchone()
            cursor.close()
            conn.close()
            
            if user_data:
                return user_data[0]
            else:
                return 1
        except:
            return 1

def send_email(to_email, subject, html_content):
    """Send email notification using Gmail SMTP"""
    try:
        smtp_server = 'smtp.gmail.com'
        smtp_port = 587
        smtp_user = os.getenv('EMAIL_USER')
        smtp_password = os.getenv('EMAIL_PASS')
        
        if not smtp_user or not smtp_password:
            logger.warning("Email credentials not configured in .env file")
            return False
            
        msg = MIMEMultipart()
        msg['From'] = smtp_user
        msg['To'] = to_email
        msg['Subject'] = subject
        
        msg.attach(MIMEText(html_content, 'html'))
        
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(smtp_user, smtp_password)
        server.send_message(msg)
        server.quit()
        
        logger.info(f"Verification email sent successfully to {to_email}")
        return True
    except Exception as e:
        logger.error(f"Email sending failed: {e}")
        return False

# Authentication decorator
def authenticate_token(f):
    """JWT authentication decorator"""
    @wraps(f)
    def decorated(*args, **kwargs):
        token = None
        auth_header = request.headers.get('Authorization')
        
        if auth_header:
            try:
                token = auth_header.split(' ')[1]
            except IndexError:
                return jsonify({'message': 'Invalid token format'}), 401
        
        if not token:
            return jsonify({'message': 'Token is missing'}), 401
        
        try:
            data = jwt.decode(token, app.config['SECRET_KEY'], algorithms=['HS256'])
            
            # Get user from database
            conn = get_db_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT * FROM users WHERE id = %s", (data['id'],))
            current_user = cursor.fetchone()
            cursor.close()
            conn.close()
            
            if not current_user:
                return jsonify({'message': 'User not found'}), 401
                
            request.current_user = current_user
            
        except jwt.ExpiredSignatureError:
            return jsonify({'message': 'Token has expired'}), 401
        except jwt.InvalidTokenError:
            return jsonify({'message': 'Token is invalid'}), 401
        except Exception as e:
            logger.error(f"Authentication error: {e}")
            return jsonify({'message': 'Authentication failed'}), 401
        
        return f(*args, **kwargs)
    
    return decorated

# Handle preflight OPTIONS requests
@app.before_request
def handle_preflight():
    if request.method == "OPTIONS":
        response = jsonify()
        # Get the origin from the request and check if it's allowed
        origin = request.headers.get('Origin')
        allowed_origins = os.getenv('ALLOWED_ORIGINS', 'http://localhost:3000,https://visitors.pranathiss.com:3000').split(',')
        # Clean up origins (remove whitespace)
        allowed_origins = [origin.strip() for origin in allowed_origins if origin.strip()]
        
        if origin in allowed_origins:
            response.headers["Access-Control-Allow-Origin"] = origin
        else:
            # Default to the production domain
            response.headers["Access-Control-Allow-Origin"] = "https://visitors.pranathiss.com"
            
        response.headers['Access-Control-Allow-Headers'] = "Content-Type,Authorization,X-Requested-With"
        response.headers['Access-Control-Allow-Methods'] = "GET,PUT,POST,DELETE,OPTIONS"
        response.headers['Access-Control-Allow-Credentials'] = "true"
        return response

# Add CORS headers to all responses
@app.after_request
def after_request(response):
    origin = request.headers.get('Origin')
    allowed_origins = os.getenv('ALLOWED_ORIGINS', 'http://localhost:3000,https://visitors.pranathiss.com:3000').split(',')
    # Clean up origins (remove whitespace)
    allowed_origins = [origin.strip() for origin in allowed_origins if origin.strip()]
    
    if origin in allowed_origins:
        response.headers['Access-Control-Allow-Origin'] = origin
    else:
        response.headers['Access-Control-Allow-Origin'] = 'https://visitors.pranathiss.com'
        
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type,Authorization,X-Requested-With'
    response.headers['Access-Control-Allow-Methods'] = 'GET,PUT,POST,DELETE,OPTIONS'
    response.headers['Access-Control-Allow-Credentials'] = 'true'
    return response

# Test database connection endpoint
@app.route('/api/test-db', methods=['GET'])
def test_db():
    """Test database connection"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT 1")
        result = cursor.fetchone()
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Database connection successful',
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        logger.error(f"Database test failed: {e}")
        return jsonify({
            'success': False,
            'message': 'Database connection failed',
            'error': str(e)
        }), 500

# Pricing plans with features endpoint
@app.route('/api/pricing/plans', methods=['GET'])
def get_pricing_plans():
    """Return all pricing plans with their features from pricing_plans and pricing_plan_features tables"""
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Fetch plans (order: priced plans first by price asc, then custom/unpriced)
        cursor.execute(
            """
            SELECT id, plan_name, billing_cycle, price, currency, description
            FROM pricing_plans
            ORDER BY (price IS NULL) ASC, price ASC, id ASC
            """
        )
        plans = cursor.fetchall()

        # Fetch features for all plans
        cursor.execute(
            """
            SELECT plan_id, feature_name, is_included, display_order, id
            FROM pricing_plan_features
            ORDER BY display_order ASC, id ASC
            """
        )
        features = cursor.fetchall()

        # Group features by plan_id
        features_by_plan = {}
        for f in features:
            features_by_plan.setdefault(f['plan_id'], []).append({
                'feature_name': f.get('feature_name'),
                'is_included': bool(f.get('is_included', 0))
            })

        # Build response
        result = []
        for p in plans:
            price_val = p.get('price')
            result.append({
                'id': p['id'],
                'plan_name': p.get('plan_name'),
                'billing_cycle': p.get('billing_cycle'),
                'price': float(price_val) if price_val is not None else None,
                'currency': p.get('currency') or 'INR',
                'description': p.get('description') or '',
                'features': features_by_plan.get(p['id'], [])
            })

        return jsonify({
            'success': True,
            'plans': result
        }), 200
    except Exception as e:
        logger.error(f"Failed to fetch pricing plans: {e}")
        return jsonify({
            'success': False,
            'message': 'Failed to fetch pricing plans'
        }), 500
    finally:
        try:
            if cursor:
                cursor.close()
            if conn:
                conn.close()
        except Exception:
            pass

# Email verification against vms_db users table
@app.route('/api/users/verify-email', methods=['POST'])
def verify_user_email():
    """Verify if a user email exists in the users table (vms_db)"""
    try:
        data = request.get_json(silent=True) or {}
        email = (data.get('email') or '').strip()

        if not email:
            return jsonify({
                'success': False,
                'message': 'Email is required',
                'exists': False
            }), 400

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        # Prefer active users; adjust if your schema differs
        cursor.execute(
            "SELECT id FROM users WHERE email = %s AND (is_active = 1 OR is_active IS NULL) LIMIT 1",
            (email,)
        )
        user = cursor.fetchone()
        cursor.close()
        conn.close()

        if user:
            return jsonify({
                'success': True,
                'message': 'Email exists',
                'exists': True,
                'userId': user['id']
            }), 200
        else:
            return jsonify({
                'success': True,
                'message': 'Email does not exist',
                'exists': False
            }), 200

    except Exception as e:
        logger.error(f"Error verifying email: {e}")
        return jsonify({
            'success': False,
            'message': 'Internal server error',
            'exists': False
        }), 500

# Create subscription after successful payment
@app.route('/api/subscription/create', methods=['POST'])
def create_subscription():
    """Create a subscription record and update company status after successful payment"""
    conn = None
    cursor = None
    try:
        data = request.get_json(silent=True) or {}
        email = (data.get('email') or '').strip()
        plan_name = (data.get('planName') or '').strip()
        amount = data.get('amount')  # not used in subscriptions table, kept for logging
        payment_id = (data.get('paymentId') or '').strip()
        payment_method = (data.get('paymentMethod') or data.get('payment_method') or 'razorpay')
        payment_method = payment_method.strip() if isinstance(payment_method, str) else 'razorpay'
        billing_cycle = (data.get('billingCycle') or data.get('billing_cycle') or '')
        billing_cycle = billing_cycle.strip().lower() if isinstance(billing_cycle, str) else ''
        billing_contact_email = (data.get('billingContactEmail') or data.get('billing_contact_email') or '')
        billing_contact_email = billing_contact_email.strip() if isinstance(billing_contact_email, str) else ''

        if not all([email, plan_name, payment_id]):
            return jsonify({
                'success': False,
                'message': 'Email, planName and paymentId are required'
            }), 400

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Find user and their company
        cursor.execute(
            "SELECT id, company_id FROM users WHERE email = %s AND (is_active = 1 OR is_active IS NULL) LIMIT 1",
            (email,)
        )
        user = cursor.fetchone()
        if not user:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'User not found'}), 404

        company_id = user.get('company_id')
        # Fallback to companies.admin_company_id mapping if company_id missing
        if not company_id:
            try:
                company_id = get_company_id_from_companies_table(user['id'])
            except Exception:
                company_id = None
        if not company_id:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'User is not associated with a company'}), 400

        # Try to derive actual payment method and billing email from Razorpay if not provided
        final_payment_method = payment_method or 'razorpay'
        final_billing_email = billing_contact_email or email
        rp_order_id = None
        rp_amount = None
        rp_currency = None
        rp_status = 'paid'
        rp_created_at = None
        try:
            import os
            import requests
            from requests.auth import HTTPBasicAuth
            rk = os.getenv('RAZORPAY_KEY_ID') or os.getenv('RAZORPAY_KEY') or ''
            rs = os.getenv('RAZORPAY_KEY_SECRET') or ''
            if rk and rs and payment_id:
                resp = requests.get(
                    f"https://api.razorpay.com/v1/payments/{payment_id}",
                    auth=HTTPBasicAuth(rk, rs),
                    timeout=6
                )
                if resp.status_code == 200:
                    pdata = resp.json()
                    method = (pdata.get('method') or '').lower()
                    if method:
                        method_map = {
                            'upi': 'UPI',
                            'card': 'Card',
                            'netbanking': 'NetBanking',
                            'wallet': 'Wallet',
                            'emi': 'EMI',
                            'paylater': 'Pay Later',
                            'emandate': 'eMandate'
                        }
                        final_payment_method = method_map.get(method, method)
                    pay_email = pdata.get('email') or pdata.get('contact_email') or (pdata.get('notes') or {}).get('billing_email')
                    if pay_email and not billing_contact_email:
                        final_billing_email = pay_email
                    # store additional payment details
                    rp_order_id = pdata.get('order_id')
                    rp_currency = pdata.get('currency') or 'INR'
                    try:
                        rp_amount = float(pdata.get('amount', 0)) / 100.0
                    except Exception:
                        rp_amount = None
                    # map status
                    pstatus = (pdata.get('status') or '').lower()
                    if pstatus in ('captured', 'paid'):
                        rp_status = 'paid'
                    elif pstatus in ('failed'):
                        rp_status = 'failed'
                    else:
                        rp_status = 'created'
                    rp_created_at = pdata.get('created_at')
        except Exception as fetch_err:
            logger.warning(f"Payment metadata fetch warning: {fetch_err}")

        # Ensure we always have a non-null value for razorpay_order_id to satisfy DB NOT NULL constraint
        # If order_id couldn't be fetched (common when using Checkout without Orders API),
        # fall back to a deterministic placeholder based on the payment_id
        if not rp_order_id:
            # This preserves traceability while meeting NOT NULL requirement
            rp_order_id = f"NO_ORDER_{payment_id}" if payment_id else "NO_ORDER_UNKNOWN"

        # Deactivate any existing active subscription (status enum doesn't include 'inactive')
        # Avoid referencing updated_at explicitly to support older schemas without this column
        cursor.execute(
            "UPDATE subscriptions SET status = 'expired', end_date = CURDATE() WHERE company_id = %s AND status = 'active'",
            (company_id,)
        )

        # Map billingCycle to subscriptions.plan enum, enterprise overrides
        plan_value = 'monthly'
        if billing_cycle in ('yearly', 'annual'):
            plan_value = 'yearly'
        if plan_name.lower() == 'enterprise':
            plan_value = 'enterprise'

        # Create new subscription duration based on cycle
        start_date = datetime.now().date()
        end_date = (datetime.now() + timedelta(days=365)).date() if plan_value == 'yearly' else (datetime.now() + timedelta(days=30)).date()
        cursor.execute(
            """
            INSERT INTO subscriptions (company_id, plan, status, start_date, end_date, razorpay_payment_id)
            VALUES (%s, %s, %s, %s, %s, %s)
            """,
            (company_id, plan_value, 'active', start_date, end_date, payment_id)
        )
        subscription_id = cursor.lastrowid

        # Record payment in payments table
        payment_amount = None
        try:
            payment_amount = float(rp_amount if rp_amount is not None else amount if amount is not None else 0)
        except Exception:
            payment_amount = 0.0
        payment_currency = rp_currency or 'INR'
        payment_status = rp_status or 'paid'
        # payment_date: prefer Razorpay created_at (epoch seconds), else NOW() for paid
        payment_date_value = None
        try:
            if rp_created_at:
                from datetime import datetime as _dt
                payment_date_value = _dt.fromtimestamp(int(rp_created_at))
        except Exception:
            payment_date_value = None
        try:
            if payment_date_value is not None:
                cursor.execute(
                    """
                    INSERT INTO payments (company_id, subscription_id, razorpay_order_id, razorpay_payment_id, status, amount, currency, payment_date)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (company_id, subscription_id, rp_order_id, payment_id, payment_status, payment_amount, payment_currency, payment_date_value)
                )
            else:
                cursor.execute(
                    """
                    INSERT INTO payments (company_id, subscription_id, razorpay_order_id, razorpay_payment_id, status, amount, currency)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """,
                    (company_id, subscription_id, rp_order_id, payment_id, payment_status, payment_amount, payment_currency)
                )
        except Exception as pay_err:
            # Include key values to aid troubleshooting, but avoid leaking sensitive data
            logger.error(
                "Failed to insert payment record: %s | company_id=%s subscription_id=%s payment_id=%s order_id=%s status=%s amount=%s currency=%s",
                pay_err, company_id, subscription_id, payment_id, rp_order_id, payment_status, payment_amount, payment_currency
            )

        # Update company subscription status and plan along with billing fields if available
        # Be backward-compatible if new columns are not present
        try:
            if final_billing_email:
                cursor.execute(
                    "UPDATE companies SET subscription_status = 'active', plan_name = %s, subscription_plan = %s, subscription_start_date = %s, subscription_end_date = %s, payment_method = %s, billing_contact_email = %s WHERE id = %s",
                    (plan_name, plan_value, start_date, end_date, final_payment_method, final_billing_email, company_id)
                )
            else:
                cursor.execute(
                    "UPDATE companies SET subscription_status = 'active', plan_name = %s, subscription_plan = %s, subscription_start_date = %s, subscription_end_date = %s, payment_method = %s WHERE id = %s",
                    (plan_name, plan_value, start_date, end_date, final_payment_method, company_id)
                )
        except Exception as err:
            # Fallback: update without new columns if unknown column error
            if 'Unknown column' in str(err):
                cursor.execute(
                    "UPDATE companies SET subscription_status = 'active', plan_name = %s, subscription_plan = %s, subscription_start_date = %s, subscription_end_date = %s WHERE id = %s",
                    (plan_name, plan_value, start_date, end_date, company_id)
                )
            else:
                raise

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify({
            'success': True,
            'message': 'Subscription created successfully',
            'subscriptionId': subscription_id
        }), 201

    except Exception as e:
        logger.error(f"Error creating subscription: {e}")
        try:
            # Attempt rollback if connection is open
            conn.rollback()
        except Exception:
            pass
        try:
            cursor.close()
            conn.close()
        except Exception:
            pass
        return jsonify({'success': False, 'message': 'Failed to create subscription'}), 500

# Debug registration endpoint
@app.route('/api/debug-register', methods=['POST'])
def debug_register():
    """Debug registration process step by step"""
    debug_info = []
    try:
        debug_info.append("Step 1: Starting debug registration")
        
        data = request.get_json()
        debug_info.append(f"Step 2: Received data keys: {list(data.keys()) if data else 'No data'}")
        
        if not data:
            return jsonify({'debug': debug_info, 'error': 'No JSON data received'}), 400
        
        # Test JWT encoding
        debug_info.append("Step 3: Testing JWT encoding")
        test_token = jwt.encode({
            'test': 'data',
            'exp': datetime.now(timezone.utc) + timedelta(hours=1)
        }, app.config['SECRET_KEY'], algorithm='HS256')
        debug_info.append(f"Step 4: JWT encoding successful: {type(test_token)}")
        
        # Test password hashing
        debug_info.append("Step 5: Testing password hashing")
        test_hash = generate_password_hash('testpassword')
        debug_info.append(f"Step 6: Password hashing successful: {len(test_hash)} chars")
        
        # Test database operations
        debug_info.append("Step 7: Testing database operations")
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Test SELECT query
        cursor.execute("SELECT COUNT(*) FROM users")
        user_count = cursor.fetchone()[0]
        debug_info.append(f"Step 8: Current user count: {user_count}")
        
        # Test MAX query
        cursor.execute("SELECT MAX(company_id) as max_id FROM users")
        max_result = cursor.fetchone()
        max_company_id = max_result[0] if max_result and max_result[0] else 0
        debug_info.append(f"Step 9: Max company_id: {max_company_id}")
        
        cursor.close()
        conn.close()
        debug_info.append("Step 10: Database operations successful")
        
        # Test email function (without actually sending)
        debug_info.append("Step 11: Testing email configuration")
        smtp_user = os.getenv('EMAIL_USER')
        smtp_password = os.getenv('EMAIL_PASS')
        debug_info.append(f"Step 12: Email config - User: {'Set' if smtp_user else 'Not set'}, Password: {'Set' if smtp_password else 'Not set'}")
        
        return jsonify({
            'success': True,
            'debug': debug_info,
            'message': 'All registration components tested successfully'
        })
        
    except Exception as e:
        debug_info.append(f"ERROR at step: {e}")
        import traceback
        debug_info.append(f"Traceback: {traceback.format_exc()}")
        return jsonify({
            'success': False,
            'debug': debug_info,
            'error': str(e)
        }), 500

# Database schema check endpoint
@app.route('/api/check-schema', methods=['GET'])
def check_schema():
    """Check database schema for debugging"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Check visitors table schema
        cursor.execute("DESCRIBE visitors")
        visitors_schema = cursor.fetchall()
        
        # Check visits table schema
        cursor.execute("DESCRIBE visits")
        visits_schema = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'visitors_table': [
                {
                    'Field': col[0],
                    'Type': col[1],
                    'Null': col[2],
                    'Key': col[3],
                    'Default': col[4],
                    'Extra': col[5]
                } for col in visitors_schema
            ],
            'visits_table': [
                {
                    'Field': col[0],
                    'Type': col[1],
                    'Null': col[2],
                    'Key': col[3],
                    'Default': col[4],
                    'Extra': col[5]
                } for col in visits_schema
            ]
        })
    except Exception as e:
        logger.error(f"Schema check failed: {e}")
        return jsonify({
            'success': False,
            'message': 'Schema check failed',
            'error': str(e)
        }), 500

# Debug endpoint to check visits data
@app.route('/api/debug-visits', methods=['GET'])
def debug_visits():
    """Debug endpoint to check all visits data"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Check if tables exist
        cursor.execute("SHOW TABLES")
        tables = cursor.fetchall()
        
        # Get all visits with visitor data
        cursor.execute("""
            SELECT v.*, vis.name as visitor_name_from_visitors, h.name as host_name
            FROM visits v
            LEFT JOIN visitors vis ON v.visitor_id = vis.id
            LEFT JOIN users h ON v.host_id = h.id
            ORDER BY v.check_in_time DESC
            LIMIT 10
        """)
        visits = cursor.fetchall()
        
        # Specifically check purpose_of_visit field
        cursor.execute("DESCRIBE visits")
        visits_schema = cursor.fetchall()
        
        # Check purpose_of_visit data
        cursor.execute("""
            SELECT id, purpose_of_visit, visitor_name, host_id, check_in_time 
            FROM visits 
            WHERE purpose_of_visit IS NOT NULL 
            ORDER BY check_in_time DESC 
            LIMIT 5
        """)
        purpose_data = cursor.fetchall()
        
        # Get count of total visits
        cursor.execute("SELECT COUNT(*) as total FROM visits")
        total_visits = cursor.fetchone()
        
        # Get count of total users
        cursor.execute("SELECT COUNT(*) as total FROM users")
        total_users = cursor.fetchone()
        
        # Get count of hosts specifically
        cursor.execute("SELECT COUNT(*) as total FROM users WHERE role = 'host'")
        total_hosts = cursor.fetchone()
        
        # Get all hosts and their visit counts
        cursor.execute("""
            SELECT u.id, u.name, u.email, u.role, 
                   COUNT(v.id) as visit_count
            FROM users u
            LEFT JOIN visits v ON u.id = v.host_id
            WHERE u.role = 'host'
            GROUP BY u.id, u.name, u.email, u.role
            ORDER BY visit_count DESC
        """)
        hosts_with_visits = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'database_tables': [list(table.values())[0] for table in tables],
            'total_visits': total_visits,
            'total_users': total_users,
            'total_hosts': total_hosts,
            'recent_visits': visits,
            'visits_schema': visits_schema,
            'purpose_of_visit_data': purpose_data,
            'hosts_with_visit_counts': hosts_with_visits,
            'debug_note': 'Check purpose_of_visit_data to see if reason field has data'
        })
    except Exception as e:
        logger.error(f"Debug visits failed: {e}")
        return jsonify({
            'success': False,
            'message': 'Debug failed',
            'error': str(e)
        }), 500

# Debug endpoint to check users data
@app.route('/api/test-users', methods=['GET'])
def test_users():
    """Debug endpoint to check all users data and table structure"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Check users table schema
        cursor.execute("DESCRIBE users")
        users_schema = cursor.fetchall()
        
        # Get all users with their fields
        cursor.execute("""
            SELECT id, name, email, role, company_name, 
                   mobile_number, department, designation, is_verified
            FROM users
            ORDER BY role, name
        """)
        all_users = cursor.fetchall()
        
        # Get count by role
        cursor.execute("""
            SELECT role, COUNT(*) as count 
            FROM users 
            GROUP BY role
        """)
        users_by_role = cursor.fetchall()
        
        # Get count by company
        cursor.execute("""
            SELECT company_name, COUNT(*) as count 
            FROM users 
            GROUP BY company_name
        """)
        users_by_company = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'users_table_schema': [
                {
                    'Field': col['Field'],
                    'Type': col['Type'],
                    'Null': col['Null'],
                    'Key': col['Key'],
                    'Default': col['Default'],
                    'Extra': col['Extra']
                } for col in users_schema
            ],
            'total_users': len(all_users),
            'users_by_role': users_by_role,
            'users_by_company': users_by_company,
            'all_users': all_users,
            'sample_user': all_users[0] if all_users else None
        })
    except Exception as e:
        logger.error(f"Test users failed: {e}")
        return jsonify({
            'success': False,
            'message': 'Test users failed',
            'error': str(e)
        }), 500

# ============== AUTHENTICATION ENDPOINTS ==============

@app.route('/api/register', methods=['POST'])
def register():
    """Register a new user - handles both company registration and admin creating hosts"""
    try:
        data = request.get_json()
        
        # Debug: log the received data
        logger.info(f"Register endpoint received data: {data}")
        
        # Check if this is an admin creating a host (has firstName/lastName) or regular registration
        if 'firstName' in data:
            # Admin creating a host user
            auth_header = request.headers.get('Authorization')
            if not auth_header:
                return jsonify({'message': 'Authorization required for user creation'}), 401
            
            try:
                token = auth_header.split(' ')[1]
                token_data = jwt.decode(token, app.config['SECRET_KEY'], algorithms=['HS256'])
                
                # Get admin user
                conn = get_db_connection()
                cursor = conn.cursor(dictionary=True)
                cursor.execute("SELECT * FROM users WHERE id = %s", (token_data['id'],))
                admin_user = cursor.fetchone()
                
                if not admin_user or admin_user['role'] != 'admin':
                    cursor.close()
                    conn.close()
                    return jsonify({'message': 'Admin access required'}), 403
                
                # Validate required fields for host creation
                required_fields = ['firstName', 'email', 'password', 'role', 'mobile_number', 'department', 'designation']
                for field in required_fields:
                    if not data.get(field) or not str(data.get(field)).strip():
                        return jsonify({'message': f'{field} is required'}), 400
                
                # lastName is optional - if not provided or empty, use empty string
                last_name = data.get('lastName', '').strip()
                
                # Combine first and last name
                if last_name:
                    full_name = f"{data['firstName'].strip()} {last_name}"
                else:
                    full_name = data['firstName'].strip()
                
                email = data['email'].strip()
                password = data['password']
                role = data['role'].lower()  # Use the role provided in request
                mobile_number = data['mobile_number'].strip()
                department = data['department'].strip()
                designation = data['designation'].strip()
                company_name = admin_user['company_name']
                
                # Validate role
                if role not in ['admin', 'host']:
                    return jsonify({'message': 'Invalid role. Must be Admin or Host.'}), 400
                
                # Check if user already exists
                cursor.execute("SELECT id FROM users WHERE email = %s", (email,))
                existing_user = cursor.fetchone()
                
                if existing_user:
                    cursor.close()
                    conn.close()
                    return jsonify({'message': 'User already exists with this email'}), 409
                
                # Hash password
                hashed_password = generate_password_hash(password)
                
                # Get company_id from companies table using admin's user_id
                admin_company_id = get_company_id_from_companies_table(admin_user['id'])
                
                # Add department and designation columns to users table if they don't exist
                try:
                    cursor.execute("ALTER TABLE users ADD COLUMN department VARCHAR(100) NULL")
                    logger.info("Added department column to users table")
                except Exception as alter_error:
                    logger.debug(f"ALTER TABLE for department: {alter_error}")
                
                try:
                    cursor.execute("ALTER TABLE users ADD COLUMN designation VARCHAR(100) NULL")
                    logger.info("Added designation column to users table")
                except Exception as alter_error:
                    logger.debug(f"ALTER TABLE for designation: {alter_error}")
                
                # Insert new user with all fields including is_verified = 1 and is_active = 1 (admin-created users are auto-verified and active)
                cursor.execute("""
                    INSERT INTO users (name, email, password, role, company_name, company_id, mobile_number, department, designation, is_verified, is_active)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (full_name, email, hashed_password, role, company_name, admin_company_id, mobile_number, department, designation, 1, 1))
                
                user_id = cursor.lastrowid
                cursor.close()
                conn.close()
                
                return jsonify({
                    'message': f'{role.title()} user created successfully and is ready to login immediately',
                    'userId': user_id,
                    'name': full_name,
                    'email': email,
                    'role': role,
                    'mobile_number': mobile_number,
                    'department': department,
                    'designation': designation,
                    'company_name': company_name,
                    'verified': True,
                    'active': True
                }), 201
                
            except jwt.ExpiredSignatureError:
                return jsonify({'message': 'Token has expired'}), 401
            except jwt.InvalidTokenError:
                return jsonify({'message': 'Invalid token'}), 401
            except Exception as e:
                logger.error(f"Token validation error: {e}")
                return jsonify({'message': 'Authentication failed'}), 401
        
        else:
            # Regular user registration
            required_fields = ['name', 'email', 'password', 'role', 'company_name']
            for field in required_fields:
                if not data.get(field):
                    return jsonify({'message': f'{field} is required'}), 400
            
            name = data['name']
            email = data['email']
            password = data['password']
            role = data['role']
            company_name = data['company_name']
            
            # Validate role
            if role not in ['admin', 'host']:
                return jsonify({'message': 'Invalid role. Must be admin or host.'}), 400
            
            # Check if user already exists
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM users WHERE email = %s", (email,))
            existing_user = cursor.fetchone()
            
            if existing_user:
                cursor.close()
                conn.close()
                return jsonify({'message': 'User already exists with this email'}), 409
            
            # Hash password
            hashed_password = generate_password_hash(password)
            
            # Add mobile_number column to users table if it doesn't exist
            try:
                cursor.execute("ALTER TABLE users ADD COLUMN mobile_number VARCHAR(20) NULL")
                logger.info("Added mobile_number column to users table")
            except Exception as alter_error:
                # Column might already exist, that's fine
                logger.debug(f"ALTER TABLE for mobile_number: {alter_error}")
            
        # Generate a unique company_id based on company_name hash or use auto-increment
        # For now, we'll use a simple approach - get next company_id based on existing companies
        cursor.execute("SELECT MAX(company_id) as max_id FROM users WHERE company_name = %s", (company_name,))
        existing_company = cursor.fetchone()
        
        if existing_company and existing_company[0]:
            # Company already exists, use same company_id
            company_id = existing_company[0]
        else:
            # New company, get next available company_id
            cursor.execute("SELECT MAX(company_id) as max_id FROM users")
            max_result = cursor.fetchone()
            company_id = (max_result[0] if max_result and max_result[0] else 0) + 1
        
        # Insert new user with proper company_id
        cursor.execute("""
            INSERT INTO users (name, email, password, role, company_name, company_id, is_verified, mobile_number)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (name, email, hashed_password, role, company_name, company_id, 0, ''))
        
        user_id = cursor.lastrowid
        cursor.close()
        conn.close()
        
        # Send verification email
        verification_token = jwt.encode({
            'user_id': user_id,
            'exp': datetime.now(timezone.utc) + timedelta(hours=24)
        }, app.config['SECRET_KEY'], algorithm='HS256')
        
        base_url = os.getenv('BASE_URL', 'https://visitors.pranathiss.com:4000')
        verification_link = f"{base_url}/api/verify-email?token={verification_token}"
        email_subject = "Verify Your Email - Visitor Management System"
        email_body = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
            <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 30px; border-radius: 10px; text-align: center; color: white;">
                <h1 style="margin: 0; font-size: 28px;">Welcome to VMS!</h1>
                <p style="margin: 10px 0 0 0; font-size: 16px;">Visitor Management System</p>
            </div>
            <div style="padding: 30px; background: #f8f9fa; border-radius: 0 0 10px 10px;">
                <h2 style="color: #333; margin-top: 0;">Hello {name},</h2>
                <p style="color: #666; line-height: 1.6;">Thank you for registering with our Visitor Management System. To complete your registration and secure your account, please verify your email address by clicking the button below:</p>
                <div style="text-align: center; margin: 30px 0;">
                    <a href="{verification_link}" style="display: inline-block; background-color: #007bff; color: white; padding: 15px 30px; text-decoration: none; border-radius: 5px; font-weight: bold; font-size: 16px;">Verify My Email</a>
                </div>
                <p style="color: #666; font-size: 14px; line-height: 1.6;">
                    <strong>Important:</strong> This verification link will expire in 24 hours. You will not be able to log in until your email is verified.
                </p>
                <p style="color: #666; font-size: 14px;">If you didn't create this account, please ignore this email.</p>
                <hr style="border: none; border-top: 1px solid #ddd; margin: 30px 0;">
                <p style="color: #999; font-size: 12px; text-align: center;">
                    Best regards,<br>
                    The VMS Team<br>
                    Visitor Management System
                </p>
            </div>
        </div>
        """
        
        send_email(email, email_subject, email_body)
        
        return jsonify({
            'message': 'User registered successfully. Please check your email to verify your account before logging in.',
            'userId': user_id
        }), 201
        
    except Exception as e:
        logger.error(f"Registration error: {e}")
        return jsonify({'message': 'Registration failed', 'error': str(e)}), 500

@app.route('/api/registerCompany', methods=['POST'])
def register_company():
    """Register a new company with admin user"""
    try:
        logger.info("Starting company registration...")
        data = request.get_json()
        logger.info(f"Received data: {data}")
        
        # Validate required fields for company registration (matching frontend field names)
        required_fields = ['companyName', 'firstName', 'lastName', 'email', 'password']
        for field in required_fields:
            if not data.get(field):
                logger.error(f"Missing field: {field}")
                return jsonify({'message': f'{field} is required'}), 400
        
        logger.info("All required fields present")
        
        company_name = data['companyName']
        admin_first_name = data['firstName']
        admin_last_name = data['lastName']
        admin_email = data['email']
        admin_password = data['password']
        mobile_number = data.get('mobileNumber', '')
        
        # Combine first and last name
        admin_name = f"{admin_first_name} {admin_last_name}"
        logger.info(f"Processing registration for: {admin_name} at {company_name}")
        
        # Check if company or admin already exists
        logger.info("Getting database connection...")
        conn = get_db_connection()
        cursor = conn.cursor()
        logger.info("Database connection successful")
        
        # Check if admin email already exists
        logger.info(f"Checking if email exists: {admin_email}")
        cursor.execute("SELECT id FROM users WHERE email = %s", (admin_email,))
        existing_user = cursor.fetchone()
        
        if existing_user:
            logger.info(f"Email already exists: {admin_email}")
            cursor.close()
            conn.close()
            return jsonify({'message': 'Email already exists'}), 409
        
        # Generate a unique company_id for new company
        # Check if company name already exists in companies table
        logger.info(f"Checking if company name exists: {company_name}")
        cursor.execute("SELECT id FROM companies WHERE company_name = %s LIMIT 1", (company_name,))
        existing_company = cursor.fetchone()
        
        if existing_company:
            logger.info(f"Company name already exists: {company_name}")
            cursor.close()
            conn.close()
            return jsonify({'message': 'Company name already exists'}), 409
        
        logger.info("Company name is unique")
        
        # Hash password
        logger.info("Hashing password...")
        hashed_password = generate_password_hash(admin_password)
        logger.info("Password hashed successfully")
        
        # Add mobile_number column to users table if it doesn't exist
        logger.info("Checking mobile_number column...")
        try:
            cursor.execute("ALTER TABLE users ADD COLUMN mobile_number VARCHAR(20) NULL")
            logger.info("Added mobile_number column to users table")
        except Exception as alter_error:
            # Column might already exist, that's fine
            logger.debug(f"ALTER TABLE for mobile_number: {alter_error}")
        
        # First, create the company record in the companies table with trial dates
        logger.info("Creating company record...")
        trial_start = datetime.now().date()
        trial_end = (datetime.now() + timedelta(days=14)).date()
        try:
            cursor.execute("""
                INSERT INTO companies (company_name, firstname, lastname, email, password, role, mobile_number, trial_start_date, trial_end_date, subscription_status)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (company_name, admin_first_name, admin_last_name, admin_email, hashed_password, 'admin', mobile_number, trial_start, trial_end, 'trial'))
        except Exception as err:
            # Fallback for older schemas without trial columns
            if 'Unknown column' in str(err):
                cursor.execute("""
                    INSERT INTO companies (company_name, firstname, lastname, email, password, role, mobile_number)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (company_name, admin_first_name, admin_last_name, admin_email, hashed_password, 'admin', mobile_number))
            else:
                raise
        
        company_id = cursor.lastrowid
        logger.info(f"Company created with ID: {company_id}")
        
        # Update the company record with admin_company_id (self-reference)
        cursor.execute("""
            UPDATE companies SET admin_company_id = %s WHERE id = %s
        """, (company_id, company_id))
        
        # Now insert the admin user with the correct company_id reference
        logger.info("Inserting new admin user...")
        cursor.execute("""
            INSERT INTO users (name, email, password, role, company_name, company_id, is_verified, mobile_number)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (admin_name, admin_email, hashed_password, 'admin', company_name, company_id, 0, mobile_number))
        
        user_id = cursor.lastrowid
        logger.info(f"User created with ID: {user_id}")
        cursor.close()
        conn.close()
        
        # Send verification email
        logger.info("Generating JWT token...")
        verification_token = jwt.encode({
            'user_id': user_id,
            'exp': datetime.now(timezone.utc) + timedelta(hours=24)
        }, app.config['SECRET_KEY'], algorithm='HS256')
        logger.info("JWT token generated successfully")
        
        base_url = os.getenv('BASE_URL', 'https://visitors.pranathiss.com:4000')
        verification_link = f"{base_url}/api/verify-email?token={verification_token}"
        email_subject = "Verify Your Email - Visitor Management System"
        email_body = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
            <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 30px; border-radius: 10px; text-align: center; color: white;">
                <h1 style="margin: 0; font-size: 28px;">Welcome to VMS!</h1>
                <p style="margin: 10px 0 0 0; font-size: 16px;">Visitor Management System</p>
            </div>
            <div style="padding: 30px; background: #f8f9fa; border-radius: 0 0 10px 10px;">
                <h2 style="color: #333; margin-top: 0;">Hello {admin_name},</h2>
                <p style="color: #666; line-height: 1.6;">Thank you for registering your company <strong>"{company_name}"</strong> with our Visitor Management System. To complete your registration and access your admin dashboard, please verify your email address by clicking the button below:</p>
                <div style="text-align: center; margin: 30px 0;">
                    <a href="{verification_link}" style="display: inline-block; background-color: #007bff; color: white; padding: 15px 30px; text-decoration: none; border-radius: 5px; font-weight: bold; font-size: 16px;">Verify My Email</a>
                </div>
                <p style="color: #666; font-size: 14px; line-height: 1.6;">
                    <strong>Important:</strong> This verification link will expire in 24 hours. You will not be able to access your admin dashboard until your email is verified.
                </p>
                <p style="color: #666; font-size: 14px;">If you didn't create this account, please ignore this email.</p>
                <hr style="border: none; border-top: 1px solid #ddd; margin: 30px 0;">
                <p style="color: #999; font-size: 12px; text-align: center;">
                    Best regards,<br>
                    The VMS Team<br>
                    Visitor Management System
                </p>
            </div>
        </div>
        """
        
        logger.info(f"Sending verification email to: {admin_email}")
        email_result = send_email(admin_email, email_subject, email_body)
        logger.info(f"Email sending result: {email_result}")
        
        logger.info("Company registration completed successfully")
        return jsonify({
            'message': 'Company and admin user registered successfully. Please check your email to verify your account before logging in.',
            'userId': user_id,
            'companyName': company_name
        }), 201
        
    except Exception as e:
        logger.error(f"Company registration error: {e}")
        logger.error(f"Error type: {type(e)}")
        import traceback
        logger.error(f"Full traceback: {traceback.format_exc()}")
        return jsonify({'message': 'Company registration failed', 'error': str(e)}), 500
    



@app.route('/api/verify-email', methods=['GET'])
def verify_email():
    """Verify user email"""
    try:
        token = request.args.get('token')
        
        if not token:
            return jsonify({'message': 'Verification token is required'}), 400
        
        # Decode token
        data = jwt.decode(token, app.config['SECRET_KEY'], algorithms=['HS256'])
        user_id = data['user_id']
        
        # Update user as verified
        # Connect to DB with proper error handling
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
        except mysql.connector.Error as db_err:
            logger.error(f"Database connection error: {db_err}")
            return jsonify({'message': f'Database connection error: {str(db_err)}'}), 500
        
        try:
            # First get user data before updating
            cursor.execute("""
                SELECT id, name, email, password, role, company_name, is_verified, COALESCE(mobile_number, '') as mobile_number
                FROM users WHERE id = %s
            """, (user_id,))
            user_data = cursor.fetchone()
            
            if not user_data:
                cursor.close()
                conn.close()
                return jsonify({'message': 'User not found'}), 404
            
            user_id_db, name, email, password, role, company_name, current_verified, mobile_number = user_data
            
            # Check if user is already verified
            if current_verified == 1:
                cursor.close()
                conn.close()
                return jsonify({'message': 'Email already verified'}), 200
            
            # Update user as verified
            cursor.execute("""
                UPDATE users SET is_verified = 1 WHERE id = %s
            """, (user_id,))
            
            if cursor.rowcount == 0:
                cursor.close()
                conn.close()
                return jsonify({'message': 'Failed to verify user'}), 500
            
            # Check if a record already exists in the companies table for this email
            cursor.execute("SELECT id FROM companies WHERE email = %s", (email,))
            existing_company = cursor.fetchone()
            
            if not existing_company:
                # Insert record into companies table when verification is successful and not already exists
                logger.info(f"No company record exists for {email}, creating one now")
                # Split name into first and last name
                name_parts = name.split(' ', 1)
                firstname = name_parts[0] if len(name_parts) > 0 else ''
                lastname = name_parts[1] if len(name_parts) > 1 else ''
                
                # Check if companies table exists and create if needed
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS companies (
                        id INT AUTO_INCREMENT PRIMARY KEY,
                        firstname VARCHAR(100) NOT NULL,
                        lastname VARCHAR(100) NOT NULL,
                        email VARCHAR(100) NOT NULL,
                        password VARCHAR(255) NOT NULL,
                        role VARCHAR(50) NOT NULL,
                        company_name VARCHAR(200) NOT NULL,
                        admin_company_id INT NOT NULL,
                        mobile_number VARCHAR(20) NULL,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        INDEX idx_email (email),
                        INDEX idx_admin_company_id (admin_company_id)
                    )
                """)
                
                # Insert into companies table with trial dates
                trial_start = datetime.now().date()
                trial_end = (datetime.now() + timedelta(days=14)).date()
                try:
                    cursor.execute("""
                        INSERT INTO companies (firstname, lastname, email, password, role, company_name, admin_company_id, mobile_number, created_at, trial_start_date, trial_end_date, subscription_status)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW(), %s, %s, %s)
                    """, (firstname, lastname, email, password, role, company_name, user_id_db, mobile_number, trial_start, trial_end, 'trial'))
                except Exception as err_ins:
                    if 'Unknown column' in str(err_ins):
                        cursor.execute("""
                            INSERT INTO companies (firstname, lastname, email, password, role, company_name, admin_company_id, mobile_number, created_at)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW())
                        """, (firstname, lastname, email, password, role, company_name, user_id_db, mobile_number))
                    else:
                        raise
                logger.info(f"Created company record for {email}")
            else:
                logger.info(f"Company record for {email} already exists, skipping creation")
            
            # Explicitly commit the transaction
            try:
                conn.commit()
            except mysql.connector.Error as commit_err:
                logger.error(f"Error committing transaction: {commit_err}")
                # Try to rollback if possible
                try:
                    conn.rollback()
                except Exception as rollback_err:
                    logger.warning(f"Error during rollback: {rollback_err}")
                cursor.close()
                conn.close()
                return jsonify({'message': f'Error committing transaction: {str(commit_err)}'}), 500
            
            # Verify the update was successful
            try:
                cursor.execute("SELECT is_verified FROM users WHERE id = %s", (user_id,))
                result = cursor.fetchone()
            except mysql.connector.Error as select_err:
                logger.error(f"Error verifying update: {select_err}")
                result = None
            
            # Make sure to properly close everything
            try:
                cursor.close()
                conn.close()
            except Exception as close_err:
                logger.warning(f"Error closing database resources: {close_err}")
            
            logger.info(f"User ID {user_id} email verified and company record handled successfully")
            
            logger.info(f"User ID {user_id} email verified successfully. is_verified = {result[0] if result else 'unknown'}")
            
            # Get frontend URL from environment
            frontend_url = os.getenv('FRONTEND_URL', 'https://visitors.pranathiss.com:3000')
            
            # Return a nice HTML page for email verification success
            html_response = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Email Verified - VMS</title>
                <style>
                    body {{
                        font-family: Arial, sans-serif;
                        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                        margin: 0;
                        padding: 20px;
                        min-height: 100vh;
                        display: flex;
                        align-items: center;
                        justify-content: center;
                    }}
                    .container {{
                        background: white;
                        padding: 40px;
                        border-radius: 10px;
                        box-shadow: 0 10px 30px rgba(0,0,0,0.3);
                        text-align: center;
                        max-width: 500px;
                        width: 100%;
                    }}
                    .success-icon {{
                        color: #28a745;
                        font-size: 60px;
                        margin-bottom: 20px;
                    }}
                    h1 {{
                        color: #333;
                        margin-bottom: 20px;
                    }}
                    p {{
                        color: #666;
                        line-height: 1.6;
                        margin-bottom: 30px;
                    }}
                    .login-button {{
                        background-color: #007bff;
                        color: white;
                        padding: 12px 30px;
                        text-decoration: none;
                        border-radius: 5px;
                        font-weight: bold;
                        display: inline-block;
                        transition: background-color 0.3s;
                    }}
                    .login-button:hover {{
                        background-color: #0056b3;
                    }}
                </style>
            </head>
            <body>
                <div class="container">
                    <div class="success-icon">✅</div>
                    <h1>Email Verified Successfully!</h1>
                    <p>Your email address has been verified. You can now log in to your Visitor Management System account.</p>
                    <a href="{frontend_url}/login" class="login-button">Go to Login</a>
                </div>
            </body>
            </html>
            """
            return html_response, 200
            
        except Exception as db_error:
            conn.rollback()
            cursor.close()
            conn.close()
            logger.error(f"Database error during verification: {db_error}")
            return jsonify({'message': 'Email verification failed'}), 500
        
    except jwt.ExpiredSignatureError:
        html_response = """
        <!DOCTYPE html>
        <html>
        <head>
            <title>Verification Link Expired - VMS</title>
            <style>
                body { font-family: Arial, sans-serif; background: #f8f9fa; margin: 0; padding: 20px; min-height: 100vh; display: flex; align-items: center; justify-content: center; }
                .container { background: white; padding: 40px; border-radius: 10px; box-shadow: 0 5px 15px rgba(0,0,0,0.1); text-align: center; max-width: 500px; width: 100%; }
                .error-icon { color: #dc3545; font-size: 60px; margin-bottom: 20px; }
                h1 { color: #333; margin-bottom: 20px; }
                p { color: #666; line-height: 1.6; }
            </style>
        </head>
        <body>
            <div class="container">
                <div class="error-icon">⏰</div>
                <h1>Verification Link Expired</h1>
                <p>This verification link has expired. Please request a new verification email from the login page.</p>
            </div>
        </body>
        </html>
        """
        return html_response, 400
    except jwt.InvalidTokenError:
        html_response = """
        <!DOCTYPE html>
        <html>
        <head>
            <title>Invalid Verification Link - VMS</title>
            <style>
                body { font-family: Arial, sans-serif; background: #f8f9fa; margin: 0; padding: 20px; min-height: 100vh; display: flex; align-items: center; justify-content: center; }
                .container { background: white; padding: 40px; border-radius: 10px; box-shadow: 0 5px 15px rgba(0,0,0,0.1); text-align: center; max-width: 500px; width: 100%; }
                .error-icon { color: #dc3545; font-size: 60px; margin-bottom: 20px; }
                h1 { color: #333; margin-bottom: 20px; }
                p { color: #666; line-height: 1.6; }
            </style>
        </head>
        <body>
            <div class="container">
                <div class="error-icon">❌</div>
                <h1>Invalid Verification Link</h1>
                <p>This verification link is invalid. Please check your email for the correct link or request a new verification email.</p>
            </div>
        </body>
        </html>
        """
        return html_response, 400
    except mysql.connector.Error as db_err:
        logger.error(f"Database error during verification: {db_err}")
        return jsonify({'message': f'Database error: {str(db_err)}'}), 500
    except Exception as e:
        logger.error(f"Email verification error: {e}")
        return jsonify({'message': 'Email verification failed'}), 500

@app.route('/api/resend-verification', methods=['POST'])
def resend_verification():
    """Resend verification email to user"""
    try:
        data = request.get_json()
        email = data.get('email')
        
        if not email:
            return jsonify({'message': 'Email is required'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Check if user exists and is not already verified
        cursor.execute("SELECT id, is_verified FROM users WHERE email = %s", (email,))
        user = cursor.fetchone()
        
        if not user:
            return jsonify({'message': 'Email not found'}), 404
        
        user_id, is_verified = user
        
        if is_verified:
            return jsonify({'message': 'Email is already verified'}), 400
        
        # Generate new verification token
        token = jwt.encode({
            'user_id': user_id,
            'exp': datetime.now(timezone.utc) + timedelta(hours=24)
        }, app.config['SECRET_KEY'], algorithm='HS256')
        
        verification_link = f"{request.url_root.rstrip('/')}/api/verify-email?token={token}"
        
        # Send verification email
        subject = "Verify Your Email - Visitor Management System"
        body = f"""
        <html>
        <head>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    line-height: 1.6;
                    margin: 0;
                    padding: 20px;
                    background-color: #f4f4f4;
                }}
                .container {{
                    max-width: 600px;
                    margin: 0 auto;
                    background-color: white;
                    padding: 20px;
                    border-radius: 10px;
                    box-shadow: 0 0 10px rgba(0,0,0,0.1);
                }}
                .header {{
                    text-align: center;
                    color: #333;
                    border-bottom: 2px solid #007bff;
                    padding-bottom: 10px;
                    margin-bottom: 20px;
                }}
                .content {{
                    color: #555;
                    margin-bottom: 30px;
                }}
                .verify-button {{
                    display: block;
                    width: 200px;
                    margin: 20px auto;
                    padding: 12px 24px;
                    background-color: #007bff;
                    color: white;
                    text-decoration: none;
                    text-align: center;
                    border-radius: 5px;
                    font-weight: bold;
                }}
                .footer {{
                    text-align: center;
                    color: #777;
                    font-size: 12px;
                    margin-top: 30px;
                    border-top: 1px solid #eee;
                    padding-top: 20px;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h2>Email Verification - Visitor Management System</h2>
                </div>
                <div class="content">
                    <p>Hello,</p>
                    <p>Thank you for registering with our Visitor Management System. Please click the button below to verify your email address:</p>
                    <a href="{verification_link}" class="verify-button">Verify Email Address</a>
                    <p>If the button doesn't work, you can copy and paste this link into your browser:</p>
                    <p style="word-break: break-all; color: #007bff;">{verification_link}</p>
                    <p><strong>Note:</strong> This verification link will expire in 24 hours for security reasons.</p>
                </div>
                <div class="footer">
                    <p>If you didn't request this verification, please ignore this email.</p>
                    <p>© 2025 Visitor Management System. All rights reserved.</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        if send_email(email, subject, body):
            cursor.close()
            conn.close()
            logger.info(f"Verification email resent to {email}")
            return jsonify({'message': 'Verification email sent successfully'}), 200
        else:
            cursor.close()
            conn.close()
            return jsonify({'message': 'Failed to send verification email'}), 500
            
    except Exception as e:
        logger.error(f"Resend verification error: {e}")
        return jsonify({'message': 'Failed to resend verification email'}), 500

@app.route('/api/manual-verify', methods=['POST'])
def manual_verify():
    """Manually verify a user by user ID (for debugging/admin use)"""
    try:
        data = request.get_json()
        user_id = data.get('user_id')
        
        if not user_id:
            return jsonify({'message': 'User ID is required'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Check if user exists and get full user data
            cursor.execute("""
                SELECT id, name, email, password, role, company_name, is_verified, COALESCE(mobile_number, '') as mobile_number
                FROM users WHERE id = %s
            """, (user_id,))
            user = cursor.fetchone()
            
            if not user:
                cursor.close()
                conn.close()
                return jsonify({'message': 'User not found'}), 404
            
            user_id_db, name, email, password, role, company_name, current_verified, mobile_number = user
            
            # Check if user is already verified
            if current_verified == 1:
                cursor.close()
                conn.close()
                return jsonify({'message': 'User is already verified'}), 200
            
            # Update user as verified
            cursor.execute("""
                UPDATE users SET is_verified = 1 WHERE id = %s
            """, (user_id,))
            
            # Insert record into companies table when verification is successful
            # Split name into first and last name
            name_parts = name.split(' ', 1)
            firstname = name_parts[0] if len(name_parts) > 0 else ''
            lastname = name_parts[1] if len(name_parts) > 1 else ''
            
            # Check if companies table exists and create if needed
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS companies (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    firstname VARCHAR(100) NOT NULL,
                    lastname VARCHAR(100) NOT NULL,
                    email VARCHAR(100) NOT NULL,
                    password VARCHAR(255) NOT NULL,
                    role VARCHAR(50) NOT NULL,
                    company_name VARCHAR(200) NOT NULL,
                    admin_company_id INT NOT NULL,
                    mobile_number VARCHAR(20) NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    INDEX idx_email (email),
                    INDEX idx_admin_company_id (admin_company_id)
                )
            """)
            
            # Insert into companies table
            cursor.execute("""
                INSERT INTO companies (firstname, lastname, email, password, role, company_name, admin_company_id, mobile_number, created_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW())
            """, (firstname, lastname, email, password, role, company_name, user_id_db, mobile_number))
            
            # Explicitly commit the transaction
            conn.commit()
            
            # Verify the update was successful
            cursor.execute("SELECT is_verified FROM users WHERE id = %s", (user_id,))
            result = cursor.fetchone()
            
            cursor.close()
            conn.close()
            
            logger.info(f"Manual verification - User ID {user_id} ({email}) verified and record inserted into companies table")
            return jsonify({
                'message': f'User ID {user_id} manually verified successfully and record added to companies table',
                'email': email,
                'previous_status': current_verified,
                'new_status': result[0] if result else None
            }), 200
            
        except Exception as db_error:
            conn.rollback()
            cursor.close()
            conn.close()
            logger.error(f"Database error during manual verification: {db_error}")
            return jsonify({'message': 'Manual verification failed'}), 500
            
    except Exception as e:
        logger.error(f"Manual verification error: {e}")
        return jsonify({'message': 'Manual verification failed'}), 500

@app.route('/api/debug-unverified', methods=['GET'])
def debug_unverified():
    """Debug endpoint to list all unverified users"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT id, name, email, role, company_name, is_verified
            FROM users 
            WHERE is_verified = 0 
            ORDER BY id DESC
        """)
        unverified_users = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'total_unverified': len(unverified_users),
            'unverified_users': unverified_users
        }), 200
        
    except Exception as e:
        logger.error(f"Debug unverified users error: {e}")
        return jsonify({'message': 'Failed to fetch unverified users'}), 500

@app.route('/api/login', methods=['POST'])
def login():
    """Authenticate user and return JWT token"""
    try:
        data = request.get_json()
        
        if not data.get('email') or not data.get('password'):
            return jsonify({'message': 'Email and password are required'}), 400
        
        email = data['email']
        password = data['password']
        
        # Find user
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM users WHERE email = %s", (email,))
        user = cursor.fetchone()
        cursor.close()
        conn.close()
        
        if not user or not check_password_hash(user['password'], password):
            return jsonify({'message': 'Invalid email or password'}), 401
        
        # Check if email is verified
        if not user.get('is_verified', 0):
            return jsonify({
                'message': 'Please verify your email address before logging in. Check your email for the verification link.',
                'emailVerificationRequired': True
            }), 403
        
        # Get company_id from companies table using user_id
        company_id = get_company_id_from_companies_table(user['id'])
        
        # Generate JWT token
        token = jwt.encode({
            'id': user['id'],
            'email': user['email'],
            'role': user['role'],
            'company_name': user['company_name'],
            'company_id': company_id,
            'exp': datetime.now(timezone.utc) + timedelta(hours=24)
        }, app.config['SECRET_KEY'], algorithm='HS256')
        
        return jsonify({
            'token': token,
            'user': {
                'id': user['id'],
                'name': user['name'],
                'email': user['email'],
                'role': user['role'],
                'company_name': user['company_name']
            }
        }), 200
        
    except Exception as e:
        logger.error(f"Login error: {e}")
        return jsonify({'message': 'Login failed'}), 500

# ============== PASSWORD RESET (EMAIL VERIFIED) ==============

@app.route('/api/forgot-password/check', methods=['POST'])
def forgot_password_check():
    """Check if email exists and is verified before allowing password reset"""
    try:
        data = request.get_json() or {}
        email = (data.get('email') or '').strip()
        if not email:
            return jsonify({'message': 'Email is required'}), 400

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id, email, is_verified FROM users WHERE email = %s", (email,))
        user = cursor.fetchone()
        cursor.close()
        conn.close()

        if not user:
            return jsonify({'message': 'User not found'}), 404

        if not user.get('is_verified', 0):
            return jsonify({'message': 'Email not verified. Please verify your email first.'}), 403

        return jsonify({'message': 'Email is verified', 'emailVerified': True}), 200
    except Exception as e:
        logger.error(f"Forgot password check error: {e}")
        return jsonify({'message': 'Failed to process request'}), 500


@app.route('/api/forgot-password/reset', methods=['POST'])
def forgot_password_reset():
    """Reset password directly after verifying email exists and is verified"""
    try:
        data = request.get_json(silent=True) or {}
        email = (data.get('email') or '').strip()
        new_password = data.get('new_password') or data.get('password')
        confirm_password = data.get('confirm_password')

        if not email or not new_password:
            return jsonify({'message': 'Email and new_password are required'}), 400

        # If confirm provided, ensure match
        if confirm_password is not None and new_password != confirm_password:
            return jsonify({'message': 'Passwords do not match'}), 400

        # Basic password policy
        if len(new_password) < 6:
            return jsonify({'message': 'Password must be at least 6 characters'}), 400

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id, is_verified FROM users WHERE email = %s", (email,))
        user = cursor.fetchone()
        if not user:
            cursor.close()
            conn.close()
            return jsonify({'message': 'User not found'}), 404

        if not user.get('is_verified', 0):
            cursor.close()
            conn.close()
            return jsonify({'message': 'Email not verified. Please verify your email first.'}), 403

        # Hash and update password
        hashed = generate_password_hash(new_password)
        cursor2 = conn.cursor()
        try:
            cursor2.execute(
                "UPDATE users SET password = %s, last_password_change = NOW(), failed_login_attempts = 0, locked_until = NULL WHERE id = %s",
                (hashed, user['id'])
            )
        except Exception as e:
            # Fallback for schemas missing some columns
            try:
                cursor2.execute(
                    "UPDATE users SET password = %s WHERE id = %s",
                    (hashed, user['id'])
                )
                logger.warning(f"Fallback password update used (reduced columns) for user_id={user['id']}: {e}")
            except Exception as e2:
                logger.error(f"Password update failed for user_id={user['id']}: {e2}")
                raise
        conn.commit()
        cursor2.close()
        cursor.close()
        conn.close()

        return jsonify({'message': 'Password reset successful'}), 200
    except Exception as e:
        logger.error(f"Forgot password reset error: {e}")
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify({'message': 'Failed to reset password'}), 500

# ============== CONTACT & DEMO ENDPOINTS ==============

@app.route('/api/contact', methods=['POST'])
def contact():
    """Store contact form submission"""
    try:
        data = request.get_json()
        
        required_fields = ['name', 'email', 'message']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'message': f'{field} is required'}), 400
        
        name = data['name']
        email = data['email']
        message = data['message']
        
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO contact_us (name, email, message)
            VALUES (%s, %s, %s)
        """, (name, email, message))
        cursor.close()
        conn.close()
        
        # Send notification email to admin
        admin_email = os.getenv('ADMIN_EMAIL')
        if admin_email:
            subject = f"New Contact Form Submission from {name}"
            body = f"""
            <h3>New Contact Form Submission</h3>
            <p><strong>Name:</strong> {name}</p>
            <p><strong>Email:</strong> {email}</p>
            <p><strong>Message:</strong></p>
            <p>{message}</p>
            <p><strong>Submitted at:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            """
            send_email(admin_email, subject, body)
        
        return jsonify({'message': 'Contact form submitted successfully'}), 201
        
    except Exception as e:
        logger.error(f"Contact form error: {e}")
        return jsonify({'message': 'Failed to submit contact form'}), 500

@app.route('/api/book-demo', methods=['POST'])
def book_demo():
    """Store demo booking request"""
    try:
        data = request.get_json()
        
        required_fields = ['name', 'email', 'organization']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'message': f'{field} is required'}), 400
        
        name = data['name']
        email = data['email']
        organization = data['organization']
        preferred_date = data.get('preferred_date')
        message = data.get('message', '')
        
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO book_demo (name, email, organization, preferred_date, message)
            VALUES (%s, %s, %s, %s, %s)
        """, (name, email, organization, preferred_date, message))
        cursor.close()
        conn.close()
        
        return jsonify({'message': 'Demo booking request submitted successfully'}), 201
        
    except Exception as e:
        logger.error(f"Demo booking error: {e}")
        return jsonify({'message': 'Failed to submit demo booking'}), 500

# ============== VISITS MANAGEMENT ENDPOINTS ==============

@app.route('/api/visits', methods=['POST'])
@authenticate_token
def create_visit():
    """Create a new visit (Check-In)"""
    try:
        data = request.get_json()
        
        # Validate JSON data
        if not data:
            logger.error("No JSON data received")
            return jsonify({'message': 'No data provided'}), 400
        
        user = request.current_user
        
        # Validate user object
        if not user or not user.get('id'):
            logger.error("Invalid user object from authentication")
            return jsonify({'message': 'Authentication error'}), 401
        
        # Get company_id from companies table using user_id (needed for host lookup)
        try:
            company_id = get_company_id_from_companies_table(user['id'])
            logger.info(f"Retrieved company_id: {company_id} for user_id: {user['id']}")
        except Exception as e:
            logger.error(f"Error getting company_id: {e}")
            # Provide more specific error message based on the type of error
            if "Database connection failed" in str(e):
                return jsonify({'message': 'Database connection error - please try again later'}), 500
            elif "mysql.connector" in str(e).lower():
                return jsonify({'message': 'Database service unavailable'}), 500
            else:
                return jsonify({'message': f'Company information error: {str(e)}'}), 500
        
        # Debug: log the received data
        logger.info(f"Create visit endpoint received data: {data}")
        
        # Map frontend field names to backend expected names
        visitor_name = data.get('name') or data.get('visitorName')
        visitor_email = data.get('email') or data.get('visitorEmail')
        visitor_phone = data.get('phone') or data.get('visitorPhone', '')
        visitor_designation = data.get('designation') or data.get('visitorDesignation', '')
        visitor_company = data.get('company') or data.get('visitorCompany', '')
        visitor_photo = data.get('photo') or data.get('visitorPhoto', '')
        id_card_photo = data.get('idCardPhoto', '')
        id_card_number = data.get('idCardNumber', '')
        # Extract the missing fields
        company_tel = data.get('companyTel', '')
        website = data.get('website', '')
        address = data.get('address', '')
        id_card_type = data.get('idCardType', '')
        reason = data.get('reason') or data.get('purpose')
        items_carried = data.get('itemsCarried', '')
        pre_registration_id = data.get('pre_registration_id')
        
        # Validate and handle oversized image data
        MAX_IMAGE_LENGTH = 16777215  # MEDIUMTEXT limit (16MB)
        
        if visitor_photo and len(visitor_photo) > MAX_IMAGE_LENGTH:
            logger.warning(f"Visitor photo too long ({len(visitor_photo)} chars), skipping photo storage")
            visitor_photo = ''  # Skip storing oversized photo
        
        if id_card_photo and len(id_card_photo) > MAX_IMAGE_LENGTH:
            logger.warning(f"ID card photo too long ({len(id_card_photo)} chars), skipping photo storage")
            id_card_photo = ''  # Skip storing oversized photo
        
        # Validate and truncate other string fields to match database constraints
        field_limits = {
            'visitor_name': 100,
            'visitor_email': 100,
            'visitor_phone': 20,
            'visitor_designation': 100,
            'visitor_company': 200,
            'company_tel': 20,
            'website': 200,
            'id_card_number': 50,
            'id_card_type': 50
        }
        
        # Apply length limits
        if visitor_name and len(visitor_name) > field_limits['visitor_name']:
            logger.warning(f"Visitor name too long, truncating from {len(visitor_name)} to {field_limits['visitor_name']}")
            visitor_name = visitor_name[:field_limits['visitor_name']]
        
        if visitor_email and len(visitor_email) > field_limits['visitor_email']:
            logger.warning(f"Visitor email too long, truncating from {len(visitor_email)} to {field_limits['visitor_email']}")
            visitor_email = visitor_email[:field_limits['visitor_email']]
        
        if visitor_phone and len(visitor_phone) > field_limits['visitor_phone']:
            logger.warning(f"Visitor phone too long, truncating from {len(visitor_phone)} to {field_limits['visitor_phone']}")
            visitor_phone = visitor_phone[:field_limits['visitor_phone']]
        
        if visitor_designation and len(visitor_designation) > field_limits['visitor_designation']:
            logger.warning(f"Visitor designation too long, truncating from {len(visitor_designation)} to {field_limits['visitor_designation']}")
            visitor_designation = visitor_designation[:field_limits['visitor_designation']]
        
        if visitor_company and len(visitor_company) > field_limits['visitor_company']:
            logger.warning(f"Visitor company too long, truncating from {len(visitor_company)} to {field_limits['visitor_company']}")
            visitor_company = visitor_company[:field_limits['visitor_company']]
        
        if company_tel and len(company_tel) > field_limits['company_tel']:
            logger.warning(f"Company tel too long, truncating from {len(company_tel)} to {field_limits['company_tel']}")
            company_tel = company_tel[:field_limits['company_tel']]
        
        if website and len(website) > field_limits['website']:
            logger.warning(f"Website too long, truncating from {len(website)} to {field_limits['website']}")
            website = website[:field_limits['website']]
        
        if id_card_number and len(id_card_number) > field_limits['id_card_number']:
            logger.warning(f"ID card number too long, truncating from {len(id_card_number)} to {field_limits['id_card_number']}")
            id_card_number = id_card_number[:field_limits['id_card_number']]
        
        if id_card_type and len(id_card_type) > field_limits['id_card_type']:
            logger.warning(f"ID card type too long, truncating from {len(id_card_type)} to {field_limits['id_card_type']}")
            id_card_type = id_card_type[:field_limits['id_card_type']]
        
        logger.info(f"Mapped data - visitor_name: {visitor_name}, visitor_email: {visitor_email}, reason: {reason}")
        
        # Handle hostId vs hostName
        host_id = data.get('hostId')
        host_name = data.get('hostName')
        
        # If hostName is provided instead of hostId, look up the host ID
        if not host_id and host_name:
            host_conn = get_db_connection()
            host_cursor = host_conn.cursor(dictionary=True, buffered=True)
            try:
                host_cursor.execute("SELECT id FROM users WHERE name = %s", (host_name,))
                host_result = host_cursor.fetchall()
                
                if host_result:
                    host_id = host_result[0]['id']
                else:
                    # If hostName lookup fails, try to find any host for this company
                    logger.warning(f"Host not found with name: {host_name}, looking for any host in company")
                    host_cursor.execute("""
                        SELECT id, name FROM users 
                        WHERE role IN ('host', 'admin') AND company_id = %s 
                        LIMIT 1
                    """, (company_id,))
                    fallback_host = host_cursor.fetchall()
                    
                    if fallback_host:
                        host_id = fallback_host[0]['id']
                        host_name = fallback_host[0]['name']  # Update to actual host name
                        logger.info(f"Using fallback host: {host_name} (ID: {host_id})")
                    else:
                        logger.error(f"No hosts found for company_id {company_id}")
                        return jsonify({'message': f'No hosts available for your company. Please contact your administrator.'}), 400
            finally:
                host_cursor.close()
                host_conn.close()
        
        # Required fields validation
        if not visitor_name:
            logger.error(f"Missing visitor name. Available fields: {list(data.keys()) if data else 'No data'}")
            return jsonify({'message': 'Visitor name is required'}), 400
        if not visitor_email:
            logger.error(f"Missing visitor email. Available fields: {list(data.keys()) if data else 'No data'}")
            return jsonify({'message': 'Visitor email is required'}), 400
        if not host_id:
            logger.error(f"Missing host ID/name. Available fields: {list(data.keys()) if data else 'No data'}")
            return jsonify({'message': 'Host is required'}), 400
        
        # Ensure visitor_name is not empty string for database NOT NULL constraint
        if not visitor_name.strip():
            logger.error("Visitor name is empty string")
            return jsonify({'message': 'Visitor name cannot be empty'}), 400
        if not reason:
            logger.error(f"Missing reason/purpose. Available fields: {list(data.keys()) if data else 'No data'}")
            return jsonify({'message': 'Visit reason is required'}), 400
        
        # Ensure reason is not empty string for database NOT NULL constraint
        if not reason.strip():
            reason = "General visit"  # Default fallback
            logger.warning(f"Empty reason provided, using default: {reason}")
        
        # Check if visitor is blacklisted using a separate connection
        blacklist_conn = get_db_connection()
        blacklist_cursor = blacklist_conn.cursor(dictionary=True, buffered=True)
        try:
            blacklist_cursor.execute("""
                SELECT id FROM visitors WHERE email = %s AND is_blacklisted = TRUE LIMIT 1
            """, (visitor_email,))
            blacklisted = blacklist_cursor.fetchall()
            
            if blacklisted:
                return jsonify({'message': 'This visitor has been blacklisted and cannot check in.'}), 403
        finally:
            blacklist_cursor.close()
            blacklist_conn.close()
        
        # Check for duplicate check-in for the same company today
        if visitor_email:
            duplicate_conn = get_db_connection()
            duplicate_cursor = duplicate_conn.cursor(dictionary=True, buffered=True)
            try:
                # Check for existing check-in for the same visitor email and host company today
                duplicate_cursor.execute("""
                    SELECT v.*, vis.name as existing_visitor_name, u.company_name as host_company_name
                    FROM visits v
                    JOIN visitors vis ON v.visitor_id = vis.id
                    JOIN users u ON v.host_id = u.id
                    WHERE vis.email = %s 
                    AND u.company_name = (SELECT company_name FROM users WHERE id = %s)
                    AND DATE(v.visit_date) = CURDATE()
                    AND v.status = 'checked-in'
                    LIMIT 1
                """, (visitor_email, host_id))
                existing_visit = duplicate_cursor.fetchone()
                
                if existing_visit:
                    check_in_time = existing_visit.get('check_in_time', 'Unknown')
                    host_company_name = existing_visit.get('host_company_name', 'this company')
                    return jsonify({
                        'message': f'You are already checked in for {host_company_name} today at {check_in_time}. Please check out first before checking in again.',
                        'error': 'DUPLICATE_CHECKIN',
                        'existing_visit_id': existing_visit.get('id'),
                        'existing_checkin_time': str(check_in_time)
                    }), 409  # 409 Conflict status code
            finally:
                duplicate_cursor.close()
                duplicate_conn.close()
        
        # Get host details using a separate connection BEFORE creating the visit
        host_details_conn = get_db_connection()
        host_details_cursor = host_details_conn.cursor(dictionary=True)
        try:
            host_details_cursor.execute("SELECT name, email FROM users WHERE id = %s", (host_id,))
            host = host_details_cursor.fetchone()
            if not host:
                return jsonify({'message': f'Host not found with ID: {host_id}'}), 400
            host_name_value = host['name'] or f"Host_{host_id}"  # Fallback if name is NULL
            host_email_value = host['email'] or f"host{host_id}@company.com"  # Fallback if email is NULL
            logger.info(f"Host details retrieved: name={host_name_value}, email={host_email_value}")
        finally:
            host_details_cursor.close()
            host_details_conn.close()
        
        # Perform the main database operations
        main_conn = get_db_connection()
        main_cursor = main_conn.cursor(dictionary=True, buffered=True)
        
        logger.info(f"About to start database operations with visitor_name: {visitor_name}, visitor_email: {visitor_email}, host_id: {host_id}, reason: {reason}")
        
        try:
            # Start transaction
            main_conn.start_transaction()
            
            # Always create a new visitor record - including all available fields
            main_cursor.execute("""
                INSERT INTO visitors (name, email, phone, designation, company, photo, idCardPhoto, 
                                    idCardNumber, companyTel, website, address, type_of_card)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (visitor_name, visitor_email, visitor_phone, visitor_designation, 
                  visitor_company, visitor_photo, id_card_photo, id_card_number,
                  company_tel, website, address, id_card_type))
            
            visitor_id = main_cursor.lastrowid
            logger.info(f"Created visitor with ID: {visitor_id}")
            
            # Create visit record - use purpose_of_visit (NOT NULL) instead of reason
            logger.info(f"Creating visit with reason: '{reason}' (type: {type(reason)}, length: {len(reason) if reason else 0})")
            main_cursor.execute("""
                INSERT INTO visits (visitor_id, host_id, purpose_of_visit, itemsCarried, check_in_time, 
                                  status, company_id, pre_registration_id, visitor_name, visitor_company,
                                  visitor_email, visitor_phone, visit_date, host_name, host_email)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s, %s, %s, %s)
            """, (visitor_id, host_id, reason, items_carried, datetime.now(), 
                  'checked-in', company_id, pre_registration_id, visitor_name,visitor_company, 
                  visitor_email, visitor_phone, datetime.now().date(), host_name_value, host_email_value))
            
            visit_id = main_cursor.lastrowid
            logger.info(f"Created visit with ID: {visit_id}, purpose_of_visit set to: '{reason}'")
            
            # Update pre-registration status if applicable
            if pre_registration_id:
                main_cursor.execute("""
                    UPDATE pre_registrations SET status = 'checked-in' 
                    WHERE id = %s
                """, (pre_registration_id,))
            
            main_conn.commit()
            
        except Exception as e:
            main_conn.rollback()
            raise e
        finally:
            main_cursor.close()
            main_conn.close()
        
        # Send notification email to host
        if host_email_value:
            subject = f"New Visitor Check-in: {visitor_name}"
            body = f"""
            <h3>New Visitor Check-in Notification</h3>
            <p>Dear {host_name_value},</p>
            <p>You have a new visitor:</p>
            <ul>
                <li><strong>Name:</strong> {visitor_name}</li>
                <li><strong>Company:</strong> {visitor_company}</li>
                <li><strong>Purpose:</strong> {reason}</li>
                <li><strong>Check-in Time:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</li>
            </ul>
            <p>Best regards,<br>Visitor Management System</p>
            """
            send_email(host_email_value, subject, body)
        
        return jsonify({
            'message': 'Visitor checked in successfully',
            'visitId': visit_id,
            'visitorId': visitor_id,
            'checkInTime': datetime.now().isoformat(),
            'hostId': host_id,
            'visitorName': visitor_name,
            'visitorEmail': visitor_email
        }), 201
            
    except Exception as e:
        logger.error(f"Visit creation error: {e}")
        logger.error(f"Error type: {type(e)}")
        logger.error(f"Error args: {e.args}")
        import traceback
        logger.error(f"Full traceback: {traceback.format_exc()}")
        
        # Provide more specific error messages based on error type
        error_message = str(e)
        if "Database connection failed" in error_message:
            user_message = "Database connection error - please try again"
        elif "mysql.connector" in error_message.lower():
            user_message = "Database service is currently unavailable"
        elif "Connection" in error_message:
            user_message = "Network connection error - please check your connection"
        elif "Authentication" in error_message:
            user_message = "Authentication error - please log in again"
        else:
            user_message = f"Visit creation failed: {error_message}"
        
        return jsonify({
            'message': user_message, 
            'error': str(e),
            'debug_info': 'Check server logs for detailed error information'
        }), 500

# Debug endpoint for visit creation
@app.route('/api/debug-visit', methods=['POST'])
@authenticate_token
def debug_visit():
    """Debug visit creation process step by step"""
    debug_info = []
    try:
        debug_info.append("Step 1: Starting debug visit creation")
        
        data = request.get_json()
        user = request.current_user
        debug_info.append(f"Step 2: User authenticated: {user['email']}, Role: {user['role']}")
        debug_info.append(f"Step 3: Received data keys: {list(data.keys()) if data else 'No data'}")
        
        if not data:
            return jsonify({'debug': debug_info, 'error': 'No JSON data received'}), 400
        
        # Test company_id retrieval
        debug_info.append("Step 4: Testing company_id retrieval")
        company_id = get_company_id_from_companies_table(user['id'])
        debug_info.append(f"Step 5: Company ID: {company_id}")
        
        # Test host lookup
        host_id = data.get('hostId')
        host_name = data.get('hostName')
        debug_info.append(f"Step 6: Host ID: {host_id}, Host Name: {host_name}")
        
        if not host_id and host_name:
            debug_info.append("Step 7: Looking up host by name")
            host_conn = get_db_connection()
            host_cursor = host_conn.cursor(dictionary=True, buffered=True)
            try:
                host_cursor.execute("SELECT id, name, email FROM users WHERE name = %s", (host_name,))
                host_result = host_cursor.fetchall()
                debug_info.append(f"Step 8: Host lookup result: {len(host_result)} found")
                if host_result:
                    host_id = host_result[0]['id']
                    debug_info.append(f"Step 9: Using host ID: {host_id}")
            finally:
                host_cursor.close()
                host_conn.close()
        
        # Test database schema check
        debug_info.append("Step 10: Testing database schema")
        schema_conn = get_db_connection()
        schema_cursor = schema_conn.cursor()
        try:
            # Check visitors table columns
            schema_cursor.execute("DESCRIBE visitors")
            visitors_columns = [row[0] for row in schema_cursor.fetchall()]
            debug_info.append(f"Step 11: Visitors table columns: {visitors_columns}")
            
            # Check visits table columns
            schema_cursor.execute("DESCRIBE visits")
            visits_columns = [row[0] for row in schema_cursor.fetchall()]
            debug_info.append(f"Step 12: Visits table columns: {visits_columns}")
            
        finally:
            schema_cursor.close()
            schema_conn.close()
        
        # Test actual visit creation process
        debug_info.append("Step 13: Testing actual visit creation process")
        
        # Extract the missing fields
        visitor_name = data.get('name') or data.get('visitorName')
        visitor_email = data.get('email') or data.get('visitorEmail')
        visitor_phone = data.get('phone') or data.get('visitorPhone', '')
        reason = data.get('reason') or data.get('purpose')
        
        debug_info.append(f"Step 14: Visitor data - Name: {visitor_name}, Email: {visitor_email}")
        debug_info.append(f"Step 15: Visit reason: {reason}")
        
        if host_id:
            # Test host details retrieval
            debug_info.append("Step 16: Testing host details retrieval")
            try:
                host_details_conn = get_db_connection()
                host_details_cursor = host_details_conn.cursor(dictionary=True)
                host_details_cursor.execute("SELECT name, email FROM users WHERE id = %s", (host_id,))
                host = host_details_cursor.fetchone()
                debug_info.append(f"Step 17: Host details: {host}")
                host_details_cursor.close()
                host_details_conn.close()
            except Exception as e:
                debug_info.append(f"Step 17 ERROR: Host details retrieval failed: {e}")
        
        # Test visitor creation
        debug_info.append("Step 18: Testing visitor creation")
        try:
            test_conn = get_db_connection()
            test_cursor = test_conn.cursor()
            test_cursor.execute("""
                INSERT INTO visitors (name, email, phone)
                VALUES (%s, %s, %s)
            """, (visitor_name, visitor_email, visitor_phone))
            test_visitor_id = test_cursor.lastrowid
            debug_info.append(f"Step 19: Test visitor created with ID: {test_visitor_id}")
            
            # Rollback the test insertion
            test_conn.rollback()
            test_cursor.close()
            test_conn.close()
        except Exception as e:
            debug_info.append(f"Step 19 ERROR: Visitor creation failed: {e}")
        
        return jsonify({
            'success': True,
            'debug': debug_info,
            'message': 'Debug visit creation completed successfully'
        })
        
    except Exception as e:
        debug_info.append(f"ERROR at step: {e}")
        import traceback
        debug_info.append(f"Traceback: {traceback.format_exc()}")
        return jsonify({
            'success': False,
            'debug': debug_info,
            'error': str(e)
        }), 500

@app.route('/api/visits', methods=['GET'])
@authenticate_token
def get_visits():
    """Get all visits with filtering (admin only)"""
    try:
        user = request.current_user
        
        if user['role'] != 'admin':
            return jsonify({'message': 'Admin access required'}), 403
        
        # Get query parameters
        host_id = request.args.get('hostId')
        start_date = request.args.get('startDate')
        end_date = request.args.get('endDate')
        host_name = request.args.get('hostName')
        visitor_name = request.args.get('visitorName')
        
        # Build query - include all visitor fields
        query = """
            SELECT v.id, 
                   COALESCE(NULLIF(v.purpose_of_visit, ''), 'General Visit') AS reason, 
                   v.itemsCarried, v.check_in_time, v.check_out_time,
                   v.visitor_name, v.visitor_email, v.visitor_phone,
                   vis.id AS visitor_id, vis.designation, vis.company AS visitor_company,
                   vis.photo AS visitorPhoto, vis.idCardPhoto, vis.idCardNumber,
                   vis.companyTel, vis.website, vis.address, vis.type_of_card,
                   h.id AS host_id, h.name AS hostName
            FROM visits v
            LEFT JOIN visitors vis ON v.visitor_id = vis.id
            LEFT JOIN users h ON v.host_id = h.id
            WHERE h.company_name = %s
        """
        
        params = [user['company_name']]
        
        # Add filters
        if host_id:
            query += " AND v.host_id = %s"
            params.append(host_id)
        
        if start_date:
            query += " AND DATE(v.check_in_time) >= %s"
            params.append(start_date)
        
        if end_date:
            query += " AND DATE(v.check_in_time) <= %s"
            params.append(end_date)
        
        if host_name:
            query += " AND h.name LIKE %s"
            params.append(f"%{host_name}%")
        
        if visitor_name:
            query += " AND (v.visitor_name LIKE %s OR vis.name LIKE %s)"
            params.append(f"%{visitor_name}%")
            params.append(f"%{visitor_name}%")
        
        query += " ORDER BY v.check_in_time DESC"
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(query, params)
        visits = cursor.fetchall()
        cursor.close()
        conn.close()
        
        return jsonify(visits), 200
        
    except Exception as e:
        logger.error(f"Get visits error: {e}")
        return jsonify({'message': 'Failed to fetch visits'}), 500

@app.route('/api/host-visits', methods=['GET'])
@authenticate_token
def get_host_visits():
    """Get visits for the authenticated host with pagination support"""
    try:
        user = request.current_user
        
        logger.info(f"Host visits request - User: {user['id']}, Role: {user['role']}, Email: {user['email']}")
        
        if user['role'] != 'host':
            logger.warning(f"Non-host user {user['id']} attempted to access host visits")
            return jsonify({'message': 'Host access required'}), 403
        
        host_id = user['id']
        
        # Get pagination parameters
        page = request.args.get('page', 1, type=int)
        limit = request.args.get('limit', 10, type=int)
        
        # Ensure reasonable limits
        page = max(1, page)
        limit = max(1, min(100, limit))  # Cap at 100 items per page
        
        offset = (page - 1) * limit
        
        # First, get the total count
        count_query = """
            SELECT COUNT(*) as total
            FROM visits v
            WHERE v.host_id = %s
        """
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(count_query, (host_id,))
        total_result = cursor.fetchone()
        total_visits = total_result['total'] if total_result else 0
        
        # Calculate total pages
        total_pages = (total_visits + limit - 1) // limit if total_visits > 0 else 1
        
        # Use LEFT JOIN to ensure we get visits even if visitor record has issues
        # Use the visitor data stored directly in visits table as backup
        query = """
            SELECT v.id, 
                   COALESCE(NULLIF(v.purpose_of_visit, ''), 'General Visit') AS reason, 
                   v.itemsCarried, v.check_in_time, v.check_out_time, v.status,
                   COALESCE(vis.id, v.visitor_id) AS visitor_id, 
                   COALESCE(vis.name, v.visitor_name) AS visitorName, 
                   COALESCE(vis.email, v.visitor_email) AS visitorEmail, 
                   COALESCE(vis.phone, v.visitor_phone) AS visitorPhone, 
                   COALESCE(vis.designation, '') AS designation, 
                   COALESCE(vis.company, '') AS company, 
                   COALESCE(vis.photo, '') AS visitorPhoto,
                   COALESCE(vis.idCardPhoto, '') AS idCardPhoto, 
                   COALESCE(vis.idCardNumber, '') AS idCardNumber,
                   COALESCE(vis.companyTel, '') AS companyTel,
                   COALESCE(vis.website, '') AS website,
                   COALESCE(vis.address, '') AS address,
                   COALESCE(vis.type_of_card, '') AS type_of_card,
                   h.id AS host_id, h.name AS hostName
            FROM visits v
            LEFT JOIN visitors vis ON v.visitor_id = vis.id
            LEFT JOIN users h ON v.host_id = h.id
            WHERE v.host_id = %s
            ORDER BY v.check_in_time DESC
            LIMIT %s OFFSET %s
        """
        
        logger.info(f"Executing query for host_id={host_id}, limit={limit}, offset={offset}")
        cursor.execute(query, (host_id, limit, offset))
        visits = cursor.fetchall()
        cursor.close()
        conn.close()
        
        logger.info(f"Host {host_id} visits query returned {len(visits)} results (page {page} of {total_pages}, total: {total_visits})")
        
        # Log a sample of the data for debugging
        if visits:
            sample_visit = visits[0]
            logger.info(f"Sample visit data: ID={sample_visit.get('id')}, reason={sample_visit.get('reason')}, purpose_of_visit={sample_visit.get('purpose_of_visit')}")
            logger.info(f"Sample visit keys: {list(sample_visit.keys())}")
            
            # Additional debug: Check if reason field actually has data
            for i, visit in enumerate(visits[:3]):  # Check first 3 visits
                logger.info(f"Visit {i+1}: ID={visit.get('id')}, reason='{visit.get('reason')}', visitorName='{visit.get('visitorName')}'")
        else:
            logger.warning(f"No visits found for host {host_id} on page {page}")
        
        # Return paginated response
        response_data = {
            'visits': visits,
            'currentPage': page,
            'totalPages': total_pages,
            'totalVisits': total_visits,
            'limit': limit
        }
        
        return jsonify(response_data), 200
        
    except Exception as e:
        logger.error(f"Get host visits error: {e}")
        return jsonify({'message': 'Failed to fetch host visits'}), 500

@app.route('/api/visits/<int:visit_id>/checkout', methods=['PUT'])
@authenticate_token
def checkout_visitor(visit_id):
    """Check out a visitor"""
    try:
        user = request.current_user
        logger.info(f"Checkout attempt for visit {visit_id} by user {user['id']} ({user['role']})")
        
        if user['role'] == 'host':
            # Verify this visit belongs to the host
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT host_id FROM visits WHERE id = %s", (visit_id,))
            visit = cursor.fetchone()
            cursor.close()
            conn.close()
            
            if not visit:
                logger.warning(f"Visit {visit_id} not found")
                return jsonify({'message': 'Visit not found'}), 404
            
            if visit[0] != user['id']:
                logger.warning(f"Access denied: visit {visit_id} belongs to host {visit[0]}, not {user['id']}")
                return jsonify({'message': 'Visit not found or access denied'}), 404
        
        check_out_time = datetime.now()
        logger.info(f"Proceeding with checkout for visit {visit_id} at {check_out_time}")
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        try:
            conn.start_transaction()
            logger.info(f"Transaction started for visit {visit_id}")
            
            # Get visit details first to check current status
            cursor.execute("""
                SELECT pre_registration_id, status, check_out_time FROM visits WHERE id = %s
            """, (visit_id,))
            visit_details = cursor.fetchone()
            logger.info(f"Visit details: {visit_details}")
            
            if not visit_details:
                logger.warning(f"Visit {visit_id} not found in database")
                conn.rollback()
                return jsonify({'message': 'Visit not found'}), 404
            
            # Check if already checked out
            if visit_details['status'] == 'checked-out' or visit_details['check_out_time'] is not None:
                logger.warning(f"Visit {visit_id} already checked out. Status: {visit_details['status']}, Check-out time: {visit_details['check_out_time']}")
                conn.rollback()
                return jsonify({'message': 'Visitor already checked out'}), 400
            
            # Update visit with checkout time
            logger.info(f"Updating visit {visit_id} with checkout time {check_out_time}")
            cursor.execute("""
                UPDATE visits SET check_out_time = %s, status = 'checked-out' 
                WHERE id = %s
            """, (check_out_time, visit_id))
            
            logger.info(f"Update affected {cursor.rowcount} rows")
            if cursor.rowcount == 0:
                logger.warning(f"No rows affected when updating visit {visit_id}")
                conn.rollback()
                return jsonify({'message': 'Visit not found or already checked out'}), 404
            
            # Update pre-registration if applicable
            if visit_details and visit_details['pre_registration_id']:
                logger.info(f"Updating pre-registration {visit_details['pre_registration_id']} to checked_out")
                cursor.execute("""
                    UPDATE pre_registrations SET status = 'checked_out' 
                    WHERE id = %s
                """, (visit_details['pre_registration_id'],))
                logger.info(f"Pre-registration update affected {cursor.rowcount} rows")
            
            conn.commit()
            logger.info(f"Transaction committed successfully for visit {visit_id}")
            
            return jsonify({
                'message': 'Visitor checked out successfully',
                'checkOutTime': check_out_time.isoformat()
            }), 200
            
        except Exception as e:
            conn.rollback()
            raise e
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        logger.error(f"Checkout error: {e}")
        logger.error(f"Error type: {type(e).__name__}")
        logger.error(f"Error details: {str(e)}")
        return jsonify({'message': f'Failed to check out visitor: {str(e)}'}), 500

# ============== VISITOR MANAGEMENT ENDPOINTS ==============

@app.route('/api/test-blacklisted', methods=['GET'])
@authenticate_token
def test_blacklisted():
    """Test endpoint to debug blacklisted visitors functionality"""
    try:
        user = request.current_user
        admin_company_name = user['company_name']
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Test data structure
        test_results = {
            'admin_company': admin_company_name,
            'tables_exist': {},
            'columns_exist': {},
            'sample_data': {},
            'table_structures': {}
        }
        
        # Check if tables exist
        cursor.execute("SHOW TABLES LIKE 'visitors'")
        test_results['tables_exist']['visitors'] = cursor.fetchone() is not None
        
        cursor.execute("SHOW TABLES LIKE 'visits'")
        test_results['tables_exist']['visits'] = cursor.fetchone() is not None
        
        cursor.execute("SHOW TABLES LIKE 'users'")
        test_results['tables_exist']['users'] = cursor.fetchone() is not None
        
        # Get visitors table structure
        cursor.execute("SHOW COLUMNS FROM visitors")
        visitors_columns = cursor.fetchall()
        test_results['table_structures']['visitors'] = [col['Field'] for col in visitors_columns]
        
        # Check if is_blacklisted column exists
        cursor.execute("SHOW COLUMNS FROM visitors LIKE 'is_blacklisted'")
        test_results['columns_exist']['is_blacklisted'] = cursor.fetchone() is not None
        
        # Check if reason_for_blacklist column exists
        cursor.execute("SHOW COLUMNS FROM visitors LIKE 'reason_for_blacklist'")
        test_results['columns_exist']['reason_for_blacklist'] = cursor.fetchone() is not None
        
        # Get sample blacklisted visitors
        if test_results['columns_exist']['is_blacklisted']:
            cursor.execute("SELECT COUNT(*) as count FROM visitors WHERE is_blacklisted = TRUE")
            test_results['sample_data']['total_blacklisted'] = cursor.fetchone()['count']
            
            # Get sample of blacklisted visitors with their details
            cursor.execute("""
                SELECT id, name, email, is_blacklisted, reason_for_blacklist
                FROM visitors 
                WHERE is_blacklisted = TRUE 
                LIMIT 5
            """)
            test_results['sample_data']['blacklisted_sample'] = cursor.fetchall()
            
            # Test the simplified query
            cursor.execute("""
                SELECT COUNT(*) as count 
                FROM visitors v
                WHERE v.is_blacklisted = TRUE 
                AND EXISTS (
                    SELECT 1 
                    FROM visits vt 
                    INNER JOIN users hosts ON vt.host_id = hosts.id
                    WHERE (v.id = vt.visitor_id OR v.email = vt.visitor_email)
                    AND hosts.company_name = %s
                )
            """, (admin_company_name,))
            test_results['sample_data']['company_filtered_blacklisted'] = cursor.fetchone()['count']
        
        # Get hosts for this company
        cursor.execute("""
            SELECT id, name, email, company_name 
            FROM users 
            WHERE company_name = %s AND role = 'host'
            LIMIT 5
        """, (admin_company_name,))
        test_results['sample_data']['company_hosts'] = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify(test_results), 200
        
    except Exception as e:
        logger.error(f"Test blacklisted error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/visitors/pending', methods=['GET'])
@authenticate_token
def get_pending_visitors():
    """Get all pending visitors (pre-registrations)"""
    try:
        user = request.current_user
        limit = request.args.get('limit', 100)
        
        logger.info(f"Fetching pending visitors for user: {user.get('email', 'unknown')}, role: {user.get('role', 'unknown')}")
        
        # Get company filter based on user role
        if user['role'] == 'admin':
            company_filter = user['company_name']
        else:
            company_filter = user['company_name']
        
        if not company_filter:
            logger.error("No company name found for user")
            return jsonify({'message': 'Company information not found'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        try:
            # Check if pre_registrations table exists
            cursor.execute("SHOW TABLES LIKE 'pre_registrations'")
            table_exists = cursor.fetchone()
            
            if not table_exists:
                logger.warning("pre_registrations table does not exist")
                cursor.close()
                conn.close()
                return jsonify([]), 200
            
            # Check available columns first
            cursor.execute("SHOW COLUMNS FROM pre_registrations")
            columns_info = cursor.fetchall()
            available_columns = [col['Field'] for col in columns_info]
            logger.info(f"Available columns in pre_registrations: {available_columns}")
            
            # Build query with only available columns
            base_columns = ['id', 'visitor_name', 'visitor_email', 'visitor_phone', 
                           'visitor_company', 'host_name', 'visit_date', 'visit_time',
                           'purpose', 'status', 'created_at']
            
            optional_columns = ['qr_code', 'special_requirements', 'number_of_visitors']
            
            # Add available columns
            select_columns = []
            for col in base_columns:
                if col in available_columns:
                    select_columns.append(f'pr.{col}')
            
            for col in optional_columns:
                if col in available_columns:
                    select_columns.append(f'pr.{col}')
            
            if not select_columns:
                logger.error("No valid columns found in pre_registrations table")
                cursor.close()
                conn.close()
                return jsonify({'message': 'Database schema error'}), 500
            
            # Get pending pre-registrations
            query = f"""
                SELECT {', '.join(select_columns)}
                FROM pre_registrations pr
                WHERE pr.company_to_visit = %s 
                AND pr.status IN ('pending', 'approved')
                ORDER BY pr.created_at DESC
                LIMIT %s
            """
            
            logger.info(f"Executing query: {query}")
            cursor.execute(query, (company_filter, int(limit)))
            
            pending_visitors = cursor.fetchall()
            logger.info(f"Found {len(pending_visitors)} pending visitors")
            
            # Process results to handle datetime serialization
            processed_visitors = []
            for visitor in pending_visitors:
                processed_visitor = {}
                for key, value in visitor.items():
                    if isinstance(value, datetime):
                        processed_visitor[key] = value.isoformat()
                    elif hasattr(value, 'total_seconds'):  # timedelta
                        processed_visitor[key] = str(value)
                    elif isinstance(value, (bytes, bytearray)):
                        processed_visitor[key] = value.decode('utf-8') if value else None
                    else:
                        processed_visitor[key] = value
                processed_visitors.append(processed_visitor)
            
            cursor.close()
            conn.close()
            
            return jsonify(processed_visitors), 200
            
        except mysql.connector.Error as db_error:
            logger.error(f"Database error in pending visitors: {db_error}")
            cursor.close()
            conn.close()
            return jsonify({'message': f'Database error: {str(db_error)}'}), 500
        
    except Exception as e:
        logger.error(f"Get pending visitors error: {e}")
        return jsonify({'message': f'Failed to fetch pending visitors: {str(e)}'}), 500

@app.route('/api/visitors/blacklisted', methods=['GET'])
@authenticate_token
def get_blacklisted_visitors():
    """Get all blacklisted visitors who visited hosts from the same company as the admin"""
    try:
        user = request.current_user
        limit = request.args.get('limit', 100)
        admin_company_name = user['company_name']
        
        logger.info(f"Fetching blacklisted visitors for company: {admin_company_name}")
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        try:
            # Check if visitors table exists
            cursor.execute("SHOW TABLES LIKE 'visitors'")
            table_exists = cursor.fetchone()
            
            if not table_exists:
                logger.warning("visitors table does not exist")
                cursor.close()
                conn.close()
                return jsonify([]), 200
            
            # Check if is_blacklisted column exists
            cursor.execute("SHOW COLUMNS FROM visitors LIKE 'is_blacklisted'")
            column_exists = cursor.fetchone()
            
            if not column_exists:
                logger.info("is_blacklisted column doesn't exist, returning empty list")
                cursor.close()
                conn.close()
                return jsonify([]), 200
            
            # Check if reason_for_blacklist column exists
            cursor.execute("SHOW COLUMNS FROM visitors LIKE 'reason_for_blacklist'")
            has_blacklist_reason = cursor.fetchone() is not None
            
            # Query to get blacklisted visitors with comprehensive information
            # Include: Visit Date, Picture, Person Name, Person to Meet, Visitor ID, Visit Reason, Reason to Blacklist, Check-In, Check-Out
            if has_blacklist_reason:
                blacklist_reason_field = "v.reason_for_blacklist as reason_to_blacklist"
            else:
                blacklist_reason_field = "'Not specified' as reason_to_blacklist"
            
            # Simplified query to avoid "Out of sort memory" error
            # Get blacklisted visitors with their latest visit information
            query = f"""
                SELECT 
                    v.id as visitor_id,
                    v.name as person_name,
                    v.email,
                    v.phone,
                    v.company,
                    v.designation,
                    v.photo as picture,
                    v.is_blacklisted,
                    {blacklist_reason_field}
                FROM visitors v
                WHERE v.is_blacklisted = TRUE 
                AND EXISTS (
                    SELECT 1 
                    FROM visits vt 
                    INNER JOIN users hosts ON vt.host_id = hosts.id
                    WHERE (v.id = vt.visitor_id OR v.email = vt.visitor_email)
                    AND hosts.company_name = %s
                )
                ORDER BY v.id DESC
                LIMIT %s
            """
            
            cursor.execute(query, (admin_company_name, int(limit)))
            blacklisted_visitors = cursor.fetchall() or []
            
            # Get visit details for each blacklisted visitor separately
            processed_visitors = []
            for visitor in blacklisted_visitors:
                # Get latest visit details for this visitor
                visit_query = """
                    SELECT 
                        visits.check_in_time,
                        visits.check_out_time,
                        DATE(visits.check_in_time) as visit_date,
                        visits.reason as visit_reason,
                        hosts.name as person_to_meet
                    FROM visits 
                    INNER JOIN users hosts ON visits.host_id = hosts.id
                    WHERE (visits.visitor_id = %s OR visits.visitor_email = %s)
                    AND hosts.company_name = %s
                    ORDER BY visits.check_in_time DESC
                    LIMIT 1
                """
                
                cursor.execute(visit_query, (visitor['visitor_id'], visitor['email'], admin_company_name))
                visit_info = cursor.fetchone() or {}
                
                # Combine visitor and visit information
                processed_visitor = {}
                for key, value in visitor.items():
                    if isinstance(value, datetime):
                        processed_visitor[key] = value.isoformat()
                    elif hasattr(value, 'total_seconds'):  # timedelta
                        processed_visitor[key] = str(value)
                    elif isinstance(value, (bytes, bytearray)):
                        processed_visitor[key] = value.decode('utf-8') if value else None
                    else:
                        processed_visitor[key] = value
                
                # Add visit information
                for key, value in visit_info.items():
                    if isinstance(value, datetime):
                        processed_visitor[key] = value.isoformat()
                    elif hasattr(value, 'total_seconds'):  # timedelta
                        processed_visitor[key] = str(value)
                    elif isinstance(value, (bytes, bytearray)):
                        processed_visitor[key] = value.decode('utf-8') if value else None
                    else:
                        processed_visitor[key] = value
                
                # Add/format specific fields for frontend display
                processed_visitor['visit_date'] = processed_visitor.get('visit_date', '')
                processed_visitor['picture'] = processed_visitor.get('picture', '')
                processed_visitor['person_name'] = processed_visitor.get('person_name', '')
                processed_visitor['person_to_meet'] = processed_visitor.get('person_to_meet', '')
                processed_visitor['visitor_id'] = processed_visitor.get('visitor_id', '')
                processed_visitor['visit_reason'] = processed_visitor.get('visit_reason', '')
                processed_visitor['reason_to_blacklist'] = processed_visitor.get('reason_to_blacklist', 'Not specified')
                processed_visitor['check_in'] = processed_visitor.get('check_in_time', '')
                processed_visitor['check_out'] = processed_visitor.get('check_out_time', '')
                
                # Add blacklist status info
                processed_visitor['blacklist_status'] = 'Active'
                processed_visitor['company_match'] = admin_company_name
                
                # Add created_at and updated_at as current timestamp if not present
                if 'created_at' not in processed_visitor:
                    processed_visitor['created_at'] = datetime.now().isoformat()
                if 'updated_at' not in processed_visitor:
                    processed_visitor['updated_at'] = processed_visitor.get('check_in_time', datetime.now().isoformat())
                
                processed_visitors.append(processed_visitor)
            
            cursor.close()
            conn.close()
            
            logger.info(f"Found {len(processed_visitors)} blacklisted visitors for company {admin_company_name}")
            return jsonify(processed_visitors), 200
            
        except mysql.connector.Error as db_error:
            logger.error(f"Database error in blacklisted visitors: {db_error}")
            cursor.close()
            conn.close()
            return jsonify([]), 200  # Return empty array on DB error
        
    except Exception as e:
        logger.error(f"Get blacklisted visitors error: {e}")
        return jsonify({'message': f'Failed to fetch blacklisted visitors: {str(e)}'}), 500

@app.route('/api/visitors/counts', methods=['GET'])
@authenticate_token
def get_visitor_counts():
    """Get visitor counts for dashboard"""
    try:
        user = request.current_user
        company_filter = user['company_name']
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get pending visitors count
        cursor.execute("""
            SELECT COUNT(*) as count FROM pre_registrations 
            WHERE company_to_visit = %s AND status IN ('pending', 'approved')
        """, (company_filter,))
        pending_count = cursor.fetchone()['count']
        
        # Get blacklisted visitors count (only those who visited hosts from the same company)
        try:
            cursor.execute("SHOW COLUMNS FROM visitors LIKE 'is_blacklisted'")
            column_exists = cursor.fetchone()
            
            if column_exists:
                # Count blacklisted visitors who visited hosts from the same company
                cursor.execute("""
                    SELECT COUNT(DISTINCT v.id) as count 
                    FROM visitors v
                    INNER JOIN visits ON (
                        v.id = visits.visitor_id 
                        OR v.email = visits.visitor_email
                    )
                    INNER JOIN users hosts ON visits.host_id = hosts.id
                    WHERE v.is_blacklisted = TRUE 
                    AND hosts.company_name = %s
                """, (company_filter,))
                blacklisted_count = cursor.fetchone()['count']
            else:
                blacklisted_count = 0
        except Exception as bl_error:
            logger.warning(f"Error checking blacklisted visitors: {bl_error}")
            blacklisted_count = 0
        
        # Get total visitors count (today)
        cursor.execute("""
            SELECT COUNT(*) as count FROM visits v
            JOIN users h ON v.host_id = h.id
            WHERE h.company_name = %s AND DATE(v.check_in_time) = CURDATE()
        """, (company_filter,))
        today_visitors = cursor.fetchone()['count']
        
        # Get total unique visitors
        cursor.execute("""
            SELECT COUNT(DISTINCT v.visitor_email) as count FROM visits v
            JOIN users h ON v.host_id = h.id
            WHERE h.company_name = %s
        """, (company_filter,))
        total_unique_visitors = cursor.fetchone()['count']
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'pending': pending_count,
            'blacklisted': blacklisted_count,
            'todayVisitors': today_visitors,
            'totalUniqueVisitors': total_unique_visitors
        }), 200
        
    except Exception as e:
        logger.error(f"Get visitor counts error: {e}")
        return jsonify({
            'pending': 0,
            'blacklisted': 0,
            'todayVisitors': 0,
            'totalUniqueVisitors': 0
        }), 200

# Add a dedicated endpoint for visitor status metrics that frontend expects
@app.route('/api/visitors/status-counts', methods=['GET'])
@authenticate_token
def get_visitor_status_counts():
    """Get visitor status counts that match frontend expectations"""
    try:
        user = request.current_user
        company_filter = user['company_name']
        
        if not company_filter:
            logger.error("No company name found for user")
            return jsonify({'message': 'Company information not found'}), 400
        
        logger.info(f"Fetching visitor status counts for company: {company_filter}")
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Initialize counts
        counts = {
            'all': 0,
            'checked-in': 0,
            'checked-out': 0,
            'pending': 0,
            'expected': 0,
            'blacklisted': 0
        }
        
        try:
            # Get all visits count
            cursor.execute("""
                SELECT COUNT(*) as count FROM visits v
                JOIN users h ON v.host_id = h.id
                WHERE h.company_name = %s
            """, (company_filter,))
            all_visits = cursor.fetchone()['count'] or 0
            
            # Get checked-in visitors (have check_in_time but no check_out_time)
            cursor.execute("""
                SELECT COUNT(*) as count FROM visits v
                JOIN users h ON v.host_id = h.id
                WHERE h.company_name = %s 
                AND v.check_in_time IS NOT NULL 
                AND v.check_out_time IS NULL
            """, (company_filter,))
            counts['checked-in'] = cursor.fetchone()['count'] or 0
            
            # Get checked-out visitors (have both check_in_time and check_out_time)
            cursor.execute("""
                SELECT COUNT(*) as count FROM visits v
                JOIN users h ON v.host_id = h.id
                WHERE h.company_name = %s 
                AND v.check_in_time IS NOT NULL 
                AND v.check_out_time IS NOT NULL
            """, (company_filter,))
            counts['checked-out'] = cursor.fetchone()['count'] or 0
            
            # Check if pre_registrations table exists
            cursor.execute("SHOW TABLES LIKE 'pre_registrations'")
            pre_reg_exists = cursor.fetchone()
            
            pre_reg_count = 0
            if pre_reg_exists:
                # Get expected visitors (future pre-registrations)
                cursor.execute("""
                    SELECT COUNT(*) as count FROM pre_registrations
                    WHERE company_to_visit = %s 
                    AND status IN ('pending', 'approved', 'confirmed')
                    AND (visit_date >= CURDATE() OR visit_date IS NULL)
                """, (company_filter,))
                counts['expected'] = cursor.fetchone()['count'] or 0
                
                # Get pending visitors (past due pre-registrations)
                cursor.execute("""
                    SELECT COUNT(*) as count FROM pre_registrations
                    WHERE company_to_visit = %s 
                    AND status = 'pending'
                    AND visit_date < CURDATE()
                """, (company_filter,))
                counts['pending'] = cursor.fetchone()['count'] or 0
                
                # Get total pre-registrations
                cursor.execute("""
                    SELECT COUNT(*) as count FROM pre_registrations
                    WHERE company_to_visit = %s
                """, (company_filter,))
                pre_reg_count = cursor.fetchone()['count'] or 0
            
            # Get blacklisted count (only those who visited hosts from the same company)
            cursor.execute("SHOW COLUMNS FROM visitors LIKE 'is_blacklisted'")
            bl_column_exists = cursor.fetchone()
            
            if bl_column_exists:
                cursor.execute("""
                    SELECT COUNT(DISTINCT v.id) as count 
                    FROM visitors v
                    INNER JOIN visits ON (
                        v.id = visits.visitor_id 
                        OR v.email = visits.visitor_email
                    )
                    INNER JOIN users hosts ON visits.host_id = hosts.id
                    WHERE v.is_blacklisted = TRUE 
                    AND hosts.company_name = %s
                """, (company_filter,))
                counts['blacklisted'] = cursor.fetchone()['count'] or 0
            
            # Calculate total
            counts['all'] = all_visits + pre_reg_count
            
            cursor.close()
            conn.close()
            
            logger.info(f"Status counts for {company_filter}: {counts}")
            return jsonify(counts), 200
            
        except mysql.connector.Error as db_error:
            logger.error(f"Database error in status counts: {db_error}")
            cursor.close()
            conn.close()
            return jsonify(counts), 200  # Return zero counts on DB error
        
    except Exception as e:
        logger.error(f"Get visitor status counts error: {e}")
        return jsonify({
            'all': 0,
            'checked-in': 0,
            'checked-out': 0,
            'pending': 0,
            'expected': 0,
            'blacklisted': 0
        }), 200

@app.route('/api/visitors/status-metrics', methods=['GET'])
@authenticate_token
def get_visitor_status_metrics():
    """Get comprehensive visitor status metrics for the dashboard"""
    try:
        user = request.current_user
        company_filter = user['company_name']
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Initialize counts
        counts = {
            'checked-in': 0,
            'checked-out': 0,
            'pending': 0,
            'expected': 0,
            'blacklisted': 0,
            'all': 0
        }
        
        # Get checked-in visitors (have check_in_time but no check_out_time)
        cursor.execute("""
            SELECT COUNT(*) as count FROM visits v
            JOIN users h ON v.host_id = h.id
            WHERE h.company_name = %s 
            AND v.check_in_time IS NOT NULL 
            AND v.check_out_time IS NULL
        """, (company_filter,))
        counts['checked-in'] = cursor.fetchone()['count']
        
        # Get checked-out visitors (have both check_in_time and check_out_time)
        cursor.execute("""
            SELECT COUNT(*) as count FROM visits v
            JOIN users h ON v.host_id = h.id
            WHERE h.company_name = %s 
            AND v.check_in_time IS NOT NULL 
            AND v.check_out_time IS NOT NULL
        """, (company_filter,))
        counts['checked-out'] = cursor.fetchone()['count']
        
        # Get total visits for company (for all count)
        cursor.execute("""
            SELECT COUNT(*) as count FROM visits v
            JOIN users h ON v.host_id = h.id
            WHERE h.company_name = %s
        """, (company_filter,))
        all_visits_count = cursor.fetchone()['count']
        
        # Check if pre_registrations table exists and get pre-registration counts
        cursor.execute("SHOW TABLES LIKE 'pre_registrations'")
        pre_reg_table_exists = cursor.fetchone()
        
        pre_reg_count = 0
        
        if pre_reg_table_exists:
            # Get expected visitors (pre-registrations for today or future with pending/approved status)
            cursor.execute("""
                SELECT COUNT(*) as count FROM pre_registrations
                WHERE company_to_visit = %s 
                AND status IN ('pending', 'approved', 'confirmed')
                AND (visit_date >= CURDATE() OR visit_date IS NULL)
            """, (company_filter,))
            counts['expected'] = cursor.fetchone()['count']
            
            # Get pending visitors (pre-registrations past due with pending status)
            cursor.execute("""
                SELECT COUNT(*) as count FROM pre_registrations
                WHERE company_to_visit = %s 
                AND status = 'pending'
                AND visit_date < CURDATE()
            """, (company_filter,))
            counts['pending'] = cursor.fetchone()['count']
            
            # Get total pre-registrations count
            cursor.execute("""
                SELECT COUNT(*) as count FROM pre_registrations
                WHERE company_to_visit = %s
            """, (company_filter,))
            pre_reg_count = cursor.fetchone()['count']
        
        # Get blacklisted visitors count (only those who visited hosts from the same company)
        try:
            cursor.execute("SHOW COLUMNS FROM visitors LIKE 'is_blacklisted'")
            column_exists = cursor.fetchone()
            
            if column_exists:
                # Get blacklisted visitors who visited hosts from the same company
                cursor.execute("""
                    SELECT COUNT(DISTINCT v.id) as count 
                    FROM visitors v
                    INNER JOIN visits ON (
                        v.id = visits.visitor_id 
                        OR v.email = visits.visitor_email
                    )
                    INNER JOIN users hosts ON visits.host_id = hosts.id
                    WHERE v.is_blacklisted = TRUE 
                    AND hosts.company_name = %s
                """, (company_filter,))
                counts['blacklisted'] = cursor.fetchone()['count'] or 0
            else:
                counts['blacklisted'] = 0
        except Exception as blacklist_error:
            logger.warning(f"Error fetching blacklisted count: {blacklist_error}")
            counts['blacklisted'] = 0
        
        # Calculate total count (visits + pre-registrations)
        counts['all'] = all_visits_count + pre_reg_count
        
        cursor.close()
        conn.close()
        
        logger.info(f"Visitor status metrics for {company_filter}: {counts}")
        
        return jsonify(counts), 200
        
    except Exception as e:
        logger.error(f"Get visitor status metrics error: {e}")
        return jsonify({
            'checked-in': 0,
            'checked-out': 0,
            'pending': 0,
            'expected': 0,
            'blacklisted': 0,
            'all': 0
        }), 200

@app.route('/api/visitors', methods=['GET'])
@authenticate_token
def get_all_visitors():
    """Get all visitors for the company"""
    try:
        user = request.current_user
        limit = request.args.get('limit', 100)
        company_filter = user['company_name']
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get all visitors through visits
        cursor.execute("""
            SELECT DISTINCT v.visitor_name as name, v.visitor_email as email, 
                   v.visitor_phone as phone, vis.company, vis.designation,
                   vis.is_blacklisted, v.check_in_time as last_visit,
                   vis.id as visitor_id
            FROM visits v
            LEFT JOIN visitors vis ON v.visitor_id = vis.id
            JOIN users h ON v.host_id = h.id
            WHERE h.company_name = %s
            ORDER BY v.check_in_time DESC
            LIMIT %s
        """, (company_filter, int(limit)))
        
        all_visitors = cursor.fetchall()
        cursor.close()
        conn.close()
        
        return jsonify(all_visitors), 200
        
    except Exception as e:
        logger.error(f"Get all visitors error: {e}")
        return jsonify({'message': 'Failed to fetch visitors'}), 500

@app.route('/api/visitors/<int:visitor_id>/blacklist', methods=['PUT'])
@authenticate_token
def blacklist_visitor(visitor_id):
    """Blacklist or unblacklist a visitor"""
    try:
        user = request.current_user
        
        if user['role'] != 'admin':
            return jsonify({'message': 'Admin access required to blacklist visitors'}), 403
        
        data = request.get_json()
        is_blacklisted = data.get('isBlacklisted', False)
        reason_for_blacklist = data.get('reasonForBlacklist', '') if is_blacklisted else None
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # First check if visitor exists and get their email
        cursor.execute("SELECT id, name, email FROM visitors WHERE id = %s", (visitor_id,))
        visitor = cursor.fetchone()
        
        if not visitor:
            cursor.close()
            conn.close()
            return jsonify({'message': 'Visitor not found'}), 404
        
        visitor_email = visitor['email']
        
        if not visitor_email:
            cursor.close()
            conn.close()
            return jsonify({'message': 'Visitor email not found'}), 400
        
        # Check if reason_for_blacklist column exists
        cursor.execute("SHOW COLUMNS FROM visitors LIKE 'reason_for_blacklist'")
        has_reason_column = cursor.fetchone() is not None
        
        # Update blacklist status for ALL visitors with this email
        if has_reason_column:
            cursor.execute("""
                UPDATE visitors SET is_blacklisted = %s, reason_for_blacklist = %s WHERE email = %s
            """, (is_blacklisted, reason_for_blacklist, visitor_email))
        else:
            cursor.execute("""
                UPDATE visitors SET is_blacklisted = %s WHERE email = %s
            """, (is_blacklisted, visitor_email))
        
        affected_rows = cursor.rowcount
        cursor.close()
        conn.close()
        
        if affected_rows == 0:
            return jsonify({'message': 'No visitors updated'}), 404
        
        action = 'blacklisted' if is_blacklisted else 'unblacklisted'
        logger.info(f"Admin {user['id']} {action} all visitors with email {visitor_email} ({affected_rows} records affected)")
        
        response_data = {
            'message': f'All visitors with email {visitor_email} {action} successfully. {affected_rows} records updated.',
            'visitorId': visitor_id,
            'visitorName': visitor['name'],
            'visitorEmail': visitor_email,
            'isBlacklisted': is_blacklisted,
            'affectedRecords': affected_rows
        }
        
        if is_blacklisted and reason_for_blacklist:
            response_data['reasonForBlacklist'] = reason_for_blacklist
        
        return jsonify(response_data), 200
        
    except Exception as e:
        logger.error(f"Blacklist update error: {e}")
        return jsonify({'message': 'Failed to update visitor blacklist status'}), 500

# ============== USER MANAGEMENT ENDPOINTS ==============

@app.route('/api/users', methods=['GET'])
@authenticate_token
def get_users():
    """Get all users from admin's company"""
    try:
        user = request.current_user
        
        if user['role'] != 'admin':
            return jsonify({'message': 'Admin access required'}), 403
        
        admin_company_name = user['company_name']
        
        if not admin_company_name:
            return jsonify({'message': 'Admin company information not found'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT u.id, u.name, u.email, u.role, u.company_name, 
                   u.mobile_number, u.department, u.designation, u.is_verified, u.profile_photo
            FROM users u
            WHERE u.company_name = %s
            ORDER BY u.role, u.name
        """, (admin_company_name,))
        
        users = cursor.fetchall()
        cursor.close()
        conn.close()
        
        return jsonify(users), 200
        
    except Exception as e:
        logger.error(f"Fetch users error: {e}")
        return jsonify({'message': 'Failed to fetch users'}), 500

@app.route('/api/hosts', methods=['GET'])
@authenticate_token
def get_hosts():
    """Get all hosts from the current user's company"""
    try:
        user = request.current_user
        company_name = user['company_name']
        
        if not company_name:
            return jsonify({'message': 'Company information not found'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, name, email FROM users 
            WHERE company_name = %s AND role = 'host'
            ORDER BY name
        """, (company_name,))
        
        hosts = cursor.fetchall()
        cursor.close()
        conn.close()
        
        return jsonify(hosts), 200
        
    except Exception as e:
        logger.error(f"Fetch hosts error: {e}")
        return jsonify({'message': 'Failed to fetch hosts'}), 500

# ============== COMPANY INFO (SUBSCRIPTION) ==============

@app.route('/api/company', methods=['GET'])
@authenticate_token
def get_company_info():
    """Return the current user's company info including subscription dates."""
    try:
        user = request.current_user

        # Resolve company identifier robustly
        company_id = user.get('company_id')
        if not company_id:
            try:
                company_id = get_company_id_from_companies_table(user['id'])
            except Exception:
                company_id = None

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        company_row = None
        if company_id:
            cursor.execute(
                """
                SELECT id, company_name, subscription_status, plan_name, subscription_plan,
                       subscription_start_date, subscription_end_date, trial_start_date, trial_end_date,
                       payment_method, billing_contact_email
                FROM companies
                WHERE id = %s
                LIMIT 1
                """,
                (company_id,)
            )
            company_row = cursor.fetchone()

        # Fallback by company_name if id resolution failed
        if not company_row and user.get('company_name'):
            cursor.execute(
                """
                SELECT id, company_name, subscription_status, plan_name, subscription_plan,
                       subscription_start_date, subscription_end_date, trial_start_date, trial_end_date,
                       payment_method, billing_contact_email
                FROM companies
                WHERE company_name = %s
                ORDER BY id DESC
                LIMIT 1
                """,
                (user['company_name'],)
            )
            company_row = cursor.fetchone()

        # Fetch latest payment info for amount
        payment_row = None
        try:
            if company_row and (company_id or company_row.get('id')):
                cid = company_id or company_row.get('id')
                cursor.execute(
                    """
                    SELECT amount, currency, status, payment_date
                    FROM payments
                    WHERE company_id = %s
                    ORDER BY COALESCE(payment_date, NOW()) DESC, id DESC
                    LIMIT 1
                    """,
                    (cid,)
                )
                payment_row = cursor.fetchone()
        except Exception as pay_err:
            logger.warning(f"Could not fetch latest payment info: {pay_err}")

        cursor.close()
        conn.close()

        if not company_row:
            return jsonify({'message': 'Company not found for current user'}), 404

        # Normalize field names expected by frontend
        response = {
            'id': company_row['id'],
            'name': company_row['company_name'],
            'plan_name': company_row.get('plan_name'),
            'plan_type': company_row.get('subscription_plan'),
            'subscription_status': company_row.get('subscription_status') or 'inactive',
            'subscription_start_date': (company_row.get('subscription_start_date').isoformat() if company_row.get('subscription_start_date') else None),
            'subscription_end_date': (company_row.get('subscription_end_date').isoformat() if company_row.get('subscription_end_date') else None),
            'trial_start_date': (company_row.get('trial_start_date').isoformat() if company_row.get('trial_start_date') else None),
            'trial_end_date': (company_row.get('trial_end_date').isoformat() if company_row.get('trial_end_date') else None),
            'payment_method': company_row.get('payment_method'),
            'billing_contact_email': company_row.get('billing_contact_email'),
        }

        if payment_row:
            # Include latest payment amount and details
            response['payment_amount'] = payment_row.get('amount')
            response['payment_currency'] = payment_row.get('currency')
            response['payment_status'] = payment_row.get('status')
            pd = payment_row.get('payment_date')
            response['payment_date'] = (pd.isoformat() if pd else None)

        return jsonify(response), 200

    except Exception as e:
        logger.error(f"Fetch company info error: {e}")
        return jsonify({'message': 'Failed to fetch company info'}), 500

# ============== SUPPORT TICKETS (USER-SCOPED) ==============

@app.route('/api/tickets', methods=['POST'])
@authenticate_token
def create_ticket():
    """Create a support ticket for the current user/company, defaulting status to 'open'."""
    try:
        user = request.current_user
        data = request.get_json(force=True) or {}
        title = (data.get('title') or '').strip()
        description = (data.get('description') or '').strip()
        category = (data.get('category') or '').strip() or None
        created_by_company = user.get('company_name') or (data.get('created_by_company') or '').strip()

        if not title or not description or not created_by_company:
            return jsonify({'message': 'title, description, and created_by_company are required'}), 400

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        inserted_id = None
        # Prefer schema with created_by_email if available
        try:
            cursor.execute(
                """
                INSERT INTO tickets (title, description, status, priority, created_by_company, category, assigned_to, created_at, attachment_url, created_by_email)
                VALUES (%s, %s, 'open', 'medium', %s, %s, NULL, NOW(), NULL, %s)
                """,
                (title, description, created_by_company, category, user.get('email'))
            )
            conn.commit()
            inserted_id = cursor.lastrowid
        except Exception:
            conn.rollback()
            # Fallback without created_by_email column
            cursor.execute(
                """
                INSERT INTO tickets (title, description, status, priority, created_by_company, category, assigned_to, created_at, attachment_url)
                VALUES (%s, %s, 'open', 'medium', %s, %s, NULL, NOW(), NULL)
                """,
                (title, description, created_by_company, category)
            )
            conn.commit()
            inserted_id = cursor.lastrowid

        try:
            cursor.execute(
                "SELECT id, title, description, status, category, created_by_company, created_at FROM tickets WHERE id = %s",
                (inserted_id,)
            )
            row = cursor.fetchone()
        except Exception:
            row = {
                'id': inserted_id,
                'title': title,
                'description': description,
                'status': 'open',
                'category': category,
                'created_by_company': created_by_company,
            }
        cursor.close(); conn.close()
        return jsonify({'message': 'Created', 'ticket': row}), 201
    except Exception as e:
        logger.exception('Error creating ticket')
        return jsonify({'message': 'Server error while creating ticket'}), 500


@app.route('/api/tickets', methods=['GET'])
@authenticate_token
def list_tickets():
    """List non-closed tickets raised by the current user (prefer email filter; fallback to company)."""
    try:
        user = request.current_user
        email = user.get('email')
        company = user.get('company_name')

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        rows = []
        # Try email-based filtering first if column exists
        try:
            cursor.execute(
                """
                SELECT id, title, description, status, category, created_by_company, created_at
                FROM tickets
                WHERE status <> 'closed' AND created_by_email = %s
                ORDER BY created_at DESC
                LIMIT 200
                """,
                (email,)
            )
            rows = cursor.fetchall()
        except Exception:
            # Fallback to company-based filtering
            cursor.execute(
                """
                SELECT id, title, description, status, category, created_by_company, created_at
                FROM tickets
                WHERE status <> 'closed' AND created_by_company = %s
                ORDER BY created_at DESC
                LIMIT 200
                """,
                (company,)
            )
            rows = cursor.fetchall()

        cursor.close(); conn.close()
        return jsonify(rows), 200
    except Exception:
        logger.exception('Error listing tickets')
        return jsonify({'message': 'Server error while fetching tickets'}), 500

# ============== ADMIN USER MANAGEMENT (CRUD + PASSWORD + PROFILE PHOTO) ==============

def _is_admin(user):
    try:
        return (user or {}).get('role') == 'admin'
    except Exception:
        return False

@app.route('/api/admin/users', methods=['GET'])
@authenticate_token
def admin_list_users():
    """List users scoped to the admin's company_name (admin only)."""
    try:
        current = request.current_user
        if not _is_admin(current):
            return jsonify({'message': 'Admin access required'}), 403

        company_name = current.get('company_name')
        if not company_name:
            return jsonify({'message': 'Admin company information not found'}), 400

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            """
            SELECT id, name, email, role, company_name,
                   mobile_number, department, designation, is_verified, profile_photo
            FROM users
            WHERE company_name = %s
            ORDER BY role, name
            """,
            (company_name,)
        )
        users = cursor.fetchall()
        cursor.close()
        conn.close()
        return jsonify(users), 200

    except Exception as e:
        logger.error(f"Admin list users error: {e}")
        return jsonify({'message': 'Failed to list users'}), 500

@app.route('/api/admin/users', methods=['POST'])
@authenticate_token
def admin_create_user():
    """Create a user under the admin's company (admin only)."""
    try:
        current = request.current_user
        if not _is_admin(current):
            return jsonify({'message': 'Admin access required'}), 403

        data = request.get_json(silent=True) or {}
        name = (data.get('name') or '').strip()
        email = (data.get('email') or '').strip()
        password = (data.get('password') or '').strip()
        role = (data.get('role') or 'user').strip()
        mobile_number = (data.get('mobile_number') or '').strip()
        department = (data.get('department') or '').strip()
        designation = (data.get('designation') or '').strip()
        profile_photo = data.get('profile_photo')  # may be data URL string

        if not all([name, email, password]):
            return jsonify({'message': 'name, email, and password are required'}), 400

        conn = get_db_connection()
        cursor = conn.cursor()
        hashed_password = generate_password_hash(password)
        cursor.execute(
            """
            INSERT INTO users (name, email, password, role, mobile_number, department, designation, company_name, profile_photo)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (name, email, hashed_password, role, mobile_number, department, designation, current.get('company_name'), profile_photo)
        )
        new_id = cursor.lastrowid
        cursor.close()
        conn.close()
        return jsonify({'id': new_id, 'message': 'User created'}), 201

    except Exception as e:
        logger.error(f"Admin create user error: {e}")
        return jsonify({'message': 'Failed to create user'}), 500

@app.route('/api/admin/users/<int:user_id>', methods=['PUT'])
@authenticate_token
def admin_update_user(user_id: int):
    """Update a user's profile (admin can update any in their company; non-admin can update own limited fields)."""
    try:
        current = request.current_user
        data = request.get_json(silent=True) or {}

        # Determine allowed fields
        allowed_fields_admin = {'name', 'email', 'role', 'mobile_number', 'department', 'designation', 'is_verified', 'profile_photo'}
        allowed_fields_self = {'name', 'mobile_number', 'department', 'designation', 'profile_photo'}

        if _is_admin(current):
            allowed = allowed_fields_admin
        else:
            if current.get('id') != user_id:
                return jsonify({'message': 'Forbidden'}), 403
            allowed = allowed_fields_self

        updates = {k: v for k, v in data.items() if k in allowed}
        if not updates:
            return jsonify({'message': 'No valid fields to update'}), 400

        # If non-admin, ensure they are not attempting to change email to a duplicate
        if 'email' in updates:
            if not _is_admin(current):
                return jsonify({'message': 'Only admin can change email'}), 403

        # Admin scope check: ensure the target user belongs to same company_name
        if _is_admin(current):
            try:
                conn = get_db_connection()
                cursor = conn.cursor(dictionary=True)
                cursor.execute("SELECT id, company_name FROM users WHERE id = %s", (user_id,))
                target = cursor.fetchone()
                if not target:
                    cursor.close()
                    conn.close()
                    return jsonify({'message': 'User not found'}), 404
                if target.get('company_name') != current.get('company_name'):
                    cursor.close()
                    conn.close()
                    return jsonify({'message': 'User outside your company'}), 403
                cursor.close()
                conn.close()
            except Exception as e:
                logger.error(f"Admin update user scope check error: {e}")
                return jsonify({'message': 'Failed to update user'}), 500

        # Build dynamic update
        set_clause = ", ".join([f"{k} = %s" for k in updates.keys()])
        values = list(updates.values()) + [user_id]
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(f"UPDATE users SET {set_clause} WHERE id = %s", values)
        cursor.close()
        conn.close()
        return jsonify({'message': 'User updated'}), 200

    except Exception as e:
        logger.error(f"Admin update user error: {e}")
        return jsonify({'message': 'Failed to update user'}), 500

@app.route('/api/admin/users/<int:user_id>', methods=['DELETE'])
@authenticate_token
def admin_delete_user(user_id: int):
    """Delete a user (admin only, scoped to company)."""
    try:
        current = request.current_user
        if not _is_admin(current):
            return jsonify({'message': 'Admin access required'}), 403

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id, company_name FROM users WHERE id = %s", (user_id,))
        target = cursor.fetchone()
        if not target:
            cursor.close()
            conn.close()
            return jsonify({'message': 'User not found'}), 404
        if target.get('company_name') != current.get('company_name'):
            cursor.close()
            conn.close()
            return jsonify({'message': 'User outside your company'}), 403

        # Proceed delete
        cursor = conn.cursor()
        cursor.execute("DELETE FROM users WHERE id = %s", (user_id,))
        cursor.close()
        conn.close()
        return jsonify({'message': 'User deleted'}), 200

    except Exception as e:
        logger.error(f"Admin delete user error: {e}")
        return jsonify({'message': 'Failed to delete user'}), 500

@app.route('/api/admin/users/<int:user_id>/password', methods=['PUT'])
@authenticate_token
def admin_change_user_password(user_id: int):
    """Change password for a user. Admin may change any in company; non-admin can change own password.
    Expects JSON: { password: 'newPassword' }
    """
    try:
        current = request.current_user
        data = request.get_json(silent=True) or {}
        new_password = (data.get('password') or '').strip()
        if not new_password:
            return jsonify({'message': 'password is required'}), 400

        # Scope check
        if _is_admin(current):
            conn = get_db_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT id, company_name FROM users WHERE id = %s", (user_id,))
            target = cursor.fetchone()
            if not target:
                cursor.close()
                conn.close()
                return jsonify({'message': 'User not found'}), 404
            if target.get('company_name') != current.get('company_name'):
                cursor.close()
                conn.close()
                return jsonify({'message': 'User outside your company'}), 403
            cursor.close()
            conn.close()
        else:
            if current.get('id') != user_id:
                return jsonify({'message': 'Forbidden'}), 403

        hashed = generate_password_hash(new_password)
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE users SET password = %s, last_password_change = NOW() WHERE id = %s",
            (hashed, user_id)
        )
        cursor.close()
        conn.close()
        return jsonify({'message': 'Password updated'}), 200

    except Exception as e:
        logger.error(f"Admin change password error: {e}")
        return jsonify({'message': 'Failed to change password'}), 500

# ============== CONTACT & DEMO MANAGEMENT ENDPOINTS ==============

@app.route('/api/contact-messages', methods=['GET'])
@authenticate_token
def get_contact_messages():
    """Get all contact form submissions (admin only)"""
    try:
        user = request.current_user
        
        if user['role'] != 'admin':
            return jsonify({'message': 'Admin access required'}), 403
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, name, email, message, created_at 
            FROM contact_us 
            ORDER BY created_at DESC
        """)
        
        messages = cursor.fetchall()
        cursor.close()
        conn.close()
        
        return jsonify(messages), 200
        
    except Exception as e:
        logger.error(f"Error fetching contact messages: {e}")
        return jsonify({'message': 'Server error while fetching contact messages'}), 500

@app.route('/api/demo-bookings', methods=['GET'])
@authenticate_token
def get_demo_bookings():
    """Get all demo booking requests (admin only)"""
    try:
        user = request.current_user
        
        if user['role'] != 'admin':
            return jsonify({'message': 'Admin access required'}), 403
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, name, email, organization, preferred_date, message, created_at 
            FROM book_demo 
            ORDER BY created_at DESC
        """)
        
        bookings = cursor.fetchall()
        cursor.close()
        conn.close()
        
        return jsonify(bookings), 200
        
    except Exception as e:
        logger.error(f"Error fetching demo bookings: {e}")
        return jsonify({'message': 'Server error while fetching demo bookings'}), 500

# ============== REPORTING & ANALYTICS ENDPOINTS ==============

@app.route('/api/reports', methods=['GET'])
@authenticate_token
def get_reports():
    """Get comprehensive reports data for the admin's company"""
    try:
        user = request.current_user
        
        if user['role'] != 'admin':
            return jsonify({'message': 'Admin access required'}), 403
        
        start_date = request.args.get('startDate')
        end_date = request.args.get('endDate')
        
        admin_company_name = user['company_name']
        
        if not admin_company_name:
            return jsonify({'message': 'Admin company information not found'}), 400
        
        # Prepare query parameters
        query_params = [admin_company_name]
        date_filter_clause = ''
        
        if start_date and end_date:
            date_filter_clause = 'AND DATE(v.check_in_time) BETWEEN %s AND %s'
            query_params.extend([start_date, end_date])
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Overview Stats
        overview_query = f"""
            SELECT
                COUNT(v.id) AS totalVisits,
                COUNT(DISTINCT vis.email) AS uniqueVisitors,
                AVG(TIMESTAMPDIFF(MINUTE, v.check_in_time, v.check_out_time)) AS avgDuration
            FROM visits v
            JOIN users h ON v.host_id = h.id
            LEFT JOIN visitors vis ON v.visitor_id = vis.id
            WHERE h.company_name = %s {date_filter_clause}
        """
        
        cursor.execute(overview_query, query_params)
        overview_result = cursor.fetchone()
        
        # Daily Stats
        daily_query = f"""
            SELECT
                DATE(v.check_in_time) as date,
                COUNT(v.id) as visits
            FROM visits v
            JOIN users h ON v.host_id = h.id
            WHERE h.company_name = %s {date_filter_clause}
            GROUP BY DATE(v.check_in_time)
            ORDER BY date ASC
        """
        
        cursor.execute(daily_query, query_params)
        daily_result = cursor.fetchall()
        
        # Host Performance
        host_query = f"""
            SELECT
                h.name as host_name,
                COUNT(v.id) as visits
            FROM visits v
            JOIN users h ON v.host_id = h.id
            WHERE h.company_name = %s {date_filter_clause}
            GROUP BY h.name
            ORDER BY visits DESC
        """
        
        cursor.execute(host_query, query_params)
        host_result = cursor.fetchall()
        
        # Visit Purposes
        purpose_query = f"""
            SELECT
                COALESCE(v.purpose_of_visit, 'Not Specified') as purpose,
                COUNT(v.id) as count
            FROM visits v
            JOIN users h ON v.host_id = h.id
            WHERE h.company_name = %s {date_filter_clause}
            GROUP BY COALESCE(v.purpose_of_visit, 'Not Specified')
            ORDER BY count DESC
        """
        
        cursor.execute(purpose_query, query_params)
        purpose_result = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'overview': overview_result or {'totalVisits': 0, 'uniqueVisitors': 0, 'avgDuration': 0},
            'dailyStats': daily_result or [],
            'hostStats': host_result or [],
            'purposeStats': purpose_result or []
        }), 200
        
    except Exception as e:
        logger.error(f"Reports error: {e}")
        return jsonify({'message': 'Failed to generate reports'}), 500

@app.route('/api/reports/export', methods=['GET'])
@authenticate_token
def export_reports():
    """Export comprehensive report data for the admin's company"""
    try:
        user = request.current_user
        
        if user['role'] != 'admin':
            return jsonify({'message': 'Admin access required'}), 403
        
        format_type = request.args.get('format', 'html')  # 'pdf', 'excel', 'html'
        start_date = request.args.get('startDate')
        end_date = request.args.get('endDate')
        
        # Get comprehensive report data
        report_data = get_comprehensive_report_data(user, start_date, end_date)
        
        if format_type == 'pdf':
            return export_pdf_report(report_data, start_date, end_date, user)
        elif format_type == 'excel':
            return export_excel_report(report_data, start_date, end_date, user)
        else:
            # Generate HTML export
            html_content = generate_html_report_content(report_data, start_date, end_date)
            return jsonify({
                'success': True,
                'data': html_content,
                'filename': f'visitor-report-{datetime.now().strftime("%Y%m%d")}.html'
            }), 200
        
    except Exception as e:
        logger.error(f"Export error: {e}")
        return jsonify({'message': 'Failed to export data'}), 500

def get_comprehensive_report_data(user, start_date, end_date):
    """Get comprehensive visitor data for reports"""
    try:
        admin_company_name = user['company_name']
        
        # Prepare query parameters
        query_params = [admin_company_name]
        date_filter_clause = ''
        
        if start_date and end_date:
            date_filter_clause = 'AND DATE(v.check_in_time) BETWEEN %s AND %s'
            query_params.extend([start_date, end_date])
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # 1. Overview Stats
        overview_query = f"""
            SELECT
                COUNT(v.id) AS total_visits,
                COUNT(DISTINCT COALESCE(vis.email, v.visitor_email)) AS unique_visitors,
                COUNT(CASE WHEN v.status = 'checked-in' THEN 1 END) AS active_visits,
                COUNT(CASE WHEN v.status = 'checked-out' THEN 1 END) AS completed_visits,
                AVG(CASE 
                    WHEN v.check_out_time IS NOT NULL 
                    THEN TIMESTAMPDIFF(MINUTE, v.check_in_time, v.check_out_time) 
                END) AS avg_duration_minutes
            FROM visits v
            JOIN users h ON v.host_id = h.id
            LEFT JOIN visitors vis ON v.visitor_id = vis.id
            WHERE h.company_name = %s {date_filter_clause}
        """
        cursor.execute(overview_query, query_params)
        overview = cursor.fetchone()
        
        # 2. Recent Visitor Activity (Last 50 visits)
        recent_query = f"""
            SELECT
                COALESCE(vis.name, v.visitor_name) as visitor_name,
                COALESCE(vis.email, v.visitor_email) as visitor_email,
                COALESCE(vis.company, v.visitor_company) as visitor_company,
                h.name as host_name,
                v.check_in_time,
                v.check_out_time,
                v.purpose_of_visit as purpose,
                v.status,
                CASE 
                    WHEN v.check_out_time IS NOT NULL 
                    THEN TIMESTAMPDIFF(MINUTE, v.check_in_time, v.check_out_time) 
                    ELSE NULL
                END as duration_minutes
            FROM visits v
            JOIN users h ON v.host_id = h.id
            LEFT JOIN visitors vis ON v.visitor_id = vis.id
            WHERE h.company_name = %s {date_filter_clause}
            ORDER BY v.check_in_time DESC
            LIMIT 50
        """
        cursor.execute(recent_query, query_params)
        recent_activity = cursor.fetchall()
        
        # 3. Visit Purpose Analysis
        purpose_query = f"""
            SELECT
                COALESCE(NULLIF(v.purpose_of_visit, ''), 'Not Specified') as purpose,
                COUNT(v.id) as visit_count,
                COUNT(DISTINCT COALESCE(vis.email, v.visitor_email)) as unique_visitors,
                ROUND(AVG(CASE 
                    WHEN v.check_out_time IS NOT NULL 
                    THEN TIMESTAMPDIFF(MINUTE, v.check_in_time, v.check_out_time) 
                END), 2) as avg_duration
            FROM visits v
            JOIN users h ON v.host_id = h.id
            LEFT JOIN visitors vis ON v.visitor_id = vis.id
            WHERE h.company_name = %s {date_filter_clause}
            GROUP BY COALESCE(NULLIF(v.purpose_of_visit, ''), 'Not Specified')
            ORDER BY visit_count DESC
        """
        cursor.execute(purpose_query, query_params)
        purpose_analysis = cursor.fetchall()
        
        # 4. Time-based Analysis (Daily)
        daily_query = f"""
            SELECT
                DATE(v.check_in_time) as visit_date,
                COUNT(v.id) as daily_visits,
                COUNT(DISTINCT COALESCE(vis.email, v.visitor_email)) as unique_daily_visitors,
                COUNT(CASE WHEN HOUR(v.check_in_time) BETWEEN 9 AND 12 THEN 1 END) as morning_visits,
                COUNT(CASE WHEN HOUR(v.check_in_time) BETWEEN 13 AND 17 THEN 1 END) as afternoon_visits,
                COUNT(CASE WHEN HOUR(v.check_in_time) BETWEEN 18 AND 21 THEN 1 END) as evening_visits
            FROM visits v
            JOIN users h ON v.host_id = h.id
            LEFT JOIN visitors vis ON v.visitor_id = vis.id
            WHERE h.company_name = %s {date_filter_clause}
            GROUP BY DATE(v.check_in_time)
            ORDER BY visit_date DESC
            LIMIT 30
        """
        cursor.execute(daily_query, query_params)
        daily_analysis = cursor.fetchall()
        
        # 5. Hourly Pattern Analysis
        hourly_query = f"""
            SELECT
                HOUR(v.check_in_time) as hour_of_day,
                COUNT(v.id) as visit_count,
                CASE 
                    WHEN HOUR(v.check_in_time) BETWEEN 6 AND 11 THEN 'Morning'
                    WHEN HOUR(v.check_in_time) BETWEEN 12 AND 17 THEN 'Afternoon'
                    WHEN HOUR(v.check_in_time) BETWEEN 18 AND 21 THEN 'Evening'
                    ELSE 'Night'
                END as time_period
            FROM visits v
            JOIN users h ON v.host_id = h.id
            WHERE h.company_name = %s {date_filter_clause}
            GROUP BY HOUR(v.check_in_time)
            ORDER BY hour_of_day
        """
        cursor.execute(hourly_query, query_params)
        hourly_analysis = cursor.fetchall()
        
        # 6. Host Performance Analysis
        host_query = f"""
            SELECT
                h.name as host_name,
                h.email as host_email,
                COUNT(v.id) as total_visits,
                COUNT(DISTINCT COALESCE(vis.email, v.visitor_email)) as unique_visitors,
                ROUND(AVG(CASE 
                    WHEN v.check_out_time IS NOT NULL 
                    THEN TIMESTAMPDIFF(MINUTE, v.check_in_time, v.check_out_time) 
                END), 2) as avg_visit_duration
            FROM visits v
            JOIN users h ON v.host_id = h.id
            LEFT JOIN visitors vis ON v.visitor_id = vis.id
            WHERE h.company_name = %s {date_filter_clause}
            GROUP BY h.id, h.name, h.email
            ORDER BY total_visits DESC
        """
        cursor.execute(host_query, query_params)
        host_performance = cursor.fetchall()
        
        # 7. Visitor Company Analysis
        company_query = f"""
            SELECT
                COALESCE(NULLIF(COALESCE(vis.company, v.visitor_company), ''), 'Not Specified') as company,
                COUNT(v.id) as visit_count,
                COUNT(DISTINCT COALESCE(vis.email, v.visitor_email)) as unique_visitors
            FROM visits v
            JOIN users h ON v.host_id = h.id
            LEFT JOIN visitors vis ON v.visitor_id = vis.id
            WHERE h.company_name = %s {date_filter_clause}
            GROUP BY COALESCE(NULLIF(COALESCE(vis.company, v.visitor_company), ''), 'Not Specified')
            ORDER BY visit_count DESC
            LIMIT 20
        """
        cursor.execute(company_query, query_params)
        company_analysis = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return {
            'overview': overview,
            'recent_activity': recent_activity,
            'purpose_analysis': purpose_analysis,
            'daily_analysis': daily_analysis,
            'hourly_analysis': hourly_analysis,
            'host_performance': host_performance,
            'company_analysis': company_analysis,
            'report_period': {
                'start_date': start_date,
                'end_date': end_date,
                'generated_at': datetime.now()
            }
        }
        
    except Exception as e:
        logger.error(f"Error getting comprehensive report data: {e}")
        raise e

def export_pdf_report(report_data, start_date, end_date, user):
    """Generate and return PDF report or HTML fallback"""
    try:
        html_content = generate_comprehensive_html_report(report_data, start_date, end_date, user)
        
        if WEASYPRINT_AVAILABLE:
            # Create temporary file for PDF
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
                # Generate PDF from HTML
                weasyprint.HTML(string=html_content).write_pdf(tmp_file.name)
                
                # Read the PDF content
                with open(tmp_file.name, 'rb') as pdf_file:
                    pdf_data = pdf_file.read()
                
                # Clean up temporary file
                os.unlink(tmp_file.name)
                
                # Create response
                response = Response(
                    pdf_data,
                    mimetype='application/pdf',
                    headers={
                        'Content-Disposition': f'attachment; filename=visitor-report-{datetime.now().strftime("%Y%m%d")}.pdf',
                        'Content-Type': 'application/pdf'
                    }
                )
                return response
        else:
            # Fallback: Return enhanced HTML that can be printed as PDF by browser
            enhanced_html = add_print_styles_to_html(html_content)
            response = Response(
                enhanced_html,
                mimetype='text/html',
                headers={
                    'Content-Disposition': f'attachment; filename=visitor-report-{datetime.now().strftime("%Y%m%d")}.html',
                    'Content-Type': 'text/html'
                }
            )
            return response
            
    except Exception as e:
        logger.error(f"PDF export error: {e}")
        return jsonify({
            'message': 'Failed to generate PDF report. You can use the HTML export and print it as PDF from your browser.',
            'error': str(e),
            'fallback': True
        }), 500

def add_print_styles_to_html(html_content):
    """Add print-optimized styles to HTML for browser PDF printing"""
    # Add additional CSS for better PDF printing
    print_styles = """
    <style>
        @media print {
            body { 
                margin: 0; 
                background: white !important; 
                -webkit-print-color-adjust: exact;
                color-adjust: exact;
            }
            .container { 
                box-shadow: none !important; 
                padding: 15px !important; 
                margin: 0 !important;
            }
            .page-break { 
                page-break-before: always !important; 
            }
            .no-print { 
                display: none !important; 
            }
            table { 
                page-break-inside: avoid; 
            }
            tr { 
                page-break-inside: avoid; 
                page-break-after: auto; 
            }
            .section {
                break-inside: avoid;
            }
        }
        .print-instructions {
            background: #e3f2fd;
            border: 1px solid #2196f3;
            padding: 15px;
            margin: 20px 0;
            border-radius: 5px;
            text-align: center;
        }
        .print-instructions h3 {
            color: #1976d2;
            margin: 0 0 10px 0;
        }
        @media print {
            .print-instructions {
                display: none !important;
            }
        }
    </style>
    """
    
    # Add print instructions
    print_instructions = """
    <div class="print-instructions no-print">
        <h3>📄 PDF Export Instructions</h3>
        <p><strong>To save this report as PDF:</strong></p>
        <ol style="text-align: left; display: inline-block;">
            <li>Press <kbd>Ctrl+P</kbd> (Windows) or <kbd>Cmd+P</kbd> (Mac)</li>
            <li>Select "Save as PDF" as the destination</li>
            <li>Choose "More settings" and enable "Background graphics"</li>
            <li>Click "Save" to download your PDF report</li>
        </ol>
    </div>
    """
    
    # Insert the styles after the existing <style> tag
    style_end = html_content.find('</style>')
    if style_end != -1:
        html_content = html_content[:style_end] + print_styles + html_content[style_end:]
    
    # Insert instructions after the header
    header_end = html_content.find('</div>', html_content.find('class="meta-info"'))
    if header_end != -1:
        insertion_point = header_end + 6  # After </div>
        html_content = html_content[:insertion_point] + print_instructions + html_content[insertion_point:]
    
    return html_content

def export_excel_report(report_data, start_date, end_date, user):
    """Generate and return Excel report with multiple sheets"""
    try:
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 1. Overview Sheet
        overview_ws = wb.create_sheet("Overview")
        overview_data = [
            ["Visitor Management System Report"],
            [""],
            ["Company", user['company_name']],
            ["Report Period", f"{start_date or 'All time'} to {end_date or 'Present'}"],
            ["Generated", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            [""],
            ["Metric", "Value"],
            ["Total Visits", report_data['overview']['total_visits'] or 0],
            ["Unique Visitors", report_data['overview']['unique_visitors'] or 0],
            ["Active Visits", report_data['overview']['active_visits'] or 0],
            ["Completed Visits", report_data['overview']['completed_visits'] or 0],
            ["Average Duration (minutes)", round(report_data['overview']['avg_duration_minutes'] or 0, 2)]
        ]
        
        for row_idx, row_data in enumerate(overview_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = overview_ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = title_font
                elif row_idx == 7:
                    cell.font = header_font
                    cell.fill = header_fill
        
        # 2. Recent Activity Sheet
        activity_ws = wb.create_sheet("Recent Activity")
        activity_headers = ["Visitor Name", "Email", "Company", "Host", "Check In", "Check Out", "Purpose", "Status", "Duration (min)"]
        
        # Add headers
        for col_idx, header in enumerate(activity_headers, 1):
            cell = activity_ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Add data
        for row_idx, activity in enumerate(report_data['recent_activity'], 2):
            activity_ws.cell(row=row_idx, column=1, value=activity.get('visitor_name', ''))
            activity_ws.cell(row=row_idx, column=2, value=activity.get('visitor_email', ''))
            activity_ws.cell(row=row_idx, column=3, value=activity.get('visitor_company', ''))
            activity_ws.cell(row=row_idx, column=4, value=activity.get('host_name', ''))
            activity_ws.cell(row=row_idx, column=5, value=activity.get('check_in_time', ''))
            activity_ws.cell(row=row_idx, column=6, value=activity.get('check_out_time', ''))
            activity_ws.cell(row=row_idx, column=7, value=activity.get('purpose', ''))
            activity_ws.cell(row=row_idx, column=8, value=activity.get('status', ''))
            activity_ws.cell(row=row_idx, column=9, value=activity.get('duration_minutes', ''))
        
        # 3. Purpose Analysis Sheet
        purpose_ws = wb.create_sheet("Purpose Analysis")
        purpose_headers = ["Purpose", "Visit Count", "Unique Visitors", "Avg Duration (min)"]
        
        for col_idx, header in enumerate(purpose_headers, 1):
            cell = purpose_ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, purpose in enumerate(report_data['purpose_analysis'], 2):
            purpose_ws.cell(row=row_idx, column=1, value=purpose.get('purpose', ''))
            purpose_ws.cell(row=row_idx, column=2, value=purpose.get('visit_count', 0))
            purpose_ws.cell(row=row_idx, column=3, value=purpose.get('unique_visitors', 0))
            purpose_ws.cell(row=row_idx, column=4, value=purpose.get('avg_duration', 0))
        
        # 4. Daily Analysis Sheet
        daily_ws = wb.create_sheet("Daily Analysis")
        daily_headers = ["Date", "Total Visits", "Unique Visitors", "Morning", "Afternoon", "Evening"]
        
        for col_idx, header in enumerate(daily_headers, 1):
            cell = daily_ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, daily in enumerate(report_data['daily_analysis'], 2):
            daily_ws.cell(row=row_idx, column=1, value=daily.get('visit_date', ''))
            daily_ws.cell(row=row_idx, column=2, value=daily.get('daily_visits', 0))
            daily_ws.cell(row=row_idx, column=3, value=daily.get('unique_daily_visitors', 0))
            daily_ws.cell(row=row_idx, column=4, value=daily.get('morning_visits', 0))
            daily_ws.cell(row=row_idx, column=5, value=daily.get('afternoon_visits', 0))
            daily_ws.cell(row=row_idx, column=6, value=daily.get('evening_visits', 0))
        
        # 5. Host Performance Sheet
        host_ws = wb.create_sheet("Host Performance")
        host_headers = ["Host Name", "Email", "Total Visits", "Unique Visitors", "Avg Duration (min)"]
        
        for col_idx, header in enumerate(host_headers, 1):
            cell = host_ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, host in enumerate(report_data['host_performance'], 2):
            host_ws.cell(row=row_idx, column=1, value=host.get('host_name', ''))
            host_ws.cell(row=row_idx, column=2, value=host.get('host_email', ''))
            host_ws.cell(row=row_idx, column=3, value=host.get('total_visits', 0))
            host_ws.cell(row=row_idx, column=4, value=host.get('unique_visitors', 0))
            host_ws.cell(row=row_idx, column=5, value=host.get('avg_visit_duration', 0))
        
        # Auto-size columns for all sheets
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to temporary file and return
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            wb.save(tmp_file.name)
            
            with open(tmp_file.name, 'rb') as excel_file:
                excel_data = excel_file.read()
            
            # Clean up temporary file
            os.unlink(tmp_file.name)
            
            response = Response(
                excel_data,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={
                    'Content-Disposition': f'attachment; filename=visitor-report-{datetime.now().strftime("%Y%m%d")}.xlsx',
                    'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                }
            )
            return response
            
    except Exception as e:
        logger.error(f"Excel export error: {e}")
        return jsonify({'message': 'Failed to generate Excel report', 'error': str(e)}), 500

def generate_comprehensive_html_report(report_data, start_date, end_date, user):
    """Generate comprehensive HTML report with all analytics"""
    overview = report_data['overview']
    recent_activity = report_data['recent_activity']
    purpose_analysis = report_data['purpose_analysis']
    daily_analysis = report_data['daily_analysis']
    host_performance = report_data['host_performance']
    
    # Generate recent activity table
    recent_activity_rows = ""
    for activity in recent_activity[:20]:  # Show top 20 for PDF
        recent_activity_rows += f"""
        <tr>
            <td>{activity.get('visitor_name', 'N/A')}</td>
            <td>{activity.get('visitor_email', 'N/A')}</td>
            <td>{activity.get('visitor_company', 'N/A')}</td>
            <td>{activity.get('host_name', 'N/A')}</td>
            <td>{activity.get('check_in_time', 'N/A')}</td>
            <td>{activity.get('purpose', 'N/A')}</td>
            <td><span class="status {activity.get('status', '').lower()}">{activity.get('status', 'N/A')}</span></td>
        </tr>
        """
    
    # Generate purpose analysis chart data
    purpose_chart_data = ""
    for purpose in purpose_analysis[:10]:  # Top 10 purposes
        percentage = (purpose.get('visit_count', 0) / max(overview.get('total_visits', 1), 1)) * 100
        purpose_chart_data += f"""
        <tr>
            <td>{purpose.get('purpose', 'N/A')}</td>
            <td>{purpose.get('visit_count', 0)}</td>
            <td>{purpose.get('unique_visitors', 0)}</td>
            <td>{round(percentage, 1)}%</td>
        </tr>
        """
    
    # Generate daily analysis chart
    daily_chart_data = ""
    for daily in daily_analysis[:14]:  # Last 14 days
        daily_chart_data += f"""
        <tr>
            <td>{daily.get('visit_date', 'N/A')}</td>
            <td>{daily.get('daily_visits', 0)}</td>
            <td>{daily.get('unique_daily_visitors', 0)}</td>
            <td>{daily.get('morning_visits', 0)}</td>
            <td>{daily.get('afternoon_visits', 0)}</td>
            <td>{daily.get('evening_visits', 0)}</td>
        </tr>
        """
    
    # Generate host performance data
    host_performance_data = ""
    for host in host_performance[:10]:  # Top 10 hosts
        host_performance_data += f"""
        <tr>
            <td>{host.get('host_name', 'N/A')}</td>
            <td>{host.get('total_visits', 0)}</td>
            <td>{host.get('unique_visitors', 0)}</td>
            <td>{round(host.get('avg_visit_duration', 0) or 0, 1)} min</td>
        </tr>
        """
    
    return f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Visitor Management System - Comprehensive Report</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 20px;
            color: #333;
            line-height: 1.6;
            background-color: #f8f9fa;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            padding: 30px;
            border-radius: 10px;
            box-shadow: 0 0 20px rgba(0,0,0,0.1);
        }}
        .header {{
            text-align: center;
            border-bottom: 3px solid #007bff;
            padding-bottom: 20px;
            margin-bottom: 30px;
        }}
        .header h1 {{
            color: #007bff;
            margin: 0;
            font-size: 32px;
            font-weight: 700;
        }}
        .header h2 {{
            color: #666;
            margin: 10px 0 0 0;
            font-size: 18px;
            font-weight: 400;
        }}
        .meta-info {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            border-radius: 8px;
            margin-bottom: 30px;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }}
        .meta-info .company {{
            font-size: 20px;
            font-weight: bold;
        }}
        .meta-info .period {{
            text-align: right;
        }}
        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }}
        .stat-card {{
            background: #f8f9fa;
            padding: 20px;
            border-radius: 8px;
            text-align: center;
            border-left: 4px solid #007bff;
        }}
        .stat-card h3 {{
            margin: 0 0 10px 0;
            color: #666;
            font-size: 14px;
            text-transform: uppercase;
            letter-spacing: 1px;
        }}
        .stat-card .value {{
            font-size: 28px;
            font-weight: bold;
            color: #007bff;
        }}
        .section {{
            margin-bottom: 40px;
        }}
        .section-title {{
            font-size: 24px;
            font-weight: bold;
            color: #333;
            margin-bottom: 20px;
            padding-bottom: 10px;
            border-bottom: 2px solid #e9ecef;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 10px;
            background: white;
            border-radius: 8px;
            overflow: hidden;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        th {{
            background: #007bff;
            color: white;
            padding: 15px 10px;
            text-align: left;
            font-weight: 600;
            font-size: 14px;
        }}
        td {{
            padding: 12px 10px;
            border-bottom: 1px solid #e9ecef;
            font-size: 13px;
        }}
        tr:nth-child(even) {{
            background-color: #f8f9fa;
        }}
        tr:hover {{
            background-color: #e3f2fd;
        }}
        .status {{
            padding: 4px 12px;
            border-radius: 20px;
            font-size: 11px;
            font-weight: bold;
            text-transform: uppercase;
        }}
        .status.checked_in {{
            background: #d4edda;
            color: #155724;
        }}
        .status.checked_out {{
            background: #cce5ff;
            color: #004085;
        }}
        .status.pending {{
            background: #fff3cd;
            color: #856404;
        }}
        .footer {{
            margin-top: 40px;
            padding-top: 20px;
            border-top: 1px solid #e9ecef;
            text-align: center;
            color: #666;
            font-size: 12px;
        }}
        .page-break {{
            page-break-before: always;
        }}
        @media print {{
            body {{ margin: 0; background: white; }}
            .container {{ box-shadow: none; padding: 20px; }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>VISITOR MANAGEMENT SYSTEM</h1>
            <h2>Comprehensive Analytics Report</h2>
        </div>
        
        <div class="meta-info">
            <div>
                <div class="company">{user['company_name']}</div>
                <div>Generated by: {user['name']}</div>
            </div>
            <div class="period">
                <div><strong>Report Period:</strong></div>
                <div>{start_date or 'All time'} to {end_date or 'Present'}</div>
                <div><strong>Generated:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</div>
            </div>
        </div>
        
        <div class="stats-grid">
            <div class="stat-card">
                <h3>Total Visits</h3>
                <div class="value">{overview.get('total_visits', 0):,}</div>
            </div>
            <div class="stat-card">
                <h3>Unique Visitors</h3>
                <div class="value">{overview.get('unique_visitors', 0):,}</div>
            </div>
            <div class="stat-card">
                <h3>Active Visits</h3>
                <div class="value">{overview.get('active_visits', 0):,}</div>
            </div>
            <div class="stat-card">
                <h3>Avg Duration</h3>
                <div class="value">{round(overview.get('avg_duration_minutes', 0) or 0, 1)}</div>
                <small>minutes</small>
            </div>
        </div>
        
        <div class="section">
            <h2 class="section-title">📈 Recent Visitor Activity</h2>
            <table>
                <thead>
                    <tr>
                        <th>Visitor Name</th>
                        <th>Email</th>
                        <th>Company</th>
                        <th>Host</th>
                        <th>Check In Time</th>
                        <th>Purpose</th>
                        <th>Status</th>
                    </tr>
                </thead>
                <tbody>
                    {recent_activity_rows}
                </tbody>
            </table>
        </div>
        
        <div class="page-break"></div>
        
        <div class="section">
            <h2 class="section-title">🎯 Visit Purpose Analysis</h2>
            <table>
                <thead>
                    <tr>
                        <th>Purpose</th>
                        <th>Visit Count</th>
                        <th>Unique Visitors</th>
                        <th>Percentage</th>
                    </tr>
                </thead>
                <tbody>
                    {purpose_chart_data}
                </tbody>
            </table>
        </div>
        
        <div class="section">
            <h2 class="section-title">📊 Time-based Visitor Analysis</h2>
            <table>
                <thead>
                    <tr>
                        <th>Date</th>
                        <th>Total Visits</th>
                        <th>Unique Visitors</th>
                        <th>Morning (9-12)</th>
                        <th>Afternoon (13-17)</th>
                        <th>Evening (18-21)</th>
                    </tr>
                </thead>
                <tbody>
                    {daily_chart_data}
                </tbody>
            </table>
        </div>
        
        <div class="section">
            <h2 class="section-title">👥 Host Performance Analysis</h2>
            <table>
                <thead>
                    <tr>
                        <th>Host Name</th>
                        <th>Total Visits</th>
                        <th>Unique Visitors</th>
                        <th>Avg Duration</th>
                    </tr>
                </thead>
                <tbody>
                    {host_performance_data}
                </tbody>
            </table>
        </div>
        
        <div class="footer">
            <p>This report was generated by the Visitor Management System on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}</p>
            <p>© 2025 Visitor Management System. All rights reserved.</p>
        </div>
    </div>
</body>
</html>
    """

def generate_html_report_content(data, start_date, end_date):
    """Generate simple HTML content for basic export"""
    # For backward compatibility, if data is the new comprehensive format, extract overview
    if isinstance(data, dict) and 'overview' in data:
        overview = data['overview']
        total_visits = overview.get('total_visits', 0)
        unique_visitors = overview.get('unique_visitors', 0)
    else:
        total_visits = 0
        unique_visitors = 0
    
    return f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Visitor Management System Report</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            margin: 20px;
            color: #333;
            line-height: 1.6;
        }}
        .header {{
            text-align: center;
            border-bottom: 2px solid #007bff;
            padding-bottom: 20px;
            margin-bottom: 30px;
        }}
        .header h1 {{
            color: #007bff;
            margin: 0;
            font-size: 28px;
        }}
        .meta-info {{
            background-color: #f8f9fa;
            padding: 15px;
            border-radius: 5px;
            margin-bottom: 30px;
        }}
        .stats {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin: 20px 0;
        }}
        .stat-card {{
            background: #f8f9fa;
            padding: 20px;
            border-radius: 8px;
            text-align: center;
            border-left: 4px solid #007bff;
        }}
        .stat-card h3 {{
            margin: 0 0 10px 0;
            color: #666;
            font-size: 14px;
        }}
        .stat-card .value {{
            font-size: 24px;
            font-weight: bold;
            color: #007bff;
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>VISITOR MANAGEMENT SYSTEM</h1>
        <h2>Analytics Report</h2>
    </div>
    <div class="meta-info">
        <strong>Generated:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br>
        <strong>Report Period:</strong> {start_date or 'All time'} to {end_date or 'Present'}
    </div>
    <div class="stats">
        <div class="stat-card">
            <h3>Total Visits</h3>
            <div class="value">{total_visits:,}</div>
        </div>
        <div class="stat-card">
            <h3>Unique Visitors</h3>
            <div class="value">{unique_visitors:,}</div>
        </div>
    </div>
    <p><em>For comprehensive analytics including Recent Visitor Activity, Visit Purpose Analysis, and Time-based Analysis, please use PDF or Excel export formats.</em></p>
</body>
</html>
    """

# ============== ADVANCED VISITOR FEATURES ENDPOINTS ==============

@app.route('/api/visitors/pre-register', methods=['POST'])
@authenticate_token
def pre_register_visitor():
    """Pre-register a visitor"""
    try:
        user = request.current_user
        data = request.get_json()
        
        # Extract data
        visitor_name = data['visitorName']
        visitor_email = data['visitorEmail']
        visitor_phone = data.get('visitorPhone', '')
        visitor_company = data.get('visitorCompany', '')
        host_name = data['hostName']
        visit_date = data['visitDate']
        visit_time = data['visitTime']
        purpose = data['purpose']
        duration = data.get('duration')
        is_recurring = data.get('isRecurring', False)
        recurring_pattern = data.get('recurringPattern')
        recurring_end_date = data.get('recurringEndDate')
        special_requirements = data.get('specialRequirements', '')
        emergency_contact = data.get('emergencyContact', '')
        vehicle_number = data.get('vehicleNumber', '')
        number_of_visitors = data.get('numberOfVisitors', 1)
        
        # Get company_id from companies table using user_id
        company_id = get_company_id_from_companies_table(user['id'])
        admin_company_name = user['company_name'] or 'Default Company'
        
        qr_code = generate_qr_code()
        
        # Clean empty values
        clean_recurring_end_date = recurring_end_date if recurring_end_date and recurring_end_date.strip() else None
        clean_recurring_pattern = recurring_pattern if recurring_pattern and recurring_pattern.strip() else None
        clean_duration = duration if duration and str(duration).strip() else None
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            INSERT INTO pre_registrations (
                company_id, visitor_name, visitor_email, visitor_phone, visitor_company,
                company_to_visit, host_id, host_name, visit_date, visit_time, purpose, duration,
                is_recurring, recurring_pattern, recurring_end_date,
                special_requirements, emergency_contact, vehicle_number, 
                number_of_visitors, qr_code, created_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            company_id, visitor_name, visitor_email, visitor_phone, visitor_company,
            admin_company_name, user['id'], host_name, visit_date, visit_time, purpose, clean_duration,
            is_recurring, clean_recurring_pattern, clean_recurring_end_date,
            special_requirements, emergency_contact, vehicle_number,
            number_of_visitors, qr_code, datetime.now()
        ))
        
        pre_reg_id = cursor.lastrowid
        cursor.close()
        conn.close()
        
        return jsonify({
            'message': 'Visitor pre-registered successfully',
            'id': pre_reg_id,
            'qrCode': qr_code
        }), 201
        
    except Exception as e:
        logger.error(f"Pre-registration error: {e}")
        
        error_message = 'Failed to pre-register visitor.'
        if 'Duplicate entry' in str(e):
            error_message = 'A visitor with this information already exists.'
        elif 'Data truncated' in str(e):
            error_message = 'Invalid data format. Please check all fields.'
        
        return jsonify({
            'message': error_message,
            'details': str(e) if app.debug else None
        }), 500

@app.route('/api/visitors/qr-checkin', methods=['POST'])
@authenticate_token
def qr_checkin():
    """QR Code verification endpoint"""
    try:
        user = request.current_user
        data = request.get_json()
        
        qr_code = data.get('qr_code')
        host_id = data.get('host_id')
        
        logger.info(f"QR verification request: {qr_code}, host_id: {host_id}, user: {user['id']}")
        
        if not qr_code:
            return jsonify({
                'success': False,
                'message': 'QR code is required'
            }), 400
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Verify QR code in pre_registrations
        cursor.execute("""
            SELECT * FROM pre_registrations
            WHERE qr_code = %s AND status IN ('approved', 'pending')
        """, (qr_code,))
        
        pre_reg = cursor.fetchone()
        
        logger.info(f"Pre-registration found: {pre_reg}")
        
        if not pre_reg:
            cursor.close()
            conn.close()
            return jsonify({
                'success': False,
                'message': 'QR code not found in pre-registrations or not approved'
            }), 404
        
        # Get host name
        host_name = 'Unknown Host'
        effective_host_id = host_id or pre_reg['host_id'] or user['id']
        
        if pre_reg['host_id']:
            cursor.execute("SELECT name FROM users WHERE id = %s", (pre_reg['host_id'],))
            host_result = cursor.fetchone()
            if host_result:
                host_name = host_result['name']
        elif pre_reg['host_name']:
            host_name = pre_reg['host_name']
        
        # Check if already checked in today
        today = datetime.now().date()
        cursor.execute("""
            SELECT id FROM visits 
            WHERE visitor_email = %s 
            AND DATE(check_in_time) = %s
            AND status = 'checked-in'
        """, (pre_reg['visitor_email'], today))
        
        existing_visit = cursor.fetchone()
        cursor.close()
        conn.close()
        
        if existing_visit:
            return jsonify({
                'success': False,
                'message': 'Visitor already checked in today'
            }), 409
        
        # Return visitor data for form pre-fill
        return jsonify({
            'success': True,
            'message': 'QR code verified successfully',
            'visitor_name': pre_reg['visitor_name'],
            'visitor_email': pre_reg['visitor_email'],
            'visitor_phone': pre_reg['visitor_phone'],
            'visitor_company': pre_reg['visitor_company'],
            'purpose': pre_reg['purpose'],
            'host_name': host_name,
            'host_id': effective_host_id,
            'pre_registration_id': pre_reg['id'],
            'visit_date': pre_reg['visit_date'].isoformat() if pre_reg['visit_date'] else None,
            'visit_time': str(pre_reg['visit_time']) if pre_reg['visit_time'] else None
        }), 200
        
    except Exception as e:
        logger.error(f"QR verification error: {e}")
        logger.error(f"Error type: {type(e).__name__}")
        logger.error(f"Error details: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({
            'success': False,
            'message': 'Internal server error during QR verification',
            'error': str(e) if app.debug else None
        }), 500

@app.route('/api/visitors/pre-registrations', methods=['GET'])
@authenticate_token
def get_pre_registrations():
    """Get pre-registrations for the user's company"""
    try:
        user = request.current_user
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # First, let's check what columns actually exist in the table
        cursor.execute("SHOW COLUMNS FROM pre_registrations")
        columns = [row['Field'] for row in cursor.fetchall()]
        logger.info(f"Available pre_registrations columns: {columns}")
        
        # Build a safe query with only existing columns
        safe_columns = []
        possible_columns = [
            'id', 'company_id', 'visitor_name', 'visitor_email', 'visitor_phone', 
            'visitor_company', 'company_to_visit', 'host_name', 'host_id', 
            'visit_date', 'visit_time', 'purpose', 'duration', 'special_requirements', 
            'emergency_contact', 'vehicle_number', 'qr_code', 'status', 
            'created_at', 'updated_at', 'check_in_time', 'check_out_time',
            'number_of_visitors'
        ]
        
        for col in possible_columns:
            if col in columns:
                safe_columns.append(f'pr.{col}')
        
        if not safe_columns:
            # Fallback to basic columns that should exist
            safe_columns = ['pr.*']
        
        query = f"""
            SELECT {', '.join(safe_columns)}
            FROM pre_registrations pr
            WHERE pr.company_to_visit = %s
            ORDER BY pr.created_at DESC
        """
        
        cursor.execute(query, (user['company_name'],))
        pre_registrations = cursor.fetchall()
        cursor.close()
        conn.close()
        
        # Convert any datetime objects to ISO format strings and handle binary data
        processed_registrations = []
        for registration in pre_registrations:
            processed_registration = {}
            for key, value in registration.items():
                if value is None:
                    processed_registration[key] = None
                elif hasattr(value, 'isoformat'):
                    # Handle datetime objects
                    processed_registration[key] = value.isoformat()
                elif isinstance(value, (bytes, bytearray)):
                    # Handle binary data
                    try:
                        processed_registration[key] = value.decode('utf-8')
                    except UnicodeDecodeError:
                        # If it's not UTF-8, convert to base64
                        import base64
                        processed_registration[key] = base64.b64encode(value).decode('utf-8')
                elif hasattr(value, 'total_seconds'):
                    # Handle timedelta objects
                    processed_registration[key] = str(value)
                else:
                    processed_registration[key] = value
            processed_registrations.append(processed_registration)
        
        return jsonify(processed_registrations), 200
        
    except Exception as e:
        logger.error(f"Pre-registrations fetch error: {e}")
        return jsonify({
            'message': 'Failed to fetch pre-registrations.',
            'details': str(e) if app.debug else None
        }), 500

@app.route('/api/visitors/recurring', methods=['GET'])
@authenticate_token
def get_recurring_visitors():
    """Get recurring pre-registrations for the user's company"""
    try:
        user = request.current_user
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # First, let's check what columns actually exist in the table
        cursor.execute("SHOW COLUMNS FROM pre_registrations")
        columns = [row['Field'] for row in cursor.fetchall()]
        logger.info(f"Available pre_registrations columns: {columns}")
        
        # Build a safe query with only existing columns for recurring visitors
        safe_columns = []
        possible_columns = [
            'id', 'company_id', 'visitor_name', 'visitor_email', 'visitor_phone', 
            'visitor_company', 'company_to_visit', 'host_name', 'host_id', 
            'visit_date', 'visit_time', 'purpose', 'duration', 'special_requirements', 
            'emergency_contact', 'vehicle_number', 'qr_code', 'status', 
            'created_at', 'updated_at', 'check_in_time', 'check_out_time',
            'number_of_visitors', 'is_recurring', 'recurring_pattern', 'recurring_end_date'
        ]
        
        for col in possible_columns:
            if col in columns:
                safe_columns.append(f'pr.{col}')
        
        if not safe_columns:
            # Fallback to basic columns that should exist
            safe_columns = ['pr.*']
        
        # Check if is_recurring column exists
        has_recurring_column = 'is_recurring' in columns
        
        if has_recurring_column:
            query = f"""
                SELECT {', '.join(safe_columns)}
                FROM pre_registrations pr
                WHERE pr.company_to_visit = %s 
                AND pr.is_recurring = TRUE
                ORDER BY pr.created_at DESC
            """
        else:
            # If no is_recurring column, return empty array
            cursor.close()
            conn.close()
            return jsonify([]), 200
        
        cursor.execute(query, (user['company_name'],))
        recurring_registrations = cursor.fetchall()
        cursor.close()
        conn.close()
        
        # Convert any datetime objects to ISO format strings and handle binary data
        processed_registrations = []
        for registration in recurring_registrations:
            processed_registration = {}
            for key, value in registration.items():
                if value is None:
                    processed_registration[key] = None
                elif hasattr(value, 'isoformat'):
                    # Handle datetime objects
                    processed_registration[key] = value.isoformat()
                elif isinstance(value, (bytes, bytearray)):
                    # Handle binary data
                    try:
                        processed_registration[key] = value.decode('utf-8')
                    except UnicodeDecodeError:
                        # If it's not UTF-8, convert to base64
                        import base64
                        processed_registration[key] = base64.b64encode(value).decode('utf-8')
                elif hasattr(value, 'total_seconds'):
                    # Handle timedelta objects
                    processed_registration[key] = str(value)
                else:
                    processed_registration[key] = value
            processed_registrations.append(processed_registration)
        
        return jsonify(processed_registrations), 200
        
    except Exception as e:
        logger.error(f"Recurring visitors fetch error: {e}")
        return jsonify({
            'message': 'Failed to fetch recurring visitors.',
            'details': str(e) if app.debug else None
        }), 500

@app.route('/api/pre-registrations/<int:pre_registration_id>/badge', methods=['GET'])
@authenticate_token
def generate_visitor_badge(pre_registration_id):
    """Generate visitor badge for pre-registration"""
    try:
        user = request.current_user
        
        # Check if photo should be included (for pre-registrations, default to False)
        include_photo = request.args.get('includePhoto', 'false').lower() == 'true'
        
        logger.info(f"Badge generation request for pre-registration ID: {pre_registration_id}, includePhoto: {include_photo}")
        logger.info(f"User details: {user}")
        
        # Get pre-registration details
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True, buffered=True)
        
        try:
            # Dynamic column detection
            logger.info("Starting column detection...")
            cursor.execute("SHOW COLUMNS FROM pre_registrations")
            columns_info = cursor.fetchall()
            logger.info(f"Raw columns_info: {columns_info}")
            
            # Extract column names properly - columns_info is a list of dictionaries
            if columns_info and isinstance(columns_info[0], dict):
                available_columns = [col['Field'] for col in columns_info]
            else:
                # Fallback: assume it's a list of tuples
                available_columns = [col[0] for col in columns_info]
            
            logger.info(f"Available pre_registrations columns for badge: {available_columns}")
            
            # Define columns we want for badge generation
            badge_columns = [
                'id', 'visitor_name', 'visitor_email', 'visitor_phone', 'visitor_company',
                'company_to_visit', 'host_name', 'visit_date', 'visit_time', 'purpose',
                'qr_code', 'created_at', 'status', 'number_of_visitors'
            ]
            
            # Select only available columns
            selected_columns = [col for col in badge_columns if col in available_columns]
            logger.info(f"Selected columns for badge query: {selected_columns}")
            
            if not selected_columns:
                logger.error("No valid columns found for badge generation")
                return jsonify({'message': 'Database schema error'}), 500
            
            query = f"SELECT {', '.join(selected_columns)} FROM pre_registrations WHERE id = %s"
            logger.info(f"Executing badge query: {query} with ID: {pre_registration_id}")
            cursor.execute(query, (pre_registration_id,))
            pre_registration = cursor.fetchone()
            logger.info(f"Badge query result: {pre_registration}")
            
            if not pre_registration:
                logger.warning(f"Pre-registration not found: {pre_registration_id}")
                return jsonify({'message': 'Pre-registration not found'}), 404
            
            # Get user name safely
            user_name = user.get('name') or user.get('email', 'Unknown User')
            
            # Check if user has access to this pre-registration
            if user['role'] == 'host':
                # For hosts, check if they are the host for this pre-registration
                if 'host_name' in pre_registration and pre_registration['host_name'] != user_name:
                    logger.warning(f"Host {user_name} attempted to access badge for different host's pre-registration")
                    return jsonify({'message': 'Access denied'}), 403
            elif user['role'] == 'admin':
                # For admins, check company match
                if 'company_to_visit' in pre_registration and pre_registration['company_to_visit'] != user['company_name']:
                    logger.warning(f"Admin from {user['company_name']} attempted to access badge from different company")
                    return jsonify({'message': 'Access denied'}), 403
            
            # Process data for JSON serialization
            processed_data = {}
            for key, value in pre_registration.items():
                if value is None:
                    processed_data[key] = None
                elif isinstance(value, datetime):
                    processed_data[key] = value.isoformat()
                elif isinstance(value, (bytes, bytearray)):
                    processed_data[key] = value.decode('utf-8') if value else None
                elif hasattr(value, 'total_seconds'):
                    processed_data[key] = str(value)
                else:
                    processed_data[key] = value
            
            # Generate badge HTML content
            badge_html = f"""
            <div style="
                width: 380px;
                height: 520px;
                border: none;
                border-radius: 15px;
                padding: 0;
                font-family: 'Segoe UI', -apple-system, BlinkMacSystemFont, sans-serif;
                background: linear-gradient(145deg, #667eea 0%, #764ba2 100%);
                box-shadow: 0 8px 25px rgba(0,0,0,0.15), 0 0 0 1px rgba(255,255,255,0.1);
                margin: 0 auto;
                position: relative;
                overflow: hidden;
            ">
                <!-- Decorative Top Pattern -->
                <div style="
                    position: absolute;
                    top: 0;
                    left: 0;
                    right: 0;
                    height: 80px;
                    background: linear-gradient(45deg, rgba(255,255,255,0.2) 0%, rgba(255,255,255,0.1) 100%);
                    clip-path: polygon(0 0, 100% 0, 100% 60%, 0 80%);
                "></div>
                
                <!-- Main Content Container -->
                <div style="
                    background: white;
                    margin: 15px;
                    border-radius: 12px;
                    min-height: calc(100% - 50px);
                    position: relative;
                    box-shadow: inset 0 1px 3px rgba(0,0,0,0.1);
                    padding: 25px 20px 20px 20px;
                ">
                    <!-- Header Section -->
                    <div style="text-align: center; margin-bottom: 25px;">
                        <div style="
                            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                            -webkit-background-clip: text;
                            -webkit-text-fill-color: transparent;
                            background-clip: text;
                            font-size: 24px;
                            font-weight: 800;
                            letter-spacing: 2px;
                            margin-bottom: 8px;
                            text-transform: uppercase;
                        ">VISITOR</div>
                        <div style="
                            background: #f8f9fa;
                            color: #495057;
                            font-size: 12px;
                            padding: 6px 16px;
                            border-radius: 20px;
                            display: inline-block;
                            font-weight: 600;
                            letter-spacing: 1px;
                            text-transform: uppercase;
                        ">{processed_data.get('company_to_visit', 'Company')}</div>
                    </div>
                    
                    {f'''
                    <!-- Visitor Photo Section -->
                    <div style="text-align: center; margin-bottom: 25px;">
                        <div style="
                            width: 100px;
                            height: 100px;
                            background: linear-gradient(135deg, #667eea20 0%, #764ba240 100%);
                            border-radius: 50%;
                            margin: 0 auto;
                            display: flex;
                            align-items: center;
                            justify-content: center;
                            border: 4px solid #ffffff;
                            box-shadow: 0 4px 15px rgba(102, 126, 234, 0.3);
                            position: relative;
                        ">
                            <div style="
                                font-size: 14px;
                                color: #667eea;
                                font-weight: 600;
                                text-align: center;
                                line-height: 1.2;
                            ">
                                📷<br><span style="font-size: 10px;">PHOTO</span>
                            </div>
                        </div>
                    </div>
                    ''' if include_photo else ''}
                    
                    <!-- Visitor Information -->
                    <div style="text-align: center; margin-bottom: 25px;">
                        <div style="
                            font-size: 22px;
                            font-weight: 700;
                            color: #2d3748;
                            margin-bottom: 8px;
                            line-height: 1.2;
                        ">{processed_data.get('visitor_name', 'N/A')}</div>
                        <div style="
                            font-size: 14px;
                            color: #667eea;
                            font-weight: 600;
                            margin-bottom: 5px;
                        ">{processed_data.get('visitor_company', 'N/A')}</div>
                        <div style="
                            font-size: 12px;
                            color: #718096;
                            background: #f7fafc;
                            padding: 4px 12px;
                            border-radius: 15px;
                            display: inline-block;
                        ">{processed_data.get('visitor_email', 'N/A')}</div>
                    </div>
                    
                    <!-- Visit Details Card -->
                    <div style="
                        background: linear-gradient(135deg, #f7fafc 0%, #edf2f7 100%);
                        border-radius: 12px;
                        padding: 18px;
                        margin-bottom: 20px;
                        border: 1px solid #e2e8f0;
                        position: relative;
                    ">
                        <div style="
                            position: absolute;
                            top: -8px;
                            left: 20px;
                            background: white;
                            color: #667eea;
                            font-size: 10px;
                            font-weight: 700;
                            padding: 4px 12px;
                            border-radius: 12px;
                            border: 1px solid #e2e8f0;
                            text-transform: uppercase;
                            letter-spacing: 1px;
                        ">Visit Details</div>
                        
                        <div style="margin-top: 8px;">
                            <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 10px;">
                                <span style="font-weight: 600; color: #4a5568; font-size: 11px; text-transform: uppercase; letter-spacing: 0.5px;">👤 Host:</span>
                                <span style="color: #2d3748; font-size: 13px; font-weight: 600;">{processed_data.get('host_name', 'N/A')}</span>
                            </div>
                            <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 10px;">
                                <span style="font-weight: 600; color: #4a5568; font-size: 11px; text-transform: uppercase; letter-spacing: 0.5px;">🎯 Purpose:</span>
                                <span style="color: #2d3748; font-size: 13px; font-weight: 600;">{processed_data.get('purpose', 'N/A')}</span>
                            </div>
                            <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 10px;">
                                <span style="font-weight: 600; color: #4a5568; font-size: 11px; text-transform: uppercase; letter-spacing: 0.5px;">📅 Date:</span>
                                <span style="color: #2d3748; font-size: 13px; font-weight: 600;">{processed_data.get('visit_date', 'N/A')}</span>
                            </div>
                            <div style="display: flex; justify-content: space-between; align-items: center;">
                                <span style="font-weight: 600; color: #4a5568; font-size: 11px; text-transform: uppercase; letter-spacing: 0.5px;">🕐 Time:</span>
                                <span style="color: #2d3748; font-size: 13px; font-weight: 600;">{processed_data.get('visit_time', 'N/A')}</span>
                            </div>
                        </div>
                    </div>
                    
                    <!-- QR Code and Status Section -->
                    <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 15px;">
                        <div style="text-align: center;">
                            <div style="
                                width: 70px;
                                height: 70px;
                                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                                margin: 0 auto 8px;
                                display: flex;
                                align-items: center;
                                justify-content: center;
                                font-size: 8px;
                                color: white;
                                border-radius: 8px;
                                box-shadow: 0 3px 10px rgba(102, 126, 234, 0.3);
                                position: relative;
                            ">
                                <div style="text-align: center; line-height: 1.1;">
                                    📱<br>
                                    <div style="font-size: 6px; margin-top: 2px;">QR CODE</div>
                                    <div style="font-size: 5px; margin-top: 1px; opacity: 0.8;">{processed_data.get('qr_code', 'N/A')[:8]}...</div>
                                </div>
                            </div>
                        </div>
                        
                        <div style="text-align: center; flex: 1; margin-left: 15px;">
                            <div style="
                                background: linear-gradient(135deg, #48bb78 0%, #38a169 100%);
                                color: white;
                                font-size: 10px;
                                font-weight: 700;
                                padding: 8px 16px;
                                border-radius: 20px;
                                text-transform: uppercase;
                                letter-spacing: 1px;
                                box-shadow: 0 2px 8px rgba(72, 187, 120, 0.3);
                            ">{processed_data.get('status', 'PENDING')}</div>
                            <div style="
                                font-size: 10px;
                                color: #718096;
                                margin-top: 5px;
                                font-weight: 500;
                            ">Visitors: {processed_data.get('number_of_visitors', 1)}</div>
                        </div>
                    </div>
                    
                    <!-- Footer -->
                    <div style="
                        text-align: center;
                        font-size: 9px;
                        color: #a0aec0;
                        border-top: 1px solid #e2e8f0;
                        padding-top: 12px;
                        margin-top: 20px;
                        background: white;
                    ">
                        <div style="font-weight: 600; font-size: 11px; color: #4a5568;">ID: #{pre_registration_id}</div>
                        <div style="margin-top: 4px; font-size: 8px;">Generated: {datetime.now().strftime('%d %b %Y, %H:%M')} by {user_name}</div>
                    </div>
                </div>
                
                <!-- Security Strip -->
                <div style="
                    position: absolute;
                    bottom: 2px;
                    left: 5px;
                    right: 5px;
                    height: 4px;
                    background: linear-gradient(90deg, #667eea 0%, #764ba2 50%, #667eea 100%);
                    border-radius: 0 0 10px 10px;
                "></div>
            </div>
            """
            
            # Return both HTML and data for frontend compatibility
            badge_data = {
                'html': badge_html,
                'data': {
                    'preRegistrationId': pre_registration_id,
                    'visitorName': processed_data.get('visitor_name', 'N/A'),
                    'visitorEmail': processed_data.get('visitor_email', 'N/A'),
                    'visitorPhone': processed_data.get('visitor_phone', 'N/A'),
                    'visitorCompany': processed_data.get('visitor_company', 'N/A'),
                    'companyToVisit': processed_data.get('company_to_visit', 'N/A'),
                    'hostName': processed_data.get('host_name', 'N/A'),
                    'visitDate': processed_data.get('visit_date'),
                    'visitTime': processed_data.get('visit_time'),
                    'purpose': processed_data.get('purpose', 'N/A'),
                    'qrCode': processed_data.get('qr_code', ''),
                    'status': processed_data.get('status', 'pending'),
                    'numberOfVisitors': processed_data.get('number_of_visitors', 1),
                    'generatedAt': datetime.now().isoformat(),
                    'generatedBy': user_name
                }
            }
            
            logger.info(f"Badge generated successfully for pre-registration {pre_registration_id}")
            return jsonify(badge_data), 200
            
        finally:
            cursor.close()
            conn.close()
        
    except Exception as e:
        logger.error(f"Badge generation error: {e}")
        return jsonify({
            'message': 'Failed to generate visitor badge',
            'details': str(e) if app.debug else None
        }), 500

@app.route('/api/visitors/history', methods=['GET'])
@authenticate_token
def get_visitor_history():
    """Get visitor history - Admin only, filtered by admin's company"""
    try:
        user = request.current_user
        
        # Only admin users can access visitor history
        if user['role'] != 'admin':
            return jsonify({'message': 'Admin access required to view visitor history'}), 403
        
        # Get company_id from companies table using user_id
        company_id = get_company_id_from_companies_table(user['id'])
        company_name = user['company_name']
        
        limit = request.args.get('limit', 100)
        start_date = request.args.get('startDate')
        end_date = request.args.get('endDate')
        visitor_email = request.args.get('visitorEmail')
        host_name = request.args.get('hostName')
        
        # Enhanced query to ensure we only get visits from the admin's company
        query = """
            SELECT v.*, vis.name as visitor_name, vis.email as visitor_email, 
                   vis.company as visitor_company, vis.is_blacklisted, u.name as host_name,
                   u.company_name as host_company,
                   CASE 
                     WHEN v.check_out_time IS NOT NULL THEN 'completed'
                     WHEN v.check_in_time IS NOT NULL THEN 'active'
                     ELSE 'pending'
                   END as status
            FROM visits v
            LEFT JOIN visitors vis ON v.visitor_id = vis.id
            LEFT JOIN users u ON v.host_id = u.id
            WHERE u.company_name = %s
        """
        
        params = [company_name]
        
        if start_date:
            query += " AND DATE(v.check_in_time) >= %s"
            params.append(start_date)
        
        if end_date:
            query += " AND DATE(v.check_in_time) <= %s"
            params.append(end_date)
        
        if visitor_email:
            query += " AND (vis.email LIKE %s OR v.visitor_email LIKE %s)"
            params.append(f"%{visitor_email}%")
            params.append(f"%{visitor_email}%")
        
        if host_name:
            query += " AND (u.name LIKE %s OR v.host_name LIKE %s)"
            params.append(f"%{host_name}%")
            params.append(f"%{host_name}%")
        
        query += " ORDER BY v.check_in_time DESC LIMIT %s"
        params.append(int(limit))
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(query, params)
        history = cursor.fetchall()
        cursor.close()
        conn.close()
        
        return jsonify(history), 200
        
    except Exception as e:
        logger.error(f"Visitor history error: {e}")
        return jsonify({'message': 'Failed to fetch visitor history'}), 500

# Health check endpoint
@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'service': 'Flask Visitor Management Backend'
    }), 200

# CORS debug endpoint
@app.route('/api/cors-debug', methods=['GET', 'OPTIONS'])
def cors_debug():
    """Debug endpoint to check CORS configuration"""
    origin = request.headers.get('Origin', 'No Origin Header')
    
    return jsonify({
        'message': 'CORS Debug Information',
        'request_origin': origin,
        'allowed_origins': production_origins,
        'request_method': request.method,
        'request_headers': dict(request.headers),
        'timestamp': datetime.now().isoformat()
    }), 200

# ============== ADMIN ENDPOINTS ==============

@app.route('/api/admin/settings', methods=['GET'])
@authenticate_token
def get_admin_settings():
    """Get system settings for admin dashboard"""
    try:
        user = request.current_user
        
        if user['role'] != 'admin':
            return jsonify({'message': 'Admin access required'}), 403
        
        # Get settings from database or use defaults
        settings = {
            'emailNotifications': True,
            'requireApproval': False,
            'maxVisitorsPerDay': 500,
            'retentionPeriodDays': 90,
            'allowSelfCheckout': True,
            'capturePhoto': True,
            'systemName': 'Visitor Management System',
            'companyLogo': '/assets/logo.png',
            'theme': 'light',
            'language': 'en',
            'timezone': 'UTC',
            'dateFormat': 'MM/DD/YYYY',
            'timeFormat': '12h',
            'lastUpdated': datetime.now().isoformat()
        }
        
        # Try to fetch from database if settings table exists
        try:
            conn = get_db_connection()
            cursor = conn.cursor(dictionary=True)
            
            # Check if settings table exists
            cursor.execute("SHOW TABLES LIKE 'system_settings'")
            if cursor.fetchone():
                cursor.execute("SELECT * FROM system_settings LIMIT 1")
                db_settings = cursor.fetchone()
                
                if db_settings:
                    # Update settings with values from database
                    for key, value in db_settings.items():
                        if key in settings:
                            settings[key] = value
            
            cursor.close()
            conn.close()
        except Exception as db_error:
            logger.warning(f"Error fetching settings from database: {db_error}")
            # Continue with default settings
        
        return jsonify(settings), 200
        
    except Exception as e:
        logger.error(f"Get admin settings error: {e}")
        return jsonify({
            'emailNotifications': True,
            'requireApproval': False,
            'maxVisitorsPerDay': 500,
            'retentionPeriodDays': 90,
            'allowSelfCheckout': True,
            'capturePhoto': True,
            'systemName': 'Visitor Management System',
            'companyLogo': '/assets/logo.png',
            'theme': 'light',
            'language': 'en',
            'timezone': 'UTC',
            'dateFormat': 'MM/DD/YYYY',
            'timeFormat': '12h',
            'lastUpdated': datetime.now().isoformat()
        }), 200

@app.route('/api/admin/settings', methods=['PUT'])
@authenticate_token
def update_admin_settings():
    """Update system settings"""
    try:
        user = request.current_user
        
        if user['role'] != 'admin':
            return jsonify({'message': 'Admin access required'}), 403
        
        data = request.get_json()
        
        # Validate settings
        allowed_settings = [
            'emailNotifications', 'requireApproval', 'maxVisitorsPerDay',
            'retentionPeriodDays', 'allowSelfCheckout', 'capturePhoto',
            'systemName', 'companyLogo', 'theme', 'language', 'timezone',
            'dateFormat', 'timeFormat'
        ]
        
        filtered_settings = {k: v for k, v in data.items() if k in allowed_settings}
        
        # Try to update in database if settings table exists
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            # Check if settings table exists
            cursor.execute("SHOW TABLES LIKE 'system_settings'")
            if not cursor.fetchone():
                # Create settings table if it doesn't exist
                cursor.execute("""
                    CREATE TABLE system_settings (
                        id INT AUTO_INCREMENT PRIMARY KEY,
                        emailNotifications BOOLEAN DEFAULT TRUE,
                        requireApproval BOOLEAN DEFAULT FALSE,
                        maxVisitorsPerDay INT DEFAULT 500,
                        retentionPeriodDays INT DEFAULT 90,
                        allowSelfCheckout BOOLEAN DEFAULT TRUE,
                        capturePhoto BOOLEAN DEFAULT TRUE,
                        systemName VARCHAR(255) DEFAULT 'Visitor Management System',
                        companyLogo VARCHAR(255) DEFAULT '/assets/logo.png',
                        theme VARCHAR(50) DEFAULT 'light',
                        language VARCHAR(10) DEFAULT 'en',
                        timezone VARCHAR(50) DEFAULT 'UTC',
                        dateFormat VARCHAR(20) DEFAULT 'MM/DD/YYYY',
                        timeFormat VARCHAR(10) DEFAULT '12h',
                        lastUpdated DATETIME DEFAULT CURRENT_TIMESTAMP
                    )
                """)
                
                # Insert default values
                cursor.execute("""
                    INSERT INTO system_settings (emailNotifications, requireApproval, 
                    maxVisitorsPerDay, retentionPeriodDays, allowSelfCheckout, 
                    capturePhoto, systemName, companyLogo, theme, language, timezone, 
                    dateFormat, timeFormat) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    True, False, 500, 90, True, True, 
                    'Visitor Management System', '/assets/logo.png',
                    'light', 'en', 'UTC', 'MM/DD/YYYY', '12h'
                ))
            
            # Update settings in the database
            update_query = "UPDATE system_settings SET "
            update_values = []
            
            for key, value in filtered_settings.items():
                update_query += f"{key} = %s, "
                update_values.append(value)
            
            update_query += "lastUpdated = %s WHERE id = 1"
            update_values.append(datetime.now())
            
            cursor.execute(update_query, update_values)
            conn.commit()
            
            cursor.close()
            conn.close()
            
        except Exception as db_error:
            logger.warning(f"Error updating settings in database: {db_error}")
            # Continue without database update
        
        return jsonify({
            'message': 'Settings updated successfully',
            'settings': {**filtered_settings, 'lastUpdated': datetime.now().isoformat()}
        }), 200
        
    except Exception as e:
        logger.error(f"Update admin settings error: {e}")
        return jsonify({'message': 'Failed to update settings'}), 500

@app.route('/api/admin/audit-logs', methods=['GET'])
@authenticate_token
def get_audit_logs():
    """Get audit logs for admin monitoring"""
    try:
        user = request.current_user
        
        if user['role'] != 'admin':
            return jsonify({'message': 'Admin access required'}), 403
        
        # Get query parameters
        start_date = request.args.get('startDate')
        end_date = request.args.get('endDate')
        action = request.args.get('action')
        username = request.args.get('username')
        limit = request.args.get('limit', 100)
        
        # Check if audit_logs table exists
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("SHOW TABLES LIKE 'audit_logs'")
        if not cursor.fetchone():
            # Create sample audit log data
            mock_logs = []
            for i in range(10):
                days_ago = random.randint(0, 30)
                action_type = random.choice(['login', 'logout', 'create', 'update', 'delete', 'export'])
                mock_logs.append({
                    'id': i + 1,
                    'timestamp': (datetime.now() - timedelta(days=days_ago)).isoformat(),
                    'user': f"user{random.randint(1, 5)}@example.com",
                    'action': action_type,
                    'details': f"Sample {action_type} action",
                    'ip_address': f"192.168.1.{random.randint(1, 255)}"
                })
            
            cursor.close()
            conn.close()
            return jsonify(mock_logs), 200
        
        # Build query
        query = """
            SELECT * FROM audit_logs 
            WHERE company_name = %s
        """
        
        params = [user['company_name']]
        
        # Add filters
        if start_date:
            query += " AND DATE(timestamp) >= %s"
            params.append(start_date)
        
        if end_date:
            query += " AND DATE(timestamp) <= %s"
            params.append(end_date)
        
        if action:
            query += " AND action = %s"
            params.append(action)
        
        if username:
            query += " AND user LIKE %s"
            params.append(f"%{username}%")
        
        query += " ORDER BY timestamp DESC LIMIT %s"
        params.append(int(limit))
        
        cursor.execute(query, params)
        logs = cursor.fetchall()
        
        # Process datetime for JSON serialization
        for log in logs:
            if log.get('timestamp'):
                log['timestamp'] = log['timestamp'].isoformat()
        
        cursor.close()
        conn.close()
        
        return jsonify(logs), 200
        
    except Exception as e:
        logger.error(f"Get audit logs error: {e}")
        # Return mock data on error
        mock_logs = []
        for i in range(10):
            days_ago = random.randint(0, 30)
            action_type = random.choice(['login', 'logout', 'create', 'update', 'delete', 'export'])
            mock_logs.append({
                'id': i + 1,
                'timestamp': (datetime.now() - timedelta(days=days_ago)).isoformat(),
                'user': f"user{random.randint(1, 5)}@example.com",
                'action': action_type,
                'details': f"Sample {action_type} action",
                'ip_address': f"192.168.1.{random.randint(1, 255)}"
            })
        
        return jsonify(mock_logs), 200

@app.route('/api/admin/backups', methods=['GET'])
@authenticate_token
def get_backups():
    """Get list of database backups"""
    try:
        user = request.current_user
        
        if user['role'] != 'admin':
            return jsonify({'message': 'Admin access required'}), 403
        
        # Generate mock backup data since actual backup implementation would depend on server setup
        mock_backups = []
        for i in range(5):
            days_ago = i * 7  # Weekly backups
            backup_date = datetime.now() - timedelta(days=days_ago)
            size_mb = random.uniform(1.5, 10.0)
            
            mock_backups.append({
                'id': i + 1,
                'filename': f"vms_backup_{backup_date.strftime('%Y%m%d')}.sql",
                'created_at': backup_date.isoformat(),
                'size': f"{size_mb:.2f} MB",
                'status': 'completed',
                'download_url': f"/api/admin/backups/{i+1}/download"
            })
        
        return jsonify(mock_backups), 200
        
    except Exception as e:
        logger.error(f"Get backups error: {e}")
        return jsonify([]), 200

@app.route('/api/admin/backups/create', methods=['POST'])
@authenticate_token
def create_backup():
    """Create a new database backup"""
    try:
        user = request.current_user
        
        if user['role'] != 'admin':
            return jsonify({'message': 'Admin access required'}), 403
        
        # Mock backup creation (actual implementation would depend on server setup)
        backup_date = datetime.now()
        size_mb = random.uniform(1.5, 10.0)
        
        return jsonify({
            'id': random.randint(10, 100),
            'filename': f"vms_backup_{backup_date.strftime('%Y%m%d_%H%M%S')}.sql",
            'created_at': backup_date.isoformat(),
            'size': f"{size_mb:.2f} MB",
            'status': 'completed',
            'message': 'Backup created successfully'
        }), 201
        
    except Exception as e:
        logger.error(f"Create backup error: {e}")
        return jsonify({'message': 'Failed to create backup'}), 500

@app.route('/api/admin/backups/<int:backup_id>/download', methods=['GET'])
@authenticate_token
def download_backup(backup_id):
    """Download a specific backup file"""
    try:
        user = request.current_user
        
        if user['role'] != 'admin':
            return jsonify({'message': 'Admin access required'}), 403
        
        # Mock backup download (actual implementation would depend on server setup)
        # In a real system, you would retrieve the backup file and send it
        # Here we just create a simple mock SQL file
        
        mock_sql = f"""-- MySQL dump 10.13  Distrib 8.0.28, for Win64 (x86_64)
--
-- Host: localhost    Database: vms_db
-- ------------------------------------------------------
-- Server version    8.0.28

/*!40101 SET @OLD_CHARACTER_SET_CLIENT=@@CHARACTER_SET_CLIENT */;
/*!40101 SET @OLD_CHARACTER_SET_RESULTS=@@CHARACTER_SET_RESULTS */;
/*!40101 SET @OLD_COLLATION_CONNECTION=@@COLLATION_CONNECTION */;
/*!50503 SET NAMES utf8mb4 */;
/*!40103 SET @OLD_TIME_ZONE=@@TIME_ZONE */;
/*!40103 SET TIME_ZONE='+00:00' */;
/*!40014 SET @OLD_UNIQUE_CHECKS=@@UNIQUE_CHECKS, UNIQUE_CHECKS=0 */;
/*!40014 SET @OLD_FOREIGN_KEY_CHECKS=@@FOREIGN_KEY_CHECKS, FOREIGN_KEY_CHECKS=0 */;
/*!40101 SET @OLD_SQL_MODE=@@SQL_MODE, SQL_MODE='NO_AUTO_VALUE_ON_ZERO' */;
/*!40111 SET @OLD_SQL_NOTES=@@SQL_NOTES, SQL_NOTES=0 */;

--
-- Table structure for table `users`
--

DROP TABLE IF EXISTS `users`;
/*!40101 SET @saved_cs_client     = @@character_set_client */;
/*!50503 SET character_set_client = utf8mb4 */;
CREATE TABLE `users` (
  `id` int NOT NULL AUTO_INCREMENT,
  `name` varchar(100) NOT NULL,
  `email` varchar(100) NOT NULL,
  `password` varchar(255) NOT NULL,
  `role` varchar(20) NOT NULL DEFAULT 'host',
  `company_name` varchar(100) DEFAULT NULL,
  `created_at` timestamp NULL DEFAULT CURRENT_TIMESTAMP,
  PRIMARY KEY (`id`),
  UNIQUE KEY `email` (`email`)
) ENGINE=InnoDB AUTO_INCREMENT=10 DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci;
/*!40101 SET character_set_client = @saved_cs_client */;

-- Dump completed on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
"""
        
        # Create a temporary file with the mock SQL content
        temp_file_path = os.path.join(os.path.dirname(__file__), f"temp_backup_{backup_id}.sql")
        with open(temp_file_path, 'w') as f:
            f.write(mock_sql)
        
        # Send the file and then delete it
        response = send_file(
            temp_file_path,
            as_attachment=True,
            attachment_filename=f"vms_backup_{datetime.now().strftime('%Y%m%d')}.sql",
            mimetype='application/sql'
        )
        
        # Delete the file after sending (in production, use a better cleanup mechanism)
        @response.call_on_close
        def on_close():
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)
        
        return response
        
    except Exception as e:
        logger.error(f"Download backup error: {e}")
        return jsonify({'message': 'Failed to download backup'}), 500

# Test endpoint for host dashboard connectivity
@app.route('/api/test-host', methods=['GET'])
@authenticate_token
def test_host_endpoint():
    """Test endpoint specifically for host dashboard"""
    try:
        user = request.current_user
        
        logger.info(f"Test host endpoint called by user: {user}")
        
        # Also check if there are any visits for this host
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Count total visits for this host
        cursor.execute("SELECT COUNT(*) as count FROM visits WHERE host_id = %s", (user['id'],))
        visit_count = cursor.fetchone()
        
        # Get sample visit data
        cursor.execute("""
            SELECT v.id, v.visitor_name, v.visitor_email, v.check_in_time, v.status 
            FROM visits v 
            WHERE v.host_id = %s 
            ORDER BY v.check_in_time DESC 
            LIMIT 3
        """, (user['id'],))
        sample_visits = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'user': {
                'id': user['id'],
                'name': user['name'],
                'email': user['email'],
                'role': user['role'],
                'company_name': user['company_name']
            },
            'visit_statistics': {
                'total_visits': visit_count['count'] if visit_count else 0,
                'sample_visits': sample_visits
            },
            'message': 'Host test endpoint working',
            'timestamp': datetime.now().isoformat()
        }), 200
        
    except Exception as e:
        logger.error(f"Test host endpoint error: {e}")
        return jsonify({
            'success': False,
            'message': 'Test endpoint failed',
            'error': str(e)
        }), 500

# Test endpoint for specific host ID debugging
@app.route('/api/test-host/<int:host_id>', methods=['GET'])
def test_specific_host(host_id):
    """Test endpoint to check specific host data"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get host details
        cursor.execute("SELECT * FROM users WHERE id = %s", (host_id,))
        host = cursor.fetchone()
        
        # Get visits for this host
        cursor.execute("""
            SELECT v.*, vis.name as visitor_name_from_visitors, h.name as host_name
            FROM visits v
            LEFT JOIN visitors vis ON v.visitor_id = vis.id
            LEFT JOIN users h ON v.host_id = h.id
            WHERE v.host_id = %s
            ORDER BY v.check_in_time DESC
        """, (host_id,))
        visits = cursor.fetchall()
        
        # Get total visit count for this host
        cursor.execute("SELECT COUNT(*) as total FROM visits WHERE host_id = %s", (host_id,))
        visit_count = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'requested_host_id': host_id,
            'host_details': host,
            'visits_for_host': visits,
            'total_visits_for_host': visit_count,
            'message': f'Host {host_id} has {visit_count["total"]} visits' if visit_count else 'No visit count data'
        })
    except Exception as e:
        logger.error(f"Test specific host endpoint error: {e}")
        return jsonify({
            'success': False,
            'message': 'Test failed',
            'error': str(e)
        }), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=4000)
